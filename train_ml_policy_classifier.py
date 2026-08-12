from pathlib import Path
from datetime import datetime
import warnings

import numpy as np
import pandas as pd

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    precision_recall_fscore_support,
)
from sklearn.model_selection import LeaveOneOut
from sklearn.preprocessing import MultiLabelBinarizer

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

STATEMENT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database_pilot_expanded.xlsx"
)

POLICY_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "policy_classification_results_expanded.xlsx"
)

OUTPUT_EXCEL = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "ml_policy_classifier_results.xlsx"
)

OUTPUT_CSV = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "ml_policy_predictions.csv"
)


# ============================================================
# CONFIGURATION
# ============================================================

RANDOM_STATE = 42

# A policy topic must have at least this many positive and
# negative statements to be included in the ML pilot.
MIN_POSITIVE_STATEMENTS = 4
MIN_NEGATIVE_STATEMENTS = 4

TFIDF_MAX_FEATURES = 500
TFIDF_NGRAM_RANGE = (1, 2)

PREDICTION_THRESHOLD = 0.50

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop with a clear error if a required file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to clean text.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def safe_divide(numerator, denominator):
    """
    Divide safely.
    """
    if denominator == 0:
        return 0.0

    return numerator / denominator


def style_worksheet(worksheet):
    """
    Apply consistent Excel formatting.
    """
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 2, 12),
            60,
        )


# ============================================================
# VERIFY INPUT FILES
# ============================================================

require_file(
    STATEMENT_FILE,
    "Expanded candidate statement workbook",
)

require_file(
    POLICY_FILE,
    "Expanded policy-classification workbook",
)


# ============================================================
# LOAD STATEMENTS
# ============================================================

statements = pd.read_excel(
    STATEMENT_FILE,
    sheet_name="CandidateStatements",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

required_statement_columns = [
    "Statement_ID",
    "Candidate_ID",
    "Statement_Type",
    "Statement_Text",
]

missing_statement_columns = [
    column
    for column in required_statement_columns
    if column not in statements.columns
]

if missing_statement_columns:
    raise ValueError(
        "CandidateStatements is missing required columns: "
        + ", ".join(missing_statement_columns)
    )

for column in required_statement_columns:
    statements[column] = statements[
        column
    ].apply(clean_text)

statements = statements[
    (statements["Statement_ID"] != "")
    & (statements["Statement_Text"] != "")
].copy()

statements.drop_duplicates(
    subset=["Statement_ID"],
    keep="first",
    inplace=True,
)

statements.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# LOAD POLICY LABELS
# ============================================================

classifications = pd.read_excel(
    POLICY_FILE,
    sheet_name="PolicyClassifications",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

required_policy_columns = [
    "Statement_ID",
    "Policy_Topic",
]

missing_policy_columns = [
    column
    for column in required_policy_columns
    if column not in classifications.columns
]

if missing_policy_columns:
    raise ValueError(
        "PolicyClassifications is missing required columns: "
        + ", ".join(missing_policy_columns)
    )

classifications["Statement_ID"] = classifications[
    "Statement_ID"
].apply(clean_text)

classifications["Policy_Topic"] = classifications[
    "Policy_Topic"
].apply(clean_text)

classifications = classifications[
    (classifications["Statement_ID"] != "")
    & (classifications["Policy_Topic"] != "")
].copy()


# ============================================================
# CREATE MULTI-LABEL TARGETS
# ============================================================

statement_topic_map = (
    classifications
    .groupby("Statement_ID")["Policy_Topic"]
    .apply(
        lambda values: sorted(
            set(values)
        )
    )
    .to_dict()
)

statements["Policy_Topics"] = statements[
    "Statement_ID"
].map(
    statement_topic_map
)

statements["Policy_Topics"] = statements[
    "Policy_Topics"
].apply(
    lambda value: (
        value
        if isinstance(value, list)
        else []
    )
)

all_topics = sorted(
    classifications[
        "Policy_Topic"
    ].unique()
)

topic_support_rows = []

for topic in all_topics:
    positive_count = int(
        statements[
            "Policy_Topics"
        ].apply(
            lambda topics: topic in topics
        ).sum()
    )

    negative_count = (
        len(statements)
        - positive_count
    )

    eligible = (
        positive_count
        >= MIN_POSITIVE_STATEMENTS
        and negative_count
        >= MIN_NEGATIVE_STATEMENTS
    )

    topic_support_rows.append(
        {
            "Policy_Topic": topic,
            "Positive_Statements": positive_count,
            "Negative_Statements": negative_count,
            "Eligible_for_ML": (
                "Yes" if eligible else "No"
            ),
        }
    )

topic_support = pd.DataFrame(
    topic_support_rows
)

eligible_topics = topic_support[
    topic_support[
        "Eligible_for_ML"
    ]
    == "Yes"
]["Policy_Topic"].tolist()

if not eligible_topics:
    raise ValueError(
        "No policy topics have enough positive and negative "
        "examples for the ML pilot. Lower the minimum support "
        "thresholds or add more statements."
    )


# ============================================================
# CREATE BINARY LABEL MATRIX
# ============================================================

filtered_label_lists = statements[
    "Policy_Topics"
].apply(
    lambda topics: [
        topic
        for topic in topics
        if topic in eligible_topics
    ]
)

label_binarizer = MultiLabelBinarizer(
    classes=eligible_topics
)

target_matrix = label_binarizer.fit_transform(
    filtered_label_lists
)

target_dataframe = pd.DataFrame(
    target_matrix,
    columns=eligible_topics,
)


# ============================================================
# LEAVE-ONE-OUT CROSS-VALIDATION
# ============================================================

leave_one_out = LeaveOneOut()

prediction_matrix = np.zeros_like(
    target_matrix,
    dtype=int,
)

probability_matrix = np.zeros_like(
    target_matrix,
    dtype=float,
)

fold_rows = []

for fold_number, (
    train_indices,
    test_indices,
) in enumerate(
    leave_one_out.split(statements),
    start=1,
):
    training_text = statements.iloc[
        train_indices
    ]["Statement_Text"]

    test_text = statements.iloc[
        test_indices
    ]["Statement_Text"]

    vectorizer = TfidfVectorizer(
        lowercase=True,
        stop_words="english",
        ngram_range=TFIDF_NGRAM_RANGE,
        max_features=TFIDF_MAX_FEATURES,
        min_df=1,
        sublinear_tf=True,
    )

    training_features = vectorizer.fit_transform(
        training_text
    )

    test_features = vectorizer.transform(
        test_text
    )

    for topic_index, topic in enumerate(
        eligible_topics
    ):
        training_labels = target_matrix[
            train_indices,
            topic_index,
        ]

        # In a very small fold, a topic may contain only one
        # class in training data. Use the observed constant
        # rather than failing.
        unique_training_labels = np.unique(
            training_labels
        )

        if len(unique_training_labels) < 2:
            predicted_probability = float(
                unique_training_labels[0]
            )

            predicted_label = int(
                predicted_probability
                >= PREDICTION_THRESHOLD
            )

        else:
            model = LogisticRegression(
                class_weight="balanced",
                max_iter=1000,
                random_state=RANDOM_STATE,
                solver="liblinear",
            )

            model.fit(
                training_features,
                training_labels,
            )

            predicted_probability = float(
                model.predict_proba(
                    test_features
                )[0, 1]
            )

            predicted_label = int(
                predicted_probability
                >= PREDICTION_THRESHOLD
            )

        test_index = int(
            test_indices[0]
        )

        probability_matrix[
            test_index,
            topic_index,
        ] = predicted_probability

        prediction_matrix[
            test_index,
            topic_index,
        ] = predicted_label

        fold_rows.append(
            {
                "Fold": fold_number,
                "Statement_ID": statements.iloc[
                    test_index
                ]["Statement_ID"],
                "Policy_Topic": topic,
                "Actual_Label": int(
                    target_matrix[
                        test_index,
                        topic_index,
                    ]
                ),
                "Predicted_Label": (
                    predicted_label
                ),
                "Predicted_Probability": round(
                    predicted_probability,
                    4,
                ),
            }
        )

fold_predictions = pd.DataFrame(
    fold_rows
)


# ============================================================
# CALCULATE TOPIC-LEVEL METRICS
# ============================================================

topic_metric_rows = []

for topic_index, topic in enumerate(
    eligible_topics
):
    actual_values = target_matrix[
        :,
        topic_index,
    ]

    predicted_values = prediction_matrix[
        :,
        topic_index,
    ]

    precision, recall, f1, _ = (
        precision_recall_fscore_support(
            actual_values,
            predicted_values,
            average="binary",
            zero_division=0,
        )
    )

    accuracy = accuracy_score(
        actual_values,
        predicted_values,
    )

    true_positive = int(
        np.logical_and(
            actual_values == 1,
            predicted_values == 1,
        ).sum()
    )

    false_positive = int(
        np.logical_and(
            actual_values == 0,
            predicted_values == 1,
        ).sum()
    )

    false_negative = int(
        np.logical_and(
            actual_values == 1,
            predicted_values == 0,
        ).sum()
    )

    true_negative = int(
        np.logical_and(
            actual_values == 0,
            predicted_values == 0,
        ).sum()
    )

    topic_metric_rows.append(
        {
            "Policy_Topic": topic,
            "Positive_Statements": int(
                actual_values.sum()
            ),
            "Negative_Statements": int(
                len(actual_values)
                - actual_values.sum()
            ),
            "True_Positive": true_positive,
            "False_Positive": false_positive,
            "False_Negative": false_negative,
            "True_Negative": true_negative,
            "Accuracy_Percent": round(
                accuracy * 100,
                2,
            ),
            "Precision_Percent": round(
                precision * 100,
                2,
            ),
            "Recall_Percent": round(
                recall * 100,
                2,
            ),
            "F1_Percent": round(
                f1 * 100,
                2,
            ),
        }
    )

topic_metrics = pd.DataFrame(
    topic_metric_rows
)

topic_metrics.sort_values(
    by=[
        "F1_Percent",
        "Positive_Statements",
    ],
    ascending=[
        False,
        False,
    ],
    inplace=True,
)

topic_metrics.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# OVERALL MULTI-LABEL METRICS
# ============================================================

micro_precision, micro_recall, micro_f1, _ = (
    precision_recall_fscore_support(
        target_matrix.ravel(),
        prediction_matrix.ravel(),
        average="binary",
        zero_division=0,
    )
)

macro_precision, macro_recall, macro_f1, _ = (
    precision_recall_fscore_support(
        target_matrix,
        prediction_matrix,
        average="macro",
        zero_division=0,
    )
)

subset_accuracy = float(
    np.mean(
        np.all(
            target_matrix
            == prediction_matrix,
            axis=1,
        )
    )
)

hamming_accuracy = float(
    np.mean(
        target_matrix
        == prediction_matrix
    )
)

overall_metrics = pd.DataFrame(
    [
        {
            "Metric": "Statements",
            "Value": len(statements),
        },
        {
            "Metric": "Eligible Policy Topics",
            "Value": len(eligible_topics),
        },
        {
            "Metric": "Cross-Validation Method",
            "Value": "Leave-One-Out",
        },
        {
            "Metric": "Text Features",
            "Value": "TF-IDF unigrams and bigrams",
        },
        {
            "Metric": "Classifier",
            "Value": (
                "One-vs-Rest Binary Logistic Regression"
            ),
        },
        {
            "Metric": "Micro Precision Percent",
            "Value": round(
                micro_precision * 100,
                2,
            ),
        },
        {
            "Metric": "Micro Recall Percent",
            "Value": round(
                micro_recall * 100,
                2,
            ),
        },
        {
            "Metric": "Micro F1 Percent",
            "Value": round(
                micro_f1 * 100,
                2,
            ),
        },
        {
            "Metric": "Macro Precision Percent",
            "Value": round(
                macro_precision * 100,
                2,
            ),
        },
        {
            "Metric": "Macro Recall Percent",
            "Value": round(
                macro_recall * 100,
                2,
            ),
        },
        {
            "Metric": "Macro F1 Percent",
            "Value": round(
                macro_f1 * 100,
                2,
            ),
        },
        {
            "Metric": "Subset Accuracy Percent",
            "Value": round(
                subset_accuracy * 100,
                2,
            ),
        },
        {
            "Metric": "Hamming Accuracy Percent",
            "Value": round(
                hamming_accuracy * 100,
                2,
            ),
        },
        {
            "Metric": "Run Date",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
    ]
)


# ============================================================
# CREATE STATEMENT-LEVEL PREDICTION OUTPUT
# ============================================================

prediction_rows = []

for statement_index, statement_row in (
    statements.iterrows()
):
    actual_topics = [
        eligible_topics[topic_index]
        for topic_index in range(
            len(eligible_topics)
        )
        if target_matrix[
            statement_index,
            topic_index,
        ] == 1
    ]

    predicted_topics = [
        eligible_topics[topic_index]
        for topic_index in range(
            len(eligible_topics)
        )
        if prediction_matrix[
            statement_index,
            topic_index,
        ] == 1
    ]

    missed_topics = sorted(
        set(actual_topics)
        - set(predicted_topics)
    )

    extra_topics = sorted(
        set(predicted_topics)
        - set(actual_topics)
    )

    prediction_rows.append(
        {
            "Statement_ID": statement_row[
                "Statement_ID"
            ],
            "Candidate_ID": statement_row[
                "Candidate_ID"
            ],
            "Statement_Type": statement_row[
                "Statement_Type"
            ],
            "Statement_Text": statement_row[
                "Statement_Text"
            ],
            "Actual_ML_Topics": "; ".join(
                actual_topics
            ),
            "Predicted_ML_Topics": "; ".join(
                predicted_topics
            ),
            "Missed_Topics": "; ".join(
                missed_topics
            ),
            "Extra_Predicted_Topics": "; ".join(
                extra_topics
            ),
            "Exact_Label_Match": (
                "Yes"
                if set(actual_topics)
                == set(predicted_topics)
                else "No"
            ),
        }
    )

statement_predictions = pd.DataFrame(
    prediction_rows
)


# ============================================================
# TRAIN FINAL MODELS ON ALL STATEMENTS
# ============================================================

final_vectorizer = TfidfVectorizer(
    lowercase=True,
    stop_words="english",
    ngram_range=TFIDF_NGRAM_RANGE,
    max_features=TFIDF_MAX_FEATURES,
    min_df=1,
    sublinear_tf=True,
)

final_features = final_vectorizer.fit_transform(
    statements["Statement_Text"]
)

feature_names = np.asarray(
    final_vectorizer.get_feature_names_out()
)

coefficient_rows = []

for topic_index, topic in enumerate(
    eligible_topics
):
    labels = target_matrix[
        :,
        topic_index,
    ]

    if len(np.unique(labels)) < 2:
        continue

    final_model = LogisticRegression(
        class_weight="balanced",
        max_iter=1000,
        random_state=RANDOM_STATE,
        solver="liblinear",
    )

    final_model.fit(
        final_features,
        labels,
    )

    coefficients = final_model.coef_[0]

    top_positive_indices = np.argsort(
        coefficients
    )[-10:][::-1]

    top_negative_indices = np.argsort(
        coefficients
    )[:10]

    for rank, feature_index in enumerate(
        top_positive_indices,
        start=1,
    ):
        coefficient_rows.append(
            {
                "Policy_Topic": topic,
                "Direction": (
                    "Positive predictor"
                ),
                "Rank": rank,
                "Term": feature_names[
                    feature_index
                ],
                "Coefficient": round(
                    float(
                        coefficients[
                            feature_index
                        ]
                    ),
                    4,
                ),
            }
        )

    for rank, feature_index in enumerate(
        top_negative_indices,
        start=1,
    ):
        coefficient_rows.append(
            {
                "Policy_Topic": topic,
                "Direction": (
                    "Negative predictor"
                ),
                "Rank": rank,
                "Term": feature_names[
                    feature_index
                ],
                "Coefficient": round(
                    float(
                        coefficients[
                            feature_index
                        ]
                    ),
                    4,
                ),
            }
        )

model_coefficients = pd.DataFrame(
    coefficient_rows
)


# ============================================================
# CREATE LIMITATIONS SHEET
# ============================================================

limitations = pd.DataFrame(
    [
        {
            "Limitation": (
                "The dataset contains only 20 statements."
            ),
            "Implication": (
                "Performance estimates have high uncertainty "
                "and should be treated as proof-of-concept results."
            ),
        },
        {
            "Limitation": (
                "Training labels were primarily produced by "
                "the rule-based classifier."
            ),
            "Implication": (
                "The ML model learns to approximate weak labels, "
                "not independently verified ground truth."
            ),
        },
        {
            "Limitation": (
                "Only topics with sufficient positive and "
                "negative examples were included."
            ),
            "Implication": (
                "The ML model does not currently cover all "
                "19 policy topics."
            ),
        },
        {
            "Limitation": (
                "Candidate statements are multi-label."
            ),
            "Implication": (
                "One statement can correctly belong to several "
                "policy categories at the same time."
            ),
        },
        {
            "Limitation": (
                "Several statements are structured summaries."
            ),
            "Implication": (
                "Results may be stronger than performance on "
                "messier real-world campaign text."
            ),
        },
    ]
)


# ============================================================
# SAVE OUTPUTS
# ============================================================

OUTPUT_EXCEL.parent.mkdir(
    parents=True,
    exist_ok=True,
)

statement_predictions.to_csv(
    OUTPUT_CSV,
    index=False,
    encoding="utf-8-sig",
)

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl",
) as writer:

    overall_metrics.to_excel(
        writer,
        sheet_name="OverallMetrics",
        index=False,
    )

    topic_support.to_excel(
        writer,
        sheet_name="TopicSupport",
        index=False,
    )

    topic_metrics.to_excel(
        writer,
        sheet_name="TopicMetrics",
        index=False,
    )

    statement_predictions.to_excel(
        writer,
        sheet_name="StatementPredictions",
        index=False,
    )

    fold_predictions.to_excel(
        writer,
        sheet_name="FoldPredictions",
        index=False,
    )

    model_coefficients.to_excel(
        writer,
        sheet_name="ModelCoefficients",
        index=False,
    )

    target_output = pd.concat(
        [
            statements[
                [
                    "Statement_ID",
                    "Candidate_ID",
                    "Statement_Type",
                    "Statement_Text",
                ]
            ],
            target_dataframe,
        ],
        axis=1,
    )

    target_output.to_excel(
        writer,
        sheet_name="TrainingLabels",
        index=False,
    )

    limitations.to_excel(
        writer,
        sheet_name="Limitations",
        index=False,
    )

    workbook = writer.book

    for worksheet in workbook.worksheets:
        style_worksheet(
            worksheet
        )


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("ML POLICY CLASSIFIER PILOT COMPLETE")
print("=" * 72)

print(f"Statement workbook:\n{STATEMENT_FILE}")
print()

print(f"Policy labels:\n{POLICY_FILE}")
print()

print(f"Excel output:\n{OUTPUT_EXCEL}")
print()

print(f"Prediction CSV:\n{OUTPUT_CSV}")
print()

print(
    f"Statements used: {len(statements)}"
)

print(
    f"Available policy topics: {len(all_topics)}"
)

print(
    f"ML-eligible policy topics: "
    f"{len(eligible_topics)}"
)

print(
    "Eligible topics: "
    + "; ".join(eligible_topics)
)

print()

print(
    f"Micro precision: "
    f"{micro_precision * 100:.2f}%"
)

print(
    f"Micro recall: "
    f"{micro_recall * 100:.2f}%"
)

print(
    f"Micro F1: "
    f"{micro_f1 * 100:.2f}%"
)

print(
    f"Macro F1: "
    f"{macro_f1 * 100:.2f}%"
)

print(
    f"Subset accuracy: "
    f"{subset_accuracy * 100:.2f}%"
)

print(
    f"Hamming accuracy: "
    f"{hamming_accuracy * 100:.2f}%"
)

print()

print(
    "Important: this is a small weakly supervised "
    "proof-of-concept, not a production model."
)

print()

print(
    "Next step: review OverallMetrics, TopicMetrics, "
    "StatementPredictions, ModelCoefficients, and "
    "Limitations."
)