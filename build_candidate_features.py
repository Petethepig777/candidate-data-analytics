from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "political_analytics_dashboard_data.xlsx"
)

OUTPUT_EXCEL = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_feature_dataset.xlsx"
)

OUTPUT_CSV = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_feature_dataset.csv"
)


# ============================================================
# CONFIGURATION
# ============================================================

IDENTIFIER_COLUMNS = [
    "Candidate_ID",
    "Ballot_Name",
]

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop with a clear message when a required file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to blank text and normalize whitespace.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def safe_percentage(numerator, denominator):
    """
    Calculate a percentage safely.
    """
    if denominator == 0:
        return 0.0

    return round(
        numerator / denominator * 100,
        2,
    )


def style_worksheet(worksheet):
    """
    Apply consistent formatting to an Excel worksheet.
    """
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            45,
        )


# ============================================================
# LOAD POLICY MATRIX
# ============================================================

require_file(
    INPUT_FILE,
    "Dashboard dataset workbook",
)

excel_file = pd.ExcelFile(
    INPUT_FILE
)

required_sheet = "CandidatePolicyMatrix"

if required_sheet not in excel_file.sheet_names:
    raise ValueError(
        f"The workbook does not contain "
        f"the '{required_sheet}' sheet."
    )

policy_matrix = pd.read_excel(
    INPUT_FILE,
    sheet_name=required_sheet,
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# VALIDATE COLUMNS
# ============================================================

missing_identifier_columns = [
    column
    for column in IDENTIFIER_COLUMNS
    if column not in policy_matrix.columns
]

if missing_identifier_columns:
    raise ValueError(
        "CandidatePolicyMatrix is missing required columns: "
        + ", ".join(missing_identifier_columns)
    )

policy_columns = [
    column
    for column in policy_matrix.columns
    if column not in IDENTIFIER_COLUMNS
]

if not policy_columns:
    raise ValueError(
        "No policy-topic columns were found."
    )


# ============================================================
# CLEAN DATA
# ============================================================

for column in IDENTIFIER_COLUMNS:
    policy_matrix[column] = policy_matrix[
        column
    ].apply(clean_text)

for column in policy_columns:
    policy_matrix[column] = pd.to_numeric(
        policy_matrix[column],
        errors="coerce",
    ).fillna(0)

policy_matrix = policy_matrix[
    policy_matrix["Candidate_ID"] != ""
].copy()

policy_matrix.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# ENGINEER CANDIDATE FEATURES
# ============================================================

feature_rows = []

for _, row in policy_matrix.iterrows():
    policy_values = row[
        policy_columns
    ].astype(float)

    total_policy_weight = float(
        policy_values.sum()
    )

    positive_policy_values = policy_values[
        policy_values > 0
    ]

    policy_diversity = int(
        positive_policy_values.count()
    )

    if total_policy_weight > 0:
        dominant_policy = str(
            policy_values.idxmax()
        )

        dominant_policy_weight = float(
            policy_values.max()
        )

        dominant_policy_percent = (
            safe_percentage(
                dominant_policy_weight,
                total_policy_weight,
            )
        )

    else:
        dominant_policy = ""
        dominant_policy_weight = 0.0
        dominant_policy_percent = 0.0

    if policy_diversity > 0:
        average_policy_weight = round(
            total_policy_weight
            / policy_diversity,
            2,
        )

    else:
        average_policy_weight = 0.0

    feature_rows.append(
        {
            "Candidate_ID": row[
                "Candidate_ID"
            ],
            "Ballot_Name": row[
                "Ballot_Name"
            ],
            "Total_Policy_Weight": round(
                total_policy_weight,
                2,
            ),
            "Policy_Diversity": (
                policy_diversity
            ),
            "Dominant_Policy": (
                dominant_policy
            ),
            "Dominant_Policy_Weight": round(
                dominant_policy_weight,
                2,
            ),
            "Dominant_Policy_Percent": (
                dominant_policy_percent
            ),
            "Average_Policy_Weight": (
                average_policy_weight
            ),
        }
    )

candidate_features = pd.DataFrame(
    feature_rows
)


# ============================================================
# ADD SIMPLE RANKINGS
# ============================================================

candidate_features[
    "Total_Policy_Weight_Rank"
] = candidate_features[
    "Total_Policy_Weight"
].rank(
    method="dense",
    ascending=False,
).astype(int)

candidate_features[
    "Policy_Diversity_Rank"
] = candidate_features[
    "Policy_Diversity"
].rank(
    method="dense",
    ascending=False,
).astype(int)


# ============================================================
# CREATE FEATURE STATISTICS
# ============================================================

numeric_feature_columns = [
    "Total_Policy_Weight",
    "Policy_Diversity",
    "Dominant_Policy_Weight",
    "Dominant_Policy_Percent",
    "Average_Policy_Weight",
]

feature_statistics_rows = []

for column in numeric_feature_columns:
    feature_statistics_rows.append(
        {
            "Feature": column,
            "Minimum": round(
                candidate_features[column].min(),
                2,
            ),
            "Maximum": round(
                candidate_features[column].max(),
                2,
            ),
            "Mean": round(
                candidate_features[column].mean(),
                2,
            ),
            "Median": round(
                candidate_features[column].median(),
                2,
            ),
            "Standard_Deviation": round(
                candidate_features[column].std(
                    ddof=0
                ),
                2,
            ),
        }
    )

feature_statistics = pd.DataFrame(
    feature_statistics_rows
)


# ============================================================
# CREATE DOMINANT-POLICY SUMMARY
# ============================================================

dominant_policy_summary = (
    candidate_features
    .groupby(
        "Dominant_Policy",
        as_index=False,
    )
    .agg(
        Candidate_Count=(
            "Candidate_ID",
            "count",
        ),
        Average_Dominant_Percent=(
            "Dominant_Policy_Percent",
            "mean",
        ),
    )
)

dominant_policy_summary[
    "Average_Dominant_Percent"
] = dominant_policy_summary[
    "Average_Dominant_Percent"
].round(2)

dominant_policy_summary.sort_values(
    by=[
        "Candidate_Count",
        "Average_Dominant_Percent",
    ],
    ascending=[
        False,
        False,
    ],
    inplace=True,
)

dominant_policy_summary.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# CREATE RUN SUMMARY
# ============================================================

highest_weight_candidate = (
    candidate_features
    .sort_values(
        by="Total_Policy_Weight",
        ascending=False,
    )
    .iloc[0]
)

highest_diversity_candidate = (
    candidate_features
    .sort_values(
        by="Policy_Diversity",
        ascending=False,
    )
    .iloc[0]
)

most_concentrated_candidate = (
    candidate_features
    .sort_values(
        by="Dominant_Policy_Percent",
        ascending=False,
    )
    .iloc[0]
)

run_summary = pd.DataFrame(
    [
        {
            "Metric": "Candidates Processed",
            "Value": len(
                candidate_features
            ),
        },
        {
            "Metric": "Policy Features Used",
            "Value": len(
                policy_columns
            ),
        },
        {
            "Metric": (
                "Candidate with Highest Total Policy Weight"
            ),
            "Value": highest_weight_candidate[
                "Ballot_Name"
            ],
        },
        {
            "Metric": (
                "Highest Total Policy Weight"
            ),
            "Value": highest_weight_candidate[
                "Total_Policy_Weight"
            ],
        },
        {
            "Metric": (
                "Candidate with Highest Policy Diversity"
            ),
            "Value": highest_diversity_candidate[
                "Ballot_Name"
            ],
        },
        {
            "Metric": (
                "Highest Policy Diversity"
            ),
            "Value": highest_diversity_candidate[
                "Policy_Diversity"
            ],
        },
        {
            "Metric": (
                "Most Concentrated Policy Profile"
            ),
            "Value": most_concentrated_candidate[
                "Ballot_Name"
            ],
        },
        {
            "Metric": (
                "Highest Dominant Policy Percent"
            ),
            "Value": most_concentrated_candidate[
                "Dominant_Policy_Percent"
            ],
        },
        {
            "Metric": "Feature Dataset Created",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
    ]
)


# ============================================================
# SAVE CSV
# ============================================================

OUTPUT_EXCEL.parent.mkdir(
    parents=True,
    exist_ok=True,
)

candidate_features.to_csv(
    OUTPUT_CSV,
    index=False,
    encoding="utf-8-sig",
)


# ============================================================
# SAVE EXCEL WORKBOOK
# ============================================================

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl",
) as writer:

    run_summary.to_excel(
        writer,
        sheet_name="RunSummary",
        index=False,
    )

    candidate_features.to_excel(
        writer,
        sheet_name="CandidateFeatures",
        index=False,
    )

    feature_statistics.to_excel(
        writer,
        sheet_name="FeatureStatistics",
        index=False,
    )

    dominant_policy_summary.to_excel(
        writer,
        sheet_name="DominantPolicySummary",
        index=False,
    )

    policy_matrix.to_excel(
        writer,
        sheet_name="PolicyVectors",
        index=False,
    )

    workbook = writer.book

    for worksheet in workbook.worksheets:
        style_worksheet(
            worksheet
        )

    feature_sheet = workbook[
        "CandidateFeatures"
    ]

    feature_headers = {
        cell.value: cell.column
        for cell in feature_sheet[1]
    }

    if (
        "Dominant_Policy_Percent"
        in feature_headers
    ):
        percent_column = feature_headers[
            "Dominant_Policy_Percent"
        ]

        percent_letter = get_column_letter(
            percent_column
        )

        for row_number in range(
            2,
            feature_sheet.max_row + 1,
        ):
            feature_sheet.cell(
                row=row_number,
                column=percent_column,
            ).number_format = '0.00"%"'

        feature_sheet.conditional_formatting.add(
            (
                f"{percent_letter}2:"
                f"{percent_letter}"
                f"{feature_sheet.max_row}"
            ),
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    for feature_name in [
        "Total_Policy_Weight",
        "Policy_Diversity",
        "Average_Policy_Weight",
    ]:
        if feature_name not in feature_headers:
            continue

        column_number = feature_headers[
            feature_name
        ]

        column_letter = get_column_letter(
            column_number
        )

        feature_sheet.conditional_formatting.add(
            (
                f"{column_letter}2:"
                f"{column_letter}"
                f"{feature_sheet.max_row}"
            ),
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="D9EAF7",
                end_type="max",
                end_color="5B9BD5",
            ),
        )


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("CANDIDATE FEATURE DATASET CREATED")
print("=" * 72)

print(f"Input workbook:\n{INPUT_FILE}")
print()

print(f"Excel output:\n{OUTPUT_EXCEL}")
print()

print(f"CSV output:\n{OUTPUT_CSV}")
print()

print(
    f"Candidates processed: "
    f"{len(candidate_features)}"
)

print(
    f"Policy features used: "
    f"{len(policy_columns)}"
)

print()

print(
    "Candidate with highest total policy weight: "
    f"{highest_weight_candidate['Ballot_Name']} "
    f"({highest_weight_candidate['Total_Policy_Weight']:.2f})"
)

print(
    "Candidate with highest policy diversity: "
    f"{highest_diversity_candidate['Ballot_Name']} "
    f"({int(highest_diversity_candidate['Policy_Diversity'])} topics)"
)

print(
    "Most concentrated policy profile: "
    f"{most_concentrated_candidate['Ballot_Name']} — "
    f"{most_concentrated_candidate['Dominant_Policy']} "
    f"({most_concentrated_candidate['Dominant_Policy_Percent']:.2f}%)"
)

print()

print(
    "Next step: review CandidateFeatures and "
    "FeatureStatistics in candidate_feature_dataset.xlsx."
)