from pathlib import Path
from datetime import datetime
import re

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "classifier_manual_review_completed.xlsx"
)

OUTPUT_EXCEL = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "classifier_evaluation.xlsx"
)

OUTPUT_CSV = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "classifier_metrics.csv"
)


# ============================================================
# CONFIGURATION
# ============================================================

VALID_LABELS = {
    "Correct",
    "Partially Correct",
    "Incorrect",
}

VALID_CONFIDENCE_LEVELS = {
    "High",
    "Medium",
    "Low",
    "Unknown",
}

LABEL_SCORE_MAP = {
    "Correct": 1.0,
    "Partially Correct": 0.5,
    "Incorrect": 0.0,
}

STRICT_CORRECT_MAP = {
    "Correct": 1,
    "Partially Correct": 0,
    "Incorrect": 0,
}

RELEVANT_MAP = {
    "Correct": 1,
    "Partially Correct": 1,
    "Incorrect": 0,
}


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_text(value):
    """
    Convert a value to clean text.

    Missing values become an empty string.
    """
    if pd.isna(value):
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )


def normalize_manual_label(value):
    """
    Normalize common manual-label variations.
    """
    text = clean_text(value).lower()

    label_map = {
        "correct": "Correct",
        "partially correct": "Partially Correct",
        "partial": "Partially Correct",
        "partially": "Partially Correct",
        "incorrect": "Incorrect",
        "wrong": "Incorrect",
    }

    return label_map.get(
        text,
        clean_text(value),
    )


def normalize_confidence(value):
    """
    Normalize confidence labels.
    """
    text = clean_text(value).lower()

    confidence_map = {
        "high": "High",
        "medium": "Medium",
        "low": "Low",
        "unknown": "Unknown",
    }

    return confidence_map.get(
        text,
        clean_text(value),
    )


def safe_percentage(numerator, denominator):
    """
    Calculate a percentage safely.
    """
    if denominator == 0:
        return 0.0

    return round(
        numerator / denominator * 100,
        2,
    )


def calculate_group_metrics(group):
    """
    Calculate classifier-evaluation metrics for a group.
    """
    reviewed_count = len(group)

    correct_count = int(
        (
            group["Manual_Label"]
            == "Correct"
        ).sum()
    )

    partial_count = int(
        (
            group["Manual_Label"]
            == "Partially Correct"
        ).sum()
    )

    incorrect_count = int(
        (
            group["Manual_Label"]
            == "Incorrect"
        ).sum()
    )

    strict_precision = safe_percentage(
        correct_count,
        reviewed_count,
    )

    relevance_rate = safe_percentage(
        correct_count + partial_count,
        reviewed_count,
    )

    weighted_precision = round(
        group["Manual_Label_Score"].mean()
        * 100,
        2,
    ) if reviewed_count > 0 else 0.0

    return pd.Series(
        {
            "Reviewed_Count": reviewed_count,
            "Correct_Count": correct_count,
            "Partially_Correct_Count": partial_count,
            "Incorrect_Count": incorrect_count,
            "Strict_Precision_Percent": strict_precision,
            "Relevant_or_Partial_Percent": relevance_rate,
            "Weighted_Precision_Percent": weighted_precision,
            "Average_Classification_Score": round(
                pd.to_numeric(
                    group[
                        "Classification_Score"
                    ],
                    errors="coerce",
                ).mean(),
                2,
            ),
            "Average_Keyword_Match_Count": round(
                pd.to_numeric(
                    group[
                        "Keyword_Match_Count"
                    ],
                    errors="coerce",
                ).mean(),
                2,
            ),
        }
    )


def split_keyword_field(value):
    """
    Split a semicolon-separated keyword field into individual keywords.
    """
    text = clean_text(value)

    if not text:
        return []

    return [
        keyword.strip()
        for keyword in text.split(";")
        if keyword.strip()
    ]


def classify_recommendation(
    reviewed_count,
    strict_precision,
    relevance_rate,
):
    """
    Generate a recommendation based on topic-level performance.
    """
    if reviewed_count < 2:
        return (
            "Collect more labeled examples before changing "
            "this topic's keyword rules."
        )

    if strict_precision >= 80:
        return (
            "Keep the current rules. Performance is strong; "
            "continue monitoring as the dataset grows."
        )

    if relevance_rate >= 80 and strict_precision < 80:
        return (
            "The topic is usually relevant but often secondary. "
            "Increase the confidence threshold or require multiple "
            "topic-specific keyword matches."
        )

    if strict_precision >= 60:
        return (
            "Review broad or ambiguous keywords and prioritize "
            "longer, more policy-specific phrases."
        )

    return (
        "Substantial refinement is needed. Remove ambiguous keywords, "
        "add context rules, and consider excluding single-keyword matches."
    )


def style_output_workbook(workbook):
    """
    Apply consistent formatting to all output worksheets.
    """
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    subheader_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            column_letter = get_column_letter(
                column_number
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                65,
            )

        worksheet.row_dimensions[1].height = 32

    if "RunSummary" in workbook.sheetnames:
        summary_sheet = workbook[
            "RunSummary"
        ]

        summary_sheet.column_dimensions[
            "A"
        ].width = 38

        summary_sheet.column_dimensions[
            "B"
        ].width = 70

    if "Recommendations" in workbook.sheetnames:
        recommendation_sheet = workbook[
            "Recommendations"
        ]

        recommendation_sheet.column_dimensions[
            "A"
        ].width = 28

        recommendation_sheet.column_dimensions[
            "H"
        ].width = 75

    percentage_sheets = {
        "OverallMetrics": [
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ],
        "TopicMetrics": [
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ],
        "ConfidenceMetrics": [
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ],
        "CandidateMetrics": [
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ],
        "StatementTypeMetrics": [
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ],
    }

    for (
        sheet_name,
        percentage_columns,
    ) in percentage_sheets.items():

        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[
            sheet_name
        ]

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        for column_name in percentage_columns:
            if column_name not in headers:
                continue

            column_number = headers[
                column_name
            ]

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = '0.00"%"'

            column_letter = get_column_letter(
                column_number
            )

            worksheet.conditional_formatting.add(
                (
                    f"{column_letter}2:"
                    f"{column_letter}{worksheet.max_row}"
                ),
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="F8696B",
                    mid_type="num",
                    mid_value=60,
                    mid_color="FFEB84",
                    end_type="num",
                    end_value=100,
                    end_color="63BE7B",
                ),
            )


# ============================================================
# LOAD MANUAL REVIEW WORKBOOK
# ============================================================

if not INPUT_FILE.exists():
    raise FileNotFoundError(
        f"Completed manual-review workbook was not found:\n{INPUT_FILE}"
    )

excel_file = pd.ExcelFile(
    INPUT_FILE
)

if "ManualReview" not in excel_file.sheet_names:
    raise ValueError(
        "The workbook does not contain a ManualReview sheet."
    )

reviews = pd.read_excel(
    INPUT_FILE,
    sheet_name="ManualReview",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# VALIDATE REQUIRED COLUMNS
# ============================================================

required_columns = [
    "Review_ID",
    "Policy_Record_ID",
    "Candidate_ID",
    "Ballot_Name",
    "Statement_ID",
    "Statement_Type",
    "Policy_Topic",
    "Position_Summary",
    "Matched_Keywords",
    "Keyword_Match_Count",
    "Classification_Score",
    "Classification_Confidence",
    "Manual_Label",
    "Reviewer_Confidence",
    "Incorrect_Keyword",
    "Suggested_Correct_Topic",
    "Reviewer_Notes",
]

missing_columns = [
    column
    for column in required_columns
    if column not in reviews.columns
]

if missing_columns:
    raise ValueError(
        "ManualReview is missing required columns: "
        + ", ".join(missing_columns)
    )


# ============================================================
# CLEAN MANUAL LABELS
# ============================================================

text_columns = [
    "Review_ID",
    "Policy_Record_ID",
    "Candidate_ID",
    "Ballot_Name",
    "Statement_ID",
    "Statement_Type",
    "Policy_Topic",
    "Position_Summary",
    "Matched_Keywords",
    "Classification_Confidence",
    "Manual_Label",
    "Reviewer_Confidence",
    "Incorrect_Keyword",
    "Suggested_Correct_Topic",
    "Reviewer_Notes",
]

for column in text_columns:
    reviews[column] = reviews[
        column
    ].apply(clean_text)

reviews["Manual_Label"] = reviews[
    "Manual_Label"
].apply(normalize_manual_label)

reviews[
    "Classification_Confidence"
] = reviews[
    "Classification_Confidence"
].apply(normalize_confidence)

reviews[
    "Reviewer_Confidence"
] = reviews[
    "Reviewer_Confidence"
].apply(normalize_confidence)


# ============================================================
# SEPARATE REVIEWED AND UNREVIEWED ROWS
# ============================================================

reviewed = reviews[
    reviews["Manual_Label"].isin(
        VALID_LABELS
    )
].copy()

unreviewed = reviews[
    ~reviews["Manual_Label"].isin(
        VALID_LABELS
    )
].copy()

invalid_labels = reviews[
    (
        reviews["Manual_Label"] != ""
    )
    & (
        ~reviews["Manual_Label"].isin(
            VALID_LABELS
        )
    )
].copy()

if reviewed.empty:
    raise ValueError(
        "No valid manual labels were found. "
        "Complete the Manual_Label column before running this script."
    )


# ============================================================
# CREATE DERIVED EVALUATION FIELDS
# ============================================================

reviewed["Manual_Label_Score"] = reviewed[
    "Manual_Label"
].map(
    LABEL_SCORE_MAP
)

reviewed["Strictly_Correct"] = reviewed[
    "Manual_Label"
].map(
    STRICT_CORRECT_MAP
)

reviewed["Relevant_or_Partial"] = reviewed[
    "Manual_Label"
].map(
    RELEVANT_MAP
)

reviewed[
    "Keyword_Match_Count"
] = pd.to_numeric(
    reviewed[
        "Keyword_Match_Count"
    ],
    errors="coerce",
)

reviewed[
    "Classification_Score"
] = pd.to_numeric(
    reviewed[
        "Classification_Score"
    ],
    errors="coerce",
)


# ============================================================
# OVERALL METRICS
# ============================================================

overall_metrics_series = (
    calculate_group_metrics(
        reviewed
    )
)

overall_metrics = pd.DataFrame(
    [
        {
            "Evaluation_Scope": (
                "All manually reviewed classifications"
            ),
            **overall_metrics_series.to_dict(),
        }
    ]
)


# ============================================================
# METRICS BY POLICY TOPIC
# ============================================================

topic_metrics = (
    reviewed
    .groupby(
        "Policy_Topic",
        dropna=False,
    )
    .apply(
        calculate_group_metrics
    )
    .reset_index()
)

topic_metrics.sort_values(
    by=[
        "Strict_Precision_Percent",
        "Reviewed_Count",
        "Policy_Topic",
    ],
    ascending=[
        True,
        False,
        True,
    ],
    inplace=True,
)

topic_metrics.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# METRICS BY CLASSIFIER CONFIDENCE
# ============================================================

confidence_metrics = (
    reviewed
    .groupby(
        "Classification_Confidence",
        dropna=False,
    )
    .apply(
        calculate_group_metrics
    )
    .reset_index()
)

confidence_order = {
    "High": 1,
    "Medium": 2,
    "Low": 3,
    "Unknown": 4,
    "": 5,
}

confidence_metrics[
    "Sort_Order"
] = confidence_metrics[
    "Classification_Confidence"
].map(
    confidence_order
).fillna(6)

confidence_metrics.sort_values(
    by="Sort_Order",
    inplace=True,
)

confidence_metrics.drop(
    columns=["Sort_Order"],
    inplace=True,
)

confidence_metrics.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# METRICS BY CANDIDATE
# ============================================================

candidate_metrics = (
    reviewed
    .groupby(
        [
            "Candidate_ID",
            "Ballot_Name",
        ],
        dropna=False,
    )
    .apply(
        calculate_group_metrics
    )
    .reset_index()
)

candidate_metrics.sort_values(
    by=[
        "Weighted_Precision_Percent",
        "Ballot_Name",
    ],
    ascending=[
        False,
        True,
    ],
    inplace=True,
)

candidate_metrics.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# METRICS BY STATEMENT TYPE
# ============================================================

statement_type_metrics = (
    reviewed
    .groupby(
        "Statement_Type",
        dropna=False,
    )
    .apply(
        calculate_group_metrics
    )
    .reset_index()
)

statement_type_metrics.sort_values(
    by=[
        "Strict_Precision_Percent",
        "Reviewed_Count",
    ],
    ascending=[
        True,
        False,
    ],
    inplace=True,
)

statement_type_metrics.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# LABEL DISTRIBUTION
# ============================================================

label_distribution = (
    reviewed[
        "Manual_Label"
    ]
    .value_counts()
    .rename_axis(
        "Manual_Label"
    )
    .reset_index(
        name="Record_Count"
    )
)

label_distribution[
    "Percent_of_Reviewed"
] = label_distribution[
    "Record_Count"
].apply(
    lambda value: safe_percentage(
        value,
        len(reviewed),
    )
)


# ============================================================
# INCORRECT AND PARTIAL PREDICTIONS
# ============================================================

problem_predictions = reviewed[
    reviewed["Manual_Label"].isin(
        [
            "Incorrect",
            "Partially Correct",
        ]
    )
].copy()

problem_predictions.sort_values(
    by=[
        "Manual_Label",
        "Classification_Confidence",
        "Policy_Topic",
        "Ballot_Name",
    ],
    inplace=True,
)

incorrect_predictions = reviewed[
    reviewed[
        "Manual_Label"
    ]
    == "Incorrect"
].copy()

incorrect_predictions.sort_values(
    by=[
        "Classification_Confidence",
        "Policy_Topic",
        "Ballot_Name",
    ],
    inplace=True,
)


# ============================================================
# KEYWORD-LEVEL ERROR ANALYSIS
# ============================================================

keyword_rows = []

for _, row in reviewed.iterrows():
    matched_keywords = split_keyword_field(
        row["Matched_Keywords"]
    )

    manually_flagged_keywords = (
        split_keyword_field(
            row["Incorrect_Keyword"]
        )
    )

    manually_flagged_normalized = {
        keyword.lower()
        for keyword in manually_flagged_keywords
    }

    for keyword in matched_keywords:
        keyword_rows.append(
            {
                "Policy_Record_ID": (
                    row["Policy_Record_ID"]
                ),
                "Policy_Topic": (
                    row["Policy_Topic"]
                ),
                "Keyword": keyword,
                "Manual_Label": (
                    row["Manual_Label"]
                ),
                "Manual_Label_Score": (
                    row[
                        "Manual_Label_Score"
                    ]
                ),
                "Explicitly_Flagged_Incorrect": (
                    "Yes"
                    if keyword.lower()
                    in manually_flagged_normalized
                    else "No"
                ),
            }
        )

keyword_detail = pd.DataFrame(
    keyword_rows
)

if keyword_detail.empty:
    keyword_metrics = pd.DataFrame(
        columns=[
            "Policy_Topic",
            "Keyword",
            "Reviewed_Uses",
            "Correct_Uses",
            "Partially_Correct_Uses",
            "Incorrect_Uses",
            "Explicitly_Flagged_Incorrect_Count",
            "Strict_Precision_Percent",
            "Relevant_or_Partial_Percent",
            "Weighted_Precision_Percent",
        ]
    )

else:
    keyword_metrics = (
        keyword_detail
        .groupby(
            [
                "Policy_Topic",
                "Keyword",
            ],
            as_index=False,
        )
        .agg(
            Reviewed_Uses=(
                "Policy_Record_ID",
                "count",
            ),
            Correct_Uses=(
                "Manual_Label",
                lambda series: int(
                    (
                        series
                        == "Correct"
                    ).sum()
                ),
            ),
            Partially_Correct_Uses=(
                "Manual_Label",
                lambda series: int(
                    (
                        series
                        == "Partially Correct"
                    ).sum()
                ),
            ),
            Incorrect_Uses=(
                "Manual_Label",
                lambda series: int(
                    (
                        series
                        == "Incorrect"
                    ).sum()
                ),
            ),
            Explicitly_Flagged_Incorrect_Count=(
                "Explicitly_Flagged_Incorrect",
                lambda series: int(
                    (
                        series
                        == "Yes"
                    ).sum()
                ),
            ),
            Average_Label_Score=(
                "Manual_Label_Score",
                "mean",
            ),
        )
    )

    keyword_metrics[
        "Strict_Precision_Percent"
    ] = keyword_metrics.apply(
        lambda row: safe_percentage(
            row["Correct_Uses"],
            row["Reviewed_Uses"],
        ),
        axis=1,
    )

    keyword_metrics[
        "Relevant_or_Partial_Percent"
    ] = keyword_metrics.apply(
        lambda row: safe_percentage(
            (
                row["Correct_Uses"]
                + row[
                    "Partially_Correct_Uses"
                ]
            ),
            row["Reviewed_Uses"],
        ),
        axis=1,
    )

    keyword_metrics[
        "Weighted_Precision_Percent"
    ] = (
        keyword_metrics[
            "Average_Label_Score"
        ]
        * 100
    ).round(2)

    keyword_metrics.drop(
        columns=[
            "Average_Label_Score"
        ],
        inplace=True,
    )

    keyword_metrics.sort_values(
        by=[
            "Incorrect_Uses",
            "Explicitly_Flagged_Incorrect_Count",
            "Strict_Precision_Percent",
            "Reviewed_Uses",
        ],
        ascending=[
            False,
            False,
            True,
            False,
        ],
        inplace=True,
    )

    keyword_metrics.reset_index(
        drop=True,
        inplace=True,
    )


# ============================================================
# TOPIC RECOMMENDATIONS
# ============================================================

recommendation_rows = []

for _, row in topic_metrics.iterrows():
    topic = row[
        "Policy_Topic"
    ]

    topic_keywords = keyword_metrics[
        keyword_metrics[
            "Policy_Topic"
        ]
        == topic
    ].copy()

    problematic_keywords = topic_keywords[
        (
            topic_keywords[
                "Incorrect_Uses"
            ]
            > 0
        )
        | (
            topic_keywords[
                "Strict_Precision_Percent"
            ]
            < 60
        )
    ]

    problematic_keyword_text = "; ".join(
        problematic_keywords[
            "Keyword"
        ]
        .head(8)
        .tolist()
    )

    recommendation_rows.append(
        {
            "Policy_Topic": topic,
            "Reviewed_Count": int(
                row["Reviewed_Count"]
            ),
            "Correct_Count": int(
                row["Correct_Count"]
            ),
            "Partially_Correct_Count": int(
                row[
                    "Partially_Correct_Count"
                ]
            ),
            "Incorrect_Count": int(
                row["Incorrect_Count"]
            ),
            "Strict_Precision_Percent": (
                row[
                    "Strict_Precision_Percent"
                ]
            ),
            "Relevant_or_Partial_Percent": (
                row[
                    "Relevant_or_Partial_Percent"
                ]
            ),
            "Problematic_Keywords": (
                problematic_keyword_text
            ),
            "Recommendation": (
                classify_recommendation(
                    int(
                        row[
                            "Reviewed_Count"
                        ]
                    ),
                    float(
                        row[
                            "Strict_Precision_Percent"
                        ]
                    ),
                    float(
                        row[
                            "Relevant_or_Partial_Percent"
                        ]
                    ),
                )
            ),
        }
    )

recommendations = pd.DataFrame(
    recommendation_rows
)

recommendations.sort_values(
    by=[
        "Strict_Precision_Percent",
        "Incorrect_Count",
        "Policy_Topic",
    ],
    ascending=[
        True,
        False,
        True,
    ],
    inplace=True,
)

recommendations.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# SUGGESTED CORRECT-TOPIC ANALYSIS
# ============================================================

suggested_topic_rows = reviewed[
    reviewed[
        "Suggested_Correct_Topic"
    ]
    != ""
].copy()

if suggested_topic_rows.empty:
    suggested_topic_summary = pd.DataFrame(
        columns=[
            "Predicted_Policy_Topic",
            "Suggested_Correct_Topic",
            "Record_Count",
        ]
    )

else:
    suggested_topic_summary = (
        suggested_topic_rows
        .groupby(
            [
                "Policy_Topic",
                "Suggested_Correct_Topic",
            ],
            as_index=False,
        )
        .size()
        .rename(
            columns={
                "Policy_Topic": (
                    "Predicted_Policy_Topic"
                ),
                "size": "Record_Count",
            }
        )
        .sort_values(
            by="Record_Count",
            ascending=False,
        )
        .reset_index(drop=True)
    )


# ============================================================
# RUN SUMMARY
# ============================================================

run_summary = pd.DataFrame(
    [
        {
            "Metric": "Input Workbook",
            "Value": str(INPUT_FILE),
        },
        {
            "Metric": "Evaluation Run Date",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
        {
            "Metric": "Total Sampled Records",
            "Value": len(reviews),
        },
        {
            "Metric": "Valid Reviewed Records",
            "Value": len(reviewed),
        },
        {
            "Metric": "Unreviewed Records",
            "Value": len(unreviewed),
        },
        {
            "Metric": "Invalid Manual Labels",
            "Value": len(invalid_labels),
        },
        {
            "Metric": "Correct Predictions",
            "Value": int(
                (
                    reviewed[
                        "Manual_Label"
                    ]
                    == "Correct"
                ).sum()
            ),
        },
        {
            "Metric": "Partially Correct Predictions",
            "Value": int(
                (
                    reviewed[
                        "Manual_Label"
                    ]
                    == "Partially Correct"
                ).sum()
            ),
        },
        {
            "Metric": "Incorrect Predictions",
            "Value": int(
                (
                    reviewed[
                        "Manual_Label"
                    ]
                    == "Incorrect"
                ).sum()
            ),
        },
        {
            "Metric": "Strict Precision Percent",
            "Value": overall_metrics.loc[
                0,
                "Strict_Precision_Percent",
            ],
        },
        {
            "Metric": (
                "Relevant or Partial Percent"
            ),
            "Value": overall_metrics.loc[
                0,
                "Relevant_or_Partial_Percent",
            ],
        },
        {
            "Metric": (
                "Weighted Precision Percent"
            ),
            "Value": overall_metrics.loc[
                0,
                "Weighted_Precision_Percent",
            ],
        },
        {
            "Metric": "Policy Topics Evaluated",
            "Value": reviewed[
                "Policy_Topic"
            ].nunique(),
        },
        {
            "Metric": "Candidates Represented",
            "Value": reviewed[
                "Candidate_ID"
            ].nunique(),
        },
    ]
)


# ============================================================
# EXPORT METRIC CSV
# ============================================================

OUTPUT_EXCEL.parent.mkdir(
    parents=True,
    exist_ok=True,
)

overall_metrics.to_csv(
    OUTPUT_CSV,
    index=False,
    encoding="utf-8-sig",
)


# ============================================================
# EXPORT EXCEL EVALUATION REPORT
# ============================================================

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl",
) as writer:

    run_summary.to_excel(
        writer,
        sheet_name="RunSummary",
        index=False,
    )

    overall_metrics.to_excel(
        writer,
        sheet_name="OverallMetrics",
        index=False,
    )

    label_distribution.to_excel(
        writer,
        sheet_name="LabelDistribution",
        index=False,
    )

    topic_metrics.to_excel(
        writer,
        sheet_name="TopicMetrics",
        index=False,
    )

    confidence_metrics.to_excel(
        writer,
        sheet_name="ConfidenceMetrics",
        index=False,
    )

    candidate_metrics.to_excel(
        writer,
        sheet_name="CandidateMetrics",
        index=False,
    )

    statement_type_metrics.to_excel(
        writer,
        sheet_name="StatementTypeMetrics",
        index=False,
    )

    keyword_metrics.to_excel(
        writer,
        sheet_name="KeywordMetrics",
        index=False,
    )

    problem_predictions.to_excel(
        writer,
        sheet_name="ProblemPredictions",
        index=False,
    )

    incorrect_predictions.to_excel(
        writer,
        sheet_name="IncorrectPredictions",
        index=False,
    )

    recommendations.to_excel(
        writer,
        sheet_name="Recommendations",
        index=False,
    )

    suggested_topic_summary.to_excel(
        writer,
        sheet_name="SuggestedTopicChanges",
        index=False,
    )

    reviewed.to_excel(
        writer,
        sheet_name="ReviewedRecords",
        index=False,
    )

    unreviewed.to_excel(
        writer,
        sheet_name="UnreviewedRecords",
        index=False,
    )

    invalid_labels.to_excel(
        writer,
        sheet_name="InvalidLabels",
        index=False,
    )

    output_workbook = writer.book

    style_output_workbook(
        output_workbook
    )


# ============================================================
# TERMINAL OUTPUT
# ============================================================

correct_count = int(
    (
        reviewed["Manual_Label"]
        == "Correct"
    ).sum()
)

partial_count = int(
    (
        reviewed["Manual_Label"]
        == "Partially Correct"
    ).sum()
)

incorrect_count = int(
    (
        reviewed["Manual_Label"]
        == "Incorrect"
    ).sum()
)

strict_precision = overall_metrics.loc[
    0,
    "Strict_Precision_Percent",
]

relevance_rate = overall_metrics.loc[
    0,
    "Relevant_or_Partial_Percent",
]

weighted_precision = overall_metrics.loc[
    0,
    "Weighted_Precision_Percent",
]

print("=" * 72)
print("CLASSIFIER EVALUATION COMPLETE")
print("=" * 72)

print(f"Input workbook:\n{INPUT_FILE}")
print()

print(f"Evaluation workbook:\n{OUTPUT_EXCEL}")
print()

print(f"Metrics CSV:\n{OUTPUT_CSV}")
print()

print(
    f"Total sampled records: {len(reviews)}"
)

print(
    f"Valid reviewed records: {len(reviewed)}"
)

print(
    f"Unreviewed records: {len(unreviewed)}"
)

print(
    f"Correct predictions: {correct_count}"
)

print(
    f"Partially correct predictions: {partial_count}"
)

print(
    f"Incorrect predictions: {incorrect_count}"
)

print()

print(
    "Strict precision: "
    f"{strict_precision:.2f}%"
)

print(
    "Relevant or partially relevant rate: "
    f"{relevance_rate:.2f}%"
)

print(
    "Weighted precision: "
    f"{weighted_precision:.2f}%"
)

print()

print(
    "Next step: review TopicMetrics, KeywordMetrics, "
    "IncorrectPredictions, and Recommendations in "
    "classifier_evaluation.xlsx."
)