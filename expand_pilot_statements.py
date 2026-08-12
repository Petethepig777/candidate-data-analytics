from pathlib import Path
from datetime import date

from openpyxl import load_workbook


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database_pilot.xlsx"
)

OUTPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database_pilot_expanded.xlsx"
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def get_headers(worksheet):
    """
    Return worksheet headers from the first row.
    """
    return [
        worksheet.cell(
            row=1,
            column=column_number,
        ).value
        for column_number in range(
            1,
            worksheet.max_column + 1,
        )
    ]


def get_column_number(
    worksheet,
    column_name,
):
    """
    Return the Excel column number for a named header.
    """
    headers = get_headers(worksheet)

    if column_name not in headers:
        raise ValueError(
            f"Column '{column_name}' was not found "
            f"in worksheet '{worksheet.title}'."
        )

    return headers.index(column_name) + 1


def get_existing_statement_ids(worksheet):
    """
    Return all existing nonblank Statement_ID values.

    This ignores formatted blank rows.
    """
    statement_id_column = get_column_number(
        worksheet,
        "Statement_ID",
    )

    statement_ids = set()

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row_number,
            column=statement_id_column,
        ).value

        if value is None:
            continue

        value_text = str(value).strip()

        if value_text:
            statement_ids.add(value_text)

    return statement_ids


def get_last_data_row(
    worksheet,
    key_column_name,
):
    """
    Find the last row containing an actual value in the
    selected key column.

    This ignores rows that contain only formatting,
    dropdowns, borders, or other Excel metadata.
    """
    key_column_number = get_column_number(
        worksheet,
        key_column_name,
    )

    last_data_row = 1

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row_number,
            column=key_column_number,
        ).value

        if value is None:
            continue

        if str(value).strip():
            last_data_row = row_number

    return last_data_row


def write_dict_to_row(
    worksheet,
    row_number,
    row_dict,
):
    """
    Write one dictionary into a specific worksheet row
    using the worksheet's header names.
    """
    headers = get_headers(worksheet)

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=row_number,
            column=column_number,
            value=row_dict.get(header, ""),
        )


def count_data_rows(
    worksheet,
    key_column_name,
):
    """
    Count actual records using a nonblank primary-key column.
    """
    key_column_number = get_column_number(
        worksheet,
        key_column_name,
    )

    count = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row_number,
            column=key_column_number,
        ).value

        if value is None:
            continue

        if str(value).strip():
            count += 1

    return count


# ============================================================
# LOAD WORKBOOK
# ============================================================

if not INPUT_FILE.exists():
    raise FileNotFoundError(
        f"Input workbook was not found:\n{INPUT_FILE}"
    )

workbook = load_workbook(INPUT_FILE)

if "CandidateStatements" not in workbook.sheetnames:
    raise ValueError(
        "CandidateStatements worksheet was not found."
    )

statements_sheet = workbook[
    "CandidateStatements"
]

existing_statement_ids = (
    get_existing_statement_ids(
        statements_sheet
    )
)


# ============================================================
# EXPANDED PILOT STATEMENTS
# ============================================================

new_statement_rows = [
    # ========================================================
    # KARLA HULT
    # ========================================================
    {
        "Statement_ID": "STMT000006",
        "Candidate_ID": "MN_SD50_001",
        "Statement_Type": "General Philosophy",
        "Statement_Text": (
            "I believe effective leadership begins by listening carefully, "
            "understanding people's experiences and then moving forward with "
            "practical and compassionate solutions. Government should protect "
            "democracy and fundamental rights, fully support public education, "
            "provide affordable and trustworthy healthcare, defend reproductive "
            "and LGBTQ+ freedom and remain accountable to the communities it serves."
        ),
        "Primary_Source_ID": "SRC000001",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign materials."
        ),
    },
    {
        "Statement_ID": "STMT000007",
        "Candidate_ID": "MN_SD50_001",
        "Statement_Type": "Goals If Elected",
        "Statement_Text": (
            "If elected, I will fight to fully fund public schools, reduce "
            "class sizes, expand school-based mental-health support, improve "
            "teacher pay and strengthen educator pensions. I will defend "
            "reproductive freedom, fertility treatment and LGBTQ+ rights, "
            "protect Medicaid and other essential healthcare coverage, lower "
            "medical costs and improve support for older adults, caregivers "
            "and families affected by Alzheimer's disease and dementia."
        ),
        "Primary_Source_ID": "SRC000001",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },
    {
        "Statement_ID": "STMT000008",
        "Candidate_ID": "MN_SD50_001",
        "Statement_Type": "Areas of Concentration",
        "Statement_Text": (
            "My principal areas of concentration will be public education, "
            "teacher support, healthcare affordability, mental-health access, "
            "Alzheimer's and dementia caregiving, reproductive freedom, "
            "LGBTQ+ equality, civil rights, voting rights and democratic accountability."
        ),
        "Primary_Source_ID": "SRC000001",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },

    # ========================================================
    # AMAL IBRAHIM
    # ========================================================
    {
        "Statement_ID": "STMT000009",
        "Candidate_ID": "MN_SD50_002",
        "Statement_Type": "General Philosophy",
        "Statement_Text": (
            "I believe every Minnesotan deserves equity, opportunity and "
            "stability regardless of immigration status, race, income or "
            "background. Government should make healthcare, housing, education, "
            "childcare and everyday necessities affordable while protecting "
            "civil liberties, workers and immigrant families."
        ),
        "Primary_Source_ID": "SRC000002",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign materials."
        ),
    },
    {
        "Statement_ID": "STMT000010",
        "Candidate_ID": "MN_SD50_002",
        "Statement_Type": "Goals If Elected",
        "Statement_Text": (
            "If elected, I will work to make housing, groceries, healthcare "
            "and childcare more affordable, support local small businesses, "
            "advocate for fair wages and expand workforce-development programs. "
            "I will fully fund public schools, improve teacher retention, "
            "expand early-childhood and technical education, defend immigrant "
            "rights and due process, strengthen tenant protections and promote "
            "clean energy and community-centered public safety."
        ),
        "Primary_Source_ID": "SRC000002",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },
    {
        "Statement_ID": "STMT000011",
        "Candidate_ID": "MN_SD50_002",
        "Statement_Type": "Areas of Concentration",
        "Statement_Text": (
            "My principal areas of concentration will be economic development, "
            "healthcare, education, affordable housing, immigrant rights, "
            "civil liberties, environmental protection, workers' rights, "
            "childcare and community safety."
        ),
        "Primary_Source_ID": "SRC000002",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },

    # ========================================================
    # NELLY KORMAN
    # ========================================================
    {
        "Statement_ID": "STMT000012",
        "Candidate_ID": "MN_SD50_003",
        "Statement_Type": "General Philosophy",
        "Statement_Text": (
            "I believe government should use a human-centered approach that "
            "treats every person with dignity and respect and puts families "
            "and communities at the center of public decision-making. "
            "Government should make housing, healthcare and education affordable, "
            "protect civil and reproductive rights, support safe communities "
            "and preserve Minnesota's environment."
        ),
        "Primary_Source_ID": "SRC000003",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign materials."
        ),
    },
    {
        "Statement_ID": "STMT000013",
        "Candidate_ID": "MN_SD50_003",
        "Statement_Type": "Goals If Elected",
        "Statement_Text": (
            "If elected, I will work to make housing, healthcare, energy and "
            "education more affordable, support workforce development and fair "
            "salaries and fully fund public schools. I will reduce special-education "
            "and multilingual-education cross-subsidies, improve teacher recruitment "
            "and student mental-health services, protect reproductive healthcare, "
            "prevent gun violence, promote safe firearm storage, protect clean water "
            "and establish responsible regulations for energy-intensive data centers."
        ),
        "Primary_Source_ID": "SRC000003",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },
    {
        "Statement_ID": "STMT000014",
        "Candidate_ID": "MN_SD50_003",
        "Statement_Type": "Areas of Concentration",
        "Statement_Text": (
            "My principal areas of concentration will be affordability, "
            "public education, civil rights, healthcare, public safety, "
            "economic opportunity, environmental protection, teacher retention, "
            "student mental health, reproductive healthcare and responsible "
            "regulation of data centers."
        ),
        "Primary_Source_ID": "SRC000003",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },

    # ========================================================
    # JOHN MCCLELLAN
    # ========================================================
    {
        "Statement_ID": "STMT000015",
        "Candidate_ID": "MN_SD50_004",
        "Statement_Type": "General Philosophy",
        "Statement_Text": (
            "I believe leadership should put people over politics and be rooted "
            "in dignity, transparency, accountability and listening. Government "
            "should give every person a fair opportunity through strong schools, "
            "affordable healthcare, safe communities, consumer protections, "
            "economic opportunity and equal justice."
        ),
        "Primary_Source_ID": "SRC000004",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign materials."
        ),
    },
    {
        "Statement_ID": "STMT000016",
        "Candidate_ID": "MN_SD50_004",
        "Statement_Type": "Goals If Elected",
        "Statement_Text": (
            "If elected, I will increase support for public schools, teacher pay, "
            "career and technical education and school mental-health services. "
            "I will expand local clinics and crisis-response systems, support "
            "first responders and victim services, strengthen renter, worker, "
            "consumer and disability protections, reform the justice system, "
            "fight junk fees and price gouging, improve veterans' services, "
            "increase affordable housing and establish safeguards for artificial intelligence."
        ),
        "Primary_Source_ID": "SRC000004",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },
    {
        "Statement_ID": "STMT000017",
        "Candidate_ID": "MN_SD50_004",
        "Statement_Type": "Areas of Concentration",
        "Statement_Text": (
            "My principal areas of concentration will be public education, "
            "mental healthcare, public safety, justice reform, civil and "
            "disability rights, consumer protection, veterans' services, "
            "affordability, housing, clean energy, artificial-intelligence "
            "regulation and government accountability."
        ),
        "Primary_Source_ID": "SRC000004",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on official campaign priorities."
        ),
    },

    # ========================================================
    # ERIN MAYE QUADE
    # ========================================================
    {
        "Statement_ID": "STMT000018",
        "Candidate_ID": "MN_SD56_001",
        "Statement_Type": "General Philosophy",
        "Statement_Text": (
            "I believe government should protect people's rights and freedoms, "
            "support children and families, preserve essential services and "
            "restore public trust through transparent and accountable leadership. "
            "My priorities include affordability, reproductive and LGBTQ+ rights, "
            "public education, healthcare, democracy, disability justice and "
            "regulation of harmful uses of artificial intelligence."
        ),
        "Primary_Source_ID": "SRC000005",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on campaign and official sources."
        ),
    },
    {
        "Statement_ID": "STMT000019",
        "Candidate_ID": "MN_SD56_001",
        "Statement_Type": "Goals If Elected",
        "Statement_Text": (
            "If reelected, I will work to make housing, childcare, healthcare "
            "and food more affordable, defend SNAP, WIC and Medicaid, strengthen "
            "election integrity and protect voting rights. I will regulate harmful "
            "uses of artificial intelligence, including election deepfakes and "
            "nonconsensual sexual images, protect reproductive freedom and LGBTQ+ "
            "rights, fully support public schools, improve literacy and reduce childhood hunger."
        ),
        "Primary_Source_ID": "SRC000005",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on campaign and official sources."
        ),
    },
    {
        "Statement_ID": "STMT000020",
        "Candidate_ID": "MN_SD56_001",
        "Statement_Type": "Areas of Concentration",
        "Statement_Text": (
            "My principal areas of concentration will be household affordability, "
            "public education and literacy, childhood hunger, childcare, "
            "reproductive freedom, LGBTQ+ rights, disability justice, healthcare, "
            "election integrity, gun-violence prevention, corporate-fraud prevention "
            "and regulation of harmful artificial-intelligence practices."
        ),
        "Primary_Source_ID": "SRC000005",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": (
            "Copy-ready synthesis based on campaign and official sources."
        ),
    },
]


# ============================================================
# ADD NEW STATEMENTS
# ============================================================

last_existing_data_row = get_last_data_row(
    statements_sheet,
    "Statement_ID",
)

next_data_row = last_existing_data_row + 1

added_count = 0
skipped_count = 0

for statement_row in new_statement_rows:
    statement_id = statement_row[
        "Statement_ID"
    ]

    if statement_id in existing_statement_ids:
        skipped_count += 1
        continue

    write_dict_to_row(
        statements_sheet,
        next_data_row,
        statement_row,
    )

    existing_statement_ids.add(
        statement_id
    )

    next_data_row += 1
    added_count += 1


# ============================================================
# UPDATE EXCEL TABLE RANGE
# ============================================================

final_data_row = get_last_data_row(
    statements_sheet,
    "Statement_ID",
)

table_name = "CandidateStatementsTable"

if table_name in statements_sheet.tables:
    table = statements_sheet.tables[
        table_name
    ]

    end_column_letter = statements_sheet.cell(
        row=1,
        column=statements_sheet.max_column,
    ).column_letter

    table.ref = (
        f"A1:{end_column_letter}"
        f"{final_data_row}"
    )


# ============================================================
# COUNT ACTUAL STATEMENT RECORDS
# ============================================================

total_statement_count = count_data_rows(
    statements_sheet,
    "Statement_ID",
)


# ============================================================
# SAVE OUTPUT
# ============================================================

OUTPUT_FILE.parent.mkdir(
    parents=True,
    exist_ok=True,
)

workbook.save(OUTPUT_FILE)


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("PILOT STATEMENT DATASET EXPANDED")
print("=" * 72)

print(f"Input workbook:\n{INPUT_FILE}")
print()

print(f"Expanded workbook:\n{OUTPUT_FILE}")
print()

print(
    f"New statements added: {added_count}"
)

print(
    f"Existing statements skipped: {skipped_count}"
)

print(
    "Total statements in expanded workbook: "
    f"{total_statement_count}"
)

print()

print(
    "Next step: update the policy classifier "
    "to read the expanded workbook."
)