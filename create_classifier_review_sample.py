from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "policy_classification_results_expanded.xlsx"
)

OUTPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "classifier_manual_review.xlsx"
)


# ============================================================
# CONFIGURATION
# ============================================================

RANDOM_SEED = 42

SAMPLES_PER_CONFIDENCE = {
    "High": 10,
    "Medium": 15,
    "Low": 20,
    "Unknown": 5,
}


# ============================================================
# LOAD CLASSIFICATION RESULTS
# ============================================================

if not INPUT_FILE.exists():
    raise FileNotFoundError(
        f"Classification workbook was not found:\n{INPUT_FILE}"
    )

classifications = pd.read_excel(
    INPUT_FILE,
    sheet_name="PolicyClassifications",
)

if classifications.empty:
    raise ValueError(
        "PolicyClassifications does not contain any records."
    )

required_columns = [
    "Policy_Record_ID",
    "Candidate_ID",
    "Ballot_Name",
    "Statement_ID",
    "Statement_Type",
    "Policy_Topic",
    "Position_Summary",
    "Matched_Keywords",
    "Keyword_Match_Count",
    "Classification_Score",
    "Classification_Confidence",
]

missing_columns = [
    column
    for column in required_columns
    if column not in classifications.columns
]

if missing_columns:
    raise ValueError(
        "Missing required classification columns: "
        + ", ".join(missing_columns)
    )


# ============================================================
# CREATE STRATIFIED REVIEW SAMPLE
# ============================================================

sample_frames = []

for confidence_level, requested_count in (
    SAMPLES_PER_CONFIDENCE.items()
):
    confidence_rows = classifications[
        classifications[
            "Classification_Confidence"
        ]
        == confidence_level
    ]

    if confidence_rows.empty:
        continue

    sample_count = min(
        requested_count,
        len(confidence_rows),
    )

    sampled_rows = confidence_rows.sample(
        n=sample_count,
        random_state=RANDOM_SEED,
    )

    sample_frames.append(sampled_rows)

if sample_frames:
    review_sample = pd.concat(
        sample_frames,
        ignore_index=True,
    )

else:
    review_sample = classifications.sample(
        n=min(40, len(classifications)),
        random_state=RANDOM_SEED,
    ).copy()


# ============================================================
# ENSURE POLICY-TOPIC COVERAGE
# ============================================================

sampled_topics = set(
    review_sample["Policy_Topic"]
)

all_topics = set(
    classifications["Policy_Topic"]
)

missing_topics = sorted(
    all_topics - sampled_topics
)

topic_additions = []

for topic in missing_topics:
    topic_rows = classifications[
        classifications["Policy_Topic"] == topic
    ]

    if topic_rows.empty:
        continue

    topic_additions.append(
        topic_rows.sample(
            n=1,
            random_state=RANDOM_SEED,
        )
    )

if topic_additions:
    review_sample = pd.concat(
        [
            review_sample,
            *topic_additions,
        ],
        ignore_index=True,
    )


# ============================================================
# REMOVE DUPLICATES AND SORT
# ============================================================

review_sample = (
    review_sample
    .drop_duplicates(
        subset=["Policy_Record_ID"]
    )
    .sort_values(
        by=[
            "Classification_Confidence",
            "Policy_Topic",
            "Ballot_Name",
        ],
        ascending=[
            True,
            True,
            True,
        ],
    )
    .reset_index(drop=True)
)


# ============================================================
# ADD MANUAL-REVIEW COLUMNS
# ============================================================

review_columns = [
    "Review_ID",
    "Policy_Record_ID",
    "Candidate_ID",
    "Ballot_Name",
    "Statement_ID",
    "Statement_Type",
    "Policy_Topic",
    "Position_Summary",
    "Matched_Keywords",
    "Keyword_Match_Count",
    "Classification_Score",
    "Classification_Confidence",
    "Manual_Label",
    "Reviewer_Confidence",
    "Incorrect_Keyword",
    "Suggested_Correct_Topic",
    "Reviewer_Notes",
]

review_output = review_sample[
    [
        "Policy_Record_ID",
        "Candidate_ID",
        "Ballot_Name",
        "Statement_ID",
        "Statement_Type",
        "Policy_Topic",
        "Position_Summary",
        "Matched_Keywords",
        "Keyword_Match_Count",
        "Classification_Score",
        "Classification_Confidence",
    ]
].copy()

review_output.insert(
    0,
    "Review_ID",
    [
        f"REV{number:04d}"
        for number in range(
            1,
            len(review_output) + 1,
        )
    ],
)

review_output["Manual_Label"] = ""
review_output["Reviewer_Confidence"] = ""
review_output["Incorrect_Keyword"] = ""
review_output["Suggested_Correct_Topic"] = ""
review_output["Reviewer_Notes"] = ""

review_output = review_output[
    review_columns
]


# ============================================================
# CREATE INSTRUCTIONS
# ============================================================

instructions = pd.DataFrame(
    [
        {
            "Field": "Manual_Label",
            "Instructions": (
                "Choose Correct when the candidate statement clearly "
                "relates to the predicted policy topic. Choose Partially "
                "Correct when the topic is relevant but weak or secondary. "
                "Choose Incorrect when the statement does not meaningfully "
                "address the predicted topic."
            ),
        },
        {
            "Field": "Reviewer_Confidence",
            "Instructions": (
                "Choose High, Medium, or Low based on how certain you are "
                "about your manual judgment."
            ),
        },
        {
            "Field": "Incorrect_Keyword",
            "Instructions": (
                "When the prediction is Incorrect or Partially Correct, "
                "enter the keyword or phrase that appears to have caused "
                "the weak classification."
            ),
        },
        {
            "Field": "Suggested_Correct_Topic",
            "Instructions": (
                "Enter a better policy topic when one is clear. Leave blank "
                "when the statement should simply not receive a topic."
            ),
        },
        {
            "Field": "Reviewer_Notes",
            "Instructions": (
                "Briefly explain ambiguous cases or why the classification "
                "should be changed."
            ),
        },
    ]
)


# ============================================================
# SAVE REVIEW WORKBOOK
# ============================================================

OUTPUT_FILE.parent.mkdir(
    parents=True,
    exist_ok=True,
)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl",
) as writer:
    review_output.to_excel(
        writer,
        sheet_name="ManualReview",
        index=False,
    )

    instructions.to_excel(
        writer,
        sheet_name="Instructions",
        index=False,
    )

    workbook = writer.book
    review_sheet = workbook["ManualReview"]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in review_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    review_sheet.freeze_panes = "A2"
    review_sheet.auto_filter.ref = (
        review_sheet.dimensions
    )

    label_validation = DataValidation(
        type="list",
        formula1=(
            '"Correct,Partially Correct,Incorrect"'
        ),
        allow_blank=True,
    )

    reviewer_confidence_validation = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
    )

    review_sheet.add_data_validation(
        label_validation
    )

    review_sheet.add_data_validation(
        reviewer_confidence_validation
    )

    label_validation.add(
        f"M2:M{review_sheet.max_row}"
    )

    reviewer_confidence_validation.add(
        f"N2:N{review_sheet.max_row}"
    )

    column_widths = {
        "A": 12,
        "B": 18,
        "C": 16,
        "D": 22,
        "E": 16,
        "F": 24,
        "G": 25,
        "H": 90,
        "I": 45,
        "J": 20,
        "K": 22,
        "L": 25,
        "M": 22,
        "N": 22,
        "O": 35,
        "P": 30,
        "Q": 55,
    }

    for column, width in column_widths.items():
        review_sheet.column_dimensions[
            column
        ].width = width

    for row in review_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    instruction_sheet = workbook[
        "Instructions"
    ]

    instruction_sheet.freeze_panes = "A2"

    instruction_sheet.column_dimensions[
        "A"
    ].width = 30

    instruction_sheet.column_dimensions[
        "B"
    ].width = 100

    for cell in instruction_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in instruction_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("CLASSIFIER MANUAL-REVIEW SAMPLE CREATED")
print("=" * 72)

print(f"Input file:\n{INPUT_FILE}")
print()

print(f"Review workbook:\n{OUTPUT_FILE}")
print()

print(
    f"Total classification records available: "
    f"{len(classifications)}"
)

print(
    f"Records selected for manual review: "
    f"{len(review_output)}"
)

print(
    f"Policy topics represented: "
    f"{review_output['Policy_Topic'].nunique()}"
)

print()

print(
    "Next step: open classifier_manual_review.xlsx, "
    "read the Instructions sheet, and complete the "
    "Manual_Label column in the ManualReview sheet."
)