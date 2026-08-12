from pathlib import Path
from datetime import date

from openpyxl import load_workbook


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = Path.home() / "Desktop" / "POLITICAL_CANDIDATE_ANALYTICS"

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database.xlsx"
)

OUTPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database_pilot.xlsx"
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clear_blank_template_row(worksheet):
    """
    Delete the initial blank row that was included in each Excel table.
    This keeps the pilot workbook clean before data is inserted.
    """
    if worksheet.max_row >= 2:
        values = [
            worksheet.cell(row=2, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]

        if all(value in (None, "") for value in values):
            worksheet.delete_rows(2, 1)


def append_dict_row(worksheet, row_dict):
    """
    Append one row using the worksheet header names.

    Missing dictionary keys are inserted as blank cells.
    """
    headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, worksheet.max_column + 1)
    ]

    worksheet.append(
        [row_dict.get(header, "") for header in headers]
    )


# ============================================================
# LOAD WORKBOOK
# ============================================================

if not INPUT_FILE.exists():
    raise FileNotFoundError(
        f"Input workbook was not found:\n{INPUT_FILE}"
    )

workbook = load_workbook(INPUT_FILE)

required_sheets = [
    "Candidates",
    "Sources",
    "CandidateFacts",
    "CandidateStatements",
    "PolicyTopics",
    "SocialMedia",
]

for sheet_name in required_sheets:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Required worksheet is missing: {sheet_name}"
        )

candidates_sheet = workbook["Candidates"]
sources_sheet = workbook["Sources"]
facts_sheet = workbook["CandidateFacts"]
statements_sheet = workbook["CandidateStatements"]
social_sheet = workbook["SocialMedia"]

for worksheet in [
    candidates_sheet,
    sources_sheet,
    facts_sheet,
    statements_sheet,
    social_sheet,
]:
    clear_blank_template_row(worksheet)


# ============================================================
# PILOT CANDIDATES
# ============================================================

candidate_rows = [
    {
        "Candidate_ID": "MN_SD50_001",
        "Ballot_Name": "Karla Hult",
        "Full_Name": "Karla Hult",
        "First_Name": "Karla",
        "Middle_Name": "",
        "Last_Name": "Hult",
        "Suffix": "",
        "Party": "Democratic-Farmer-Labor",
        "Office": "State Senator",
        "State": "Minnesota",
        "District": "50",
        "Incumbent": "No",
        "Candidate_Status": "Primary Candidate",
        "Filing_Date": date(2026, 5, 19),
        "DOB": "",
        "Birth_Year": "",
        "Birthplace": "",
        "Residence_City": "Edina",
        "Residence_State": "Minnesota",
        "Campaign_Website": "https://www.karlahultforsenate.com/",
        "Campaign_Email": "Karla@KarlaHultforSenate.com",
        "Campaign_Phone": "612-437-7721",
        "Campaign_Address": "P.O. Box 24544, Edina, MN 55424",
        "Research_Status": "Complete",
        "Researcher": "Youwei Chen",
        "Date_First_Researched": date(2026, 8, 4),
        "Date_Last_Updated": date(2026, 8, 5),
        "Candidate_Notes": (
            "Exact date of birth was not found. "
            "First-time candidate with a journalism and Alzheimer’s advocacy background."
        ),
    },
    {
        "Candidate_ID": "MN_SD50_002",
        "Ballot_Name": "Amal Ibrahim",
        "Full_Name": "Amal Ibrahim",
        "First_Name": "Amal",
        "Middle_Name": "",
        "Last_Name": "Ibrahim",
        "Suffix": "",
        "Party": "Democratic-Farmer-Labor",
        "Office": "State Senator",
        "State": "Minnesota",
        "District": "50",
        "Incumbent": "No",
        "Candidate_Status": "Primary Candidate",
        "Filing_Date": date(2026, 6, 2),
        "DOB": "",
        "Birth_Year": "",
        "Birthplace": "Somalia",
        "Residence_City": "Bloomington",
        "Residence_State": "Minnesota",
        "Campaign_Website": "https://www.amalformn.com/",
        "Campaign_Email": "amalformn@gmail.com",
        "Campaign_Phone": "",
        "Campaign_Address": "",
        "Research_Status": "Complete",
        "Researcher": "Youwei Chen",
        "Date_First_Researched": date(2026, 8, 4),
        "Date_Last_Updated": date(2026, 8, 5),
        "Candidate_Notes": (
            "Exact date of birth, age, campaign address, and current campaign phone were not found."
        ),
    },
    {
        "Candidate_ID": "MN_SD50_003",
        "Ballot_Name": "Nelly Korman",
        "Full_Name": "Nelly Korman",
        "First_Name": "Nelly",
        "Middle_Name": "",
        "Last_Name": "Korman",
        "Suffix": "",
        "Party": "Democratic-Farmer-Labor",
        "Office": "State Senator",
        "State": "Minnesota",
        "District": "50",
        "Incumbent": "No",
        "Candidate_Status": "Primary Candidate",
        "Filing_Date": date(2026, 5, 19),
        "DOB": date(1974, 7, 1),
        "Birth_Year": 1974,
        "Birthplace": "Colombia",
        "Residence_City": "Bloomington",
        "Residence_State": "Minnesota",
        "Campaign_Website": "https://www.kormanforminnesota.com/",
        "Campaign_Email": "kormanforsenate@gmail.com",
        "Campaign_Phone": "612-470-7384",
        "Campaign_Address": "P.O. Box 20284, Bloomington, MN 55420",
        "Research_Status": "Complete",
        "Researcher": "Youwei Chen",
        "Date_First_Researched": date(2026, 8, 4),
        "Date_Last_Updated": date(2026, 8, 5),
        "Candidate_Notes": (
            "DOB was derived from first-party campaign material celebrating her "
            "52nd birthday on July 1, 2026."
        ),
    },
    {
        "Candidate_ID": "MN_SD50_004",
        "Ballot_Name": "John McClellan",
        "Full_Name": "John McClellan",
        "First_Name": "John",
        "Middle_Name": "",
        "Last_Name": "McClellan",
        "Suffix": "",
        "Party": "Democratic-Farmer-Labor",
        "Office": "State Senator",
        "State": "Minnesota",
        "District": "50",
        "Incumbent": "No",
        "Candidate_Status": "Primary Candidate",
        "Filing_Date": date(2026, 5, 19),
        "DOB": "",
        "Birth_Year": "",
        "Birthplace": "",
        "Residence_City": "Bloomington",
        "Residence_State": "Minnesota",
        "Campaign_Website": "https://mcclellanforsd50.com/",
        "Campaign_Email": "mcclellanforsd50@gmail.com",
        "Campaign_Phone": "952-297-0159",
        "Campaign_Address": "3900 West 102nd Street, Bloomington, MN 55437",
        "Research_Status": "Complete",
        "Researcher": "Youwei Chen",
        "Date_First_Researched": date(2026, 8, 4),
        "Date_Last_Updated": date(2026, 8, 5),
        "Candidate_Notes": (
            "Some sources use Johnathon McClellan, but official filing records use John McClellan."
        ),
    },
    {
        "Candidate_ID": "MN_SD56_001",
        "Ballot_Name": "Erin Maye Quade",
        "Full_Name": "Erin K. Maye Quade",
        "First_Name": "Erin",
        "Middle_Name": "K.",
        "Last_Name": "Maye Quade",
        "Suffix": "",
        "Party": "Democratic-Farmer-Labor",
        "Office": "State Senator",
        "State": "Minnesota",
        "District": "56",
        "Incumbent": "Yes",
        "Candidate_Status": "Primary Candidate",
        "Filing_Date": "",
        "DOB": date(1986, 3, 12),
        "Birth_Year": 1986,
        "Birthplace": "",
        "Residence_City": "Apple Valley",
        "Residence_State": "Minnesota",
        "Campaign_Website": "https://erinmayequade.com/",
        "Campaign_Email": "",
        "Campaign_Phone": "",
        "Campaign_Address": "",
        "Research_Status": "Complete",
        "Researcher": "Youwei Chen",
        "Date_First_Researched": date(2026, 8, 4),
        "Date_Last_Updated": date(2026, 8, 5),
        "Candidate_Notes": (
            "Official Minnesota Legislative Reference Library lists DOB as 1986-03-12."
        ),
    },
]

for row in candidate_rows:
    append_dict_row(candidates_sheet, row)


# ============================================================
# SOURCES
# ============================================================

source_rows = [
    {
        "Source_ID": "SRC000001",
        "Candidate_ID": "MN_SD50_001",
        "Source_Title": "Karla Hult for Senate",
        "Source_Type": "Campaign Website",
        "Publisher_or_Organization": "Karla Hult for Senate",
        "Domain": "karlahultforsenate.com",
        "URL": "https://www.karlahultforsenate.com/",
        "Publication_Date": "",
        "Access_Date": date(2026, 8, 5),
        "Author": "",
        "Reliability_Level": "High",
        "Primary_or_Secondary": "Primary",
        "Link_Status": "Active",
        "Verified": "Yes",
        "Archive_URL": "",
        "Source_Notes": "Official campaign website.",
    },
    {
        "Source_ID": "SRC000002",
        "Candidate_ID": "MN_SD50_002",
        "Source_Title": "Amal Ibrahim for Minnesota",
        "Source_Type": "Campaign Website",
        "Publisher_or_Organization": "Amal for MN",
        "Domain": "amalformn.com",
        "URL": "https://www.amalformn.com/",
        "Publication_Date": "",
        "Access_Date": date(2026, 8, 5),
        "Author": "",
        "Reliability_Level": "High",
        "Primary_or_Secondary": "Primary",
        "Link_Status": "Active",
        "Verified": "Yes",
        "Archive_URL": "",
        "Source_Notes": "Official campaign website.",
    },
    {
        "Source_ID": "SRC000003",
        "Candidate_ID": "MN_SD50_003",
        "Source_Title": "Nelly Korman for Minnesota",
        "Source_Type": "Campaign Website",
        "Publisher_or_Organization": "Nelly Korman for Minnesota",
        "Domain": "kormanforminnesota.com",
        "URL": "https://www.kormanforminnesota.com/",
        "Publication_Date": "",
        "Access_Date": date(2026, 8, 5),
        "Author": "",
        "Reliability_Level": "High",
        "Primary_or_Secondary": "Primary",
        "Link_Status": "Active",
        "Verified": "Yes",
        "Archive_URL": "",
        "Source_Notes": "Official campaign website.",
    },
    {
        "Source_ID": "SRC000004",
        "Candidate_ID": "MN_SD50_004",
        "Source_Title": "McClellan for Senate District 50",
        "Source_Type": "Campaign Website",
        "Publisher_or_Organization": "McClellan for SD50",
        "Domain": "mcclellanforsd50.com",
        "URL": "https://mcclellanforsd50.com/",
        "Publication_Date": "",
        "Access_Date": date(2026, 8, 5),
        "Author": "",
        "Reliability_Level": "High",
        "Primary_or_Secondary": "Primary",
        "Link_Status": "Active",
        "Verified": "Yes",
        "Archive_URL": "",
        "Source_Notes": "Official campaign website.",
    },
    {
        "Source_ID": "SRC000005",
        "Candidate_ID": "MN_SD56_001",
        "Source_Title": "Erin Maye Quade Legislative Biography",
        "Source_Type": "Government",
        "Publisher_or_Organization": "Minnesota Legislative Reference Library",
        "Domain": "lrl.mn.gov",
        "URL": "https://www.lrl.mn.gov/legdb/fulldetail?id=15473",
        "Publication_Date": "",
        "Access_Date": date(2026, 8, 5),
        "Author": "",
        "Reliability_Level": "High",
        "Primary_or_Secondary": "Primary",
        "Link_Status": "Active",
        "Verified": "Yes",
        "Archive_URL": "",
        "Source_Notes": "Official legislative biography and DOB source.",
    },
]

for row in source_rows:
    append_dict_row(sources_sheet, row)


# ============================================================
# CANDIDATE FACTS
# ============================================================

fact_rows = [
    {
        "Fact_ID": "FACT000001",
        "Candidate_ID": "MN_SD50_001",
        "Fact_Category": "Professional Experience",
        "Fact_Field": "Occupation",
        "Fact_Value": "Journalist, educator, podcast host, and Alzheimer’s-family advocate",
        "Source_ID": "SRC000001",
        "Verification_Status": "Verified",
        "Confidence_Level": "High",
        "Is_Inferred": "No",
        "Conflicting_Source_Flag": "No",
        "Missing_Value_Flag": "No",
        "Date_Verified": date(2026, 8, 5),
        "Fact_Notes": "",
    },
    {
        "Fact_ID": "FACT000002",
        "Candidate_ID": "MN_SD50_002",
        "Fact_Category": "Biography",
        "Fact_Field": "Immigration History",
        "Fact_Value": "Immigrated from Somalia to the United States in 1999",
        "Source_ID": "SRC000002",
        "Verification_Status": "Verified",
        "Confidence_Level": "High",
        "Is_Inferred": "No",
        "Conflicting_Source_Flag": "No",
        "Missing_Value_Flag": "No",
        "Date_Verified": date(2026, 8, 5),
        "Fact_Notes": "",
    },
    {
        "Fact_ID": "FACT000003",
        "Candidate_ID": "MN_SD50_003",
        "Fact_Category": "Identity",
        "Fact_Field": "Date of Birth",
        "Fact_Value": "July 1, 1974",
        "Source_ID": "SRC000003",
        "Verification_Status": "Partially Verified",
        "Confidence_Level": "Medium",
        "Is_Inferred": "Yes",
        "Conflicting_Source_Flag": "No",
        "Missing_Value_Flag": "No",
        "Date_Verified": date(2026, 8, 5),
        "Fact_Notes": (
            "Derived from official campaign material stating she celebrated "
            "her 52nd birthday on July 1, 2026."
        ),
    },
    {
        "Fact_ID": "FACT000004",
        "Candidate_ID": "MN_SD50_004",
        "Fact_Category": "Professional Experience",
        "Fact_Field": "Occupation",
        "Fact_Value": "Retired Minneapolis firefighter and community advocate",
        "Source_ID": "SRC000004",
        "Verification_Status": "Verified",
        "Confidence_Level": "High",
        "Is_Inferred": "No",
        "Conflicting_Source_Flag": "No",
        "Missing_Value_Flag": "No",
        "Date_Verified": date(2026, 8, 5),
        "Fact_Notes": "",
    },
    {
        "Fact_ID": "FACT000005",
        "Candidate_ID": "MN_SD56_001",
        "Fact_Category": "Identity",
        "Fact_Field": "Date of Birth",
        "Fact_Value": "March 12, 1986",
        "Source_ID": "SRC000005",
        "Verification_Status": "Verified",
        "Confidence_Level": "High",
        "Is_Inferred": "No",
        "Conflicting_Source_Flag": "No",
        "Missing_Value_Flag": "No",
        "Date_Verified": date(2026, 8, 5),
        "Fact_Notes": "Verified through an official Minnesota legislative source.",
    },
]

for row in fact_rows:
    append_dict_row(facts_sheet, row)


# ============================================================
# LONG-FORM STATEMENTS
# ============================================================

statement_rows = [
    {
        "Statement_ID": "STMT000001",
        "Candidate_ID": "MN_SD50_001",
        "Statement_Type": "Why Running",
        "Statement_Text": (
            "I am running because our democracy, fundamental rights and shared "
            "humanity are at stake, and this moment requires leaders who can "
            "listen, communicate clearly and earn public trust."
        ),
        "Primary_Source_ID": "SRC000001",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": "Copy-ready synthesis based on official campaign materials.",
    },
    {
        "Statement_ID": "STMT000002",
        "Candidate_ID": "MN_SD50_002",
        "Statement_Type": "Why Running",
        "Statement_Text": (
            "I am running because too many Minnesota families are being left "
            "behind while the costs of housing, groceries, healthcare and "
            "childcare continue to rise."
        ),
        "Primary_Source_ID": "SRC000002",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": "Copy-ready synthesis based on official campaign materials.",
    },
    {
        "Statement_ID": "STMT000003",
        "Candidate_ID": "MN_SD50_003",
        "Statement_Type": "Why Running",
        "Statement_Text": (
            "I am running because my years as a public-school teacher, mother, "
            "immigrant, school-board member and planning commissioner have shown "
            "me both the strengths of our communities and the challenges families face."
        ),
        "Primary_Source_ID": "SRC000003",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": "Copy-ready synthesis based on official campaign materials.",
    },
    {
        "Statement_ID": "STMT000004",
        "Candidate_ID": "MN_SD50_004",
        "Statement_Type": "Why Running",
        "Statement_Text": (
            "I am running because District 50 deserves leadership that puts "
            "people before politics and replaces broken promises with transparency, "
            "accountability and practical results."
        ),
        "Primary_Source_ID": "SRC000004",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": "Copy-ready synthesis based on official campaign materials.",
    },
    {
        "Statement_ID": "STMT000005",
        "Candidate_ID": "MN_SD56_001",
        "Statement_Type": "Why Running",
        "Statement_Text": (
            "I am running for reelection because Minnesotans’ rights, freedoms "
            "and essential public services face serious threats, and District 56 "
            "needs an experienced advocate who understands both the legislative "
            "process and the community."
        ),
        "Primary_Source_ID": "SRC000005",
        "Additional_Source_IDs": "",
        "Statement_Date": date(2026, 8, 5),
        "Verified": "Partially",
        "Manual_or_Generated": "AI-Assisted",
        "Review_Status": "Approved",
        "Reviewer_Notes": "Copy-ready synthesis based on campaign and official sources.",
    },
]

for row in statement_rows:
    append_dict_row(statements_sheet, row)


# ============================================================
# SOCIAL MEDIA
# ============================================================

social_rows = [
    {
        "Social_ID": "SOC000001",
        "Candidate_ID": "MN_SD50_001",
        "Platform": "Instagram",
        "Account_Name": "@karlahult",
        "Profile_URL": "https://www.instagram.com/karlahult/",
        "Official_Status": "Likely Official",
        "Account_Type": "Personal Public",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "Public account connected to campaign activity.",
    },
    {
        "Social_ID": "SOC000002",
        "Candidate_ID": "MN_SD50_001",
        "Platform": "X or Twitter",
        "Account_Name": "@karlahult",
        "Profile_URL": "https://x.com/karlahult",
        "Official_Status": "Likely Official",
        "Account_Type": "Personal Public",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "",
    },
    {
        "Social_ID": "SOC000003",
        "Candidate_ID": "MN_SD50_002",
        "Platform": "Facebook",
        "Account_Name": "Amal for MN",
        "Profile_URL": "https://www.facebook.com/AmalForMN/",
        "Official_Status": "Verified Official",
        "Account_Type": "Campaign",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "Some content remains from the 2020 campaign.",
    },
    {
        "Social_ID": "SOC000004",
        "Candidate_ID": "MN_SD50_002",
        "Platform": "Instagram",
        "Account_Name": "@amalformn",
        "Profile_URL": "https://www.instagram.com/amalformn/",
        "Official_Status": "Verified Official",
        "Account_Type": "Campaign",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "Profile may still contain older House campaign information.",
    },
    {
        "Social_ID": "SOC000005",
        "Candidate_ID": "MN_SD50_003",
        "Platform": "Instagram",
        "Account_Name": "@kormanforsenate",
        "Profile_URL": "https://www.instagram.com/kormanforsenate/",
        "Official_Status": "Verified Official",
        "Account_Type": "Campaign",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "",
    },
    {
        "Social_ID": "SOC000006",
        "Candidate_ID": "MN_SD50_004",
        "Platform": "Facebook",
        "Account_Name": "McClellan for SD50",
        "Profile_URL": "https://www.facebook.com/p/McClellan-for-SD50-61585583159248/",
        "Official_Status": "Verified Official",
        "Account_Type": "Campaign",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "",
    },
    {
        "Social_ID": "SOC000007",
        "Candidate_ID": "MN_SD56_001",
        "Platform": "Instagram",
        "Account_Name": "@erinmayequade",
        "Profile_URL": "https://www.instagram.com/erinmayequade/",
        "Official_Status": "Verified Official",
        "Account_Type": "Personal Public",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "",
    },
    {
        "Social_ID": "SOC000008",
        "Candidate_ID": "MN_SD56_001",
        "Platform": "X or Twitter",
        "Account_Name": "@ErinMayeQuade",
        "Profile_URL": "https://x.com/ErinMayeQuade",
        "Official_Status": "Verified Official",
        "Account_Type": "Personal Public",
        "Active_Status": "Active",
        "Last_Checked": date(2026, 8, 5),
        "Follower_Count": "",
        "Notes": "",
    },
]

for row in social_rows:
    append_dict_row(social_sheet, row)


# ============================================================
# UPDATE TABLE RANGES
# ============================================================

table_sheet_pairs = [
    (candidates_sheet, "CandidatesTable"),
    (sources_sheet, "SourcesTable"),
    (facts_sheet, "CandidateFactsTable"),
    (statements_sheet, "CandidateStatementsTable"),
    (social_sheet, "SocialMediaTable"),
]

for worksheet, table_name in table_sheet_pairs:
    if table_name in worksheet.tables:
        table = worksheet.tables[table_name]

        end_column_letter = worksheet.cell(
            row=1,
            column=worksheet.max_column,
        ).column_letter

        table.ref = f"A1:{end_column_letter}{worksheet.max_row}"


# ============================================================
# SAVE PILOT WORKBOOK
# ============================================================

workbook.save(OUTPUT_FILE)

print("=" * 72)
print("PILOT CANDIDATE DATASET CREATED")
print("=" * 72)
print(f"Saved to:\n{OUTPUT_FILE}")
print()
print(f"Candidates added: {len(candidate_rows)}")
print(f"Sources added: {len(source_rows)}")
print(f"Facts added: {len(fact_rows)}")
print(f"Statements added: {len(statement_rows)}")
print(f"Social-media records added: {len(social_rows)}")
print()
print("Next step: open the pilot workbook and inspect the populated sheets.")