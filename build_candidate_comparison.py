from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

DASHBOARD_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "political_analytics_dashboard_data.xlsx"
)

FEATURE_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_feature_dataset.xlsx"
)

SIMILARITY_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_similarity_analysis.xlsx"
)

OUTPUT_EXCEL = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_comparison.xlsx"
)

OUTPUT_MARKDOWN = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_comparison.md"
)


# ============================================================
# SELECT THE TWO CANDIDATES HERE
# ============================================================

CANDIDATE_A_NAME = "Amal Ibrahim"
CANDIDATE_B_NAME = "Nelly Korman"


# ============================================================
# FORMATTING
# ============================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop with a clear message when an input file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to blank text and normalize whitespace.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def safe_float(value, default=0.0):
    """
    Convert one value to float safely.
    """
    try:
        numeric_value = pd.to_numeric(
            value,
            errors="coerce",
        )

        if pd.isna(numeric_value):
            return default

        return float(numeric_value)

    except Exception:
        return default


def safe_int(value, default=0):
    """
    Convert one value to integer safely.
    """
    return int(
        round(
            safe_float(
                value,
                default,
            )
        )
    )


def find_candidate_row(
    dataframe,
    candidate_name,
    name_column="Ballot_Name",
):
    """
    Return one candidate row using a case-insensitive exact match.
    """
    if name_column not in dataframe.columns:
        raise ValueError(
            f"Column '{name_column}' was not found."
        )

    normalized_target = clean_text(
        candidate_name
    ).lower()

    normalized_names = (
        dataframe[name_column]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.lower()
    )

    matching_rows = dataframe[
        normalized_names
        == normalized_target
    ]

    if matching_rows.empty:
        available_names = sorted(
            dataframe[name_column]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        raise ValueError(
            f"Candidate '{candidate_name}' was not found.\n"
            f"Available candidates: "
            + ", ".join(available_names)
        )

    return matching_rows.iloc[0]


def split_semicolon_text(value):
    """
    Split a semicolon-separated text field.
    """
    text = clean_text(value)

    if not text:
        return []

    return [
        item.strip()
        for item in text.split(";")
        if item.strip()
    ]


def style_worksheet(worksheet):
    """
    Apply consistent styling to an output worksheet.
    """
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32

    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            55,
        )


# ============================================================
# VERIFY INPUT FILES
# ============================================================

require_file(
    DASHBOARD_FILE,
    "Dashboard dataset workbook",
)

require_file(
    FEATURE_FILE,
    "Candidate feature workbook",
)

require_file(
    SIMILARITY_FILE,
    "Candidate similarity workbook",
)


# ============================================================
# LOAD DATA
# ============================================================

candidate_overview = pd.read_excel(
    DASHBOARD_FILE,
    sheet_name="CandidateOverview",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_topic_summary = pd.read_excel(
    DASHBOARD_FILE,
    sheet_name="CandidateTopicSummary",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

policy_matrix = pd.read_excel(
    DASHBOARD_FILE,
    sheet_name="CandidatePolicyMatrix",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_features = pd.read_excel(
    FEATURE_FILE,
    sheet_name="CandidateFeatures",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_pairs = pd.read_excel(
    SIMILARITY_FILE,
    sheet_name="CandidatePairs",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# FIND SELECTED CANDIDATES
# ============================================================

candidate_a_overview = find_candidate_row(
    candidate_overview,
    CANDIDATE_A_NAME,
)

candidate_b_overview = find_candidate_row(
    candidate_overview,
    CANDIDATE_B_NAME,
)

candidate_a_features = find_candidate_row(
    candidate_features,
    CANDIDATE_A_NAME,
)

candidate_b_features = find_candidate_row(
    candidate_features,
    CANDIDATE_B_NAME,
)

candidate_a_vector = find_candidate_row(
    policy_matrix,
    CANDIDATE_A_NAME,
)

candidate_b_vector = find_candidate_row(
    policy_matrix,
    CANDIDATE_B_NAME,
)


# ============================================================
# FIND THE PAIR SIMILARITY RECORD
# ============================================================

pair_match = candidate_pairs[
    (
        (
            candidate_pairs["Candidate_A"]
            .astype(str)
            .str.strip()
            .str.lower()
            == CANDIDATE_A_NAME.lower()
        )
        & (
            candidate_pairs["Candidate_B"]
            .astype(str)
            .str.strip()
            .str.lower()
            == CANDIDATE_B_NAME.lower()
        )
    )
    |
    (
        (
            candidate_pairs["Candidate_A"]
            .astype(str)
            .str.strip()
            .str.lower()
            == CANDIDATE_B_NAME.lower()
        )
        & (
            candidate_pairs["Candidate_B"]
            .astype(str)
            .str.strip()
            .str.lower()
            == CANDIDATE_A_NAME.lower()
        )
    )
]

if pair_match.empty:
    raise ValueError(
        f"No similarity record was found for "
        f"{CANDIDATE_A_NAME} and {CANDIDATE_B_NAME}."
    )

pair_row = pair_match.iloc[0]


# ============================================================
# POLICY VECTOR COLUMNS
# ============================================================

identifier_columns = {
    "Candidate_ID",
    "Ballot_Name",
}

policy_columns = [
    column
    for column in policy_matrix.columns
    if column not in identifier_columns
]

for column in policy_columns:
    candidate_a_vector[column] = safe_float(
        candidate_a_vector[column]
    )

    candidate_b_vector[column] = safe_float(
        candidate_b_vector[column]
    )


# ============================================================
# BUILD POLICY COMPARISON TABLE
# ============================================================

policy_comparison_rows = []

for policy_topic in policy_columns:
    weight_a = safe_float(
        candidate_a_vector[policy_topic]
    )

    weight_b = safe_float(
        candidate_b_vector[policy_topic]
    )

    present_a = weight_a > 0
    present_b = weight_b > 0

    if present_a and present_b:
        comparison_status = "Shared"

    elif present_a and not present_b:
        comparison_status = (
            f"Unique to {CANDIDATE_A_NAME}"
        )

    elif present_b and not present_a:
        comparison_status = (
            f"Unique to {CANDIDATE_B_NAME}"
        )

    else:
        comparison_status = "Not emphasized"

    policy_comparison_rows.append(
        {
            "Policy_Topic": policy_topic,
            CANDIDATE_A_NAME: weight_a,
            CANDIDATE_B_NAME: weight_b,
            "Absolute_Difference": round(
                abs(weight_a - weight_b),
                2,
            ),
            "Comparison_Status": (
                comparison_status
            ),
        }
    )

policy_comparison = pd.DataFrame(
    policy_comparison_rows
)

policy_comparison.sort_values(
    by=[
        "Comparison_Status",
        "Absolute_Difference",
        "Policy_Topic",
    ],
    ascending=[
        True,
        False,
        True,
    ],
    inplace=True,
)

policy_comparison.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# GET TOP POLICY TOPICS
# ============================================================

def get_top_topics(
    candidate_name,
    number_of_topics=5,
):
    """
    Return the top policy topics for one candidate.
    """
    candidate_rows = (
        candidate_topic_summary[
            candidate_topic_summary[
                "Ballot_Name"
            ]
            .astype(str)
            .str.strip()
            .str.lower()
            == candidate_name.lower()
        ]
        .copy()
    )

    if candidate_rows.empty:
        return []

    candidate_rows[
        "Total_Keyword_Matches"
    ] = pd.to_numeric(
        candidate_rows[
            "Total_Keyword_Matches"
        ],
        errors="coerce",
    ).fillna(0)

    return (
        candidate_rows
        .sort_values(
            by="Total_Keyword_Matches",
            ascending=False,
        )
        .head(number_of_topics)[
            "Policy_Topic"
        ]
        .astype(str)
        .tolist()
    )


top_topics_a = get_top_topics(
    CANDIDATE_A_NAME
)

top_topics_b = get_top_topics(
    CANDIDATE_B_NAME
)


# ============================================================
# BUILD FEATURE COMPARISON
# ============================================================

feature_rows = [
    {
        "Metric": "Research Completeness Percent",
        CANDIDATE_A_NAME: safe_float(
            candidate_a_overview.get(
                "Research_Completeness_Percent",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_float(
            candidate_b_overview.get(
                "Research_Completeness_Percent",
                0,
            )
        ),
    },
    {
        "Metric": "Source Count",
        CANDIDATE_A_NAME: safe_int(
            candidate_a_overview.get(
                "Source_Count",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_int(
            candidate_b_overview.get(
                "Source_Count",
                0,
            )
        ),
    },
    {
        "Metric": "Verified Source Count",
        CANDIDATE_A_NAME: safe_int(
            candidate_a_overview.get(
                "Verified_Source_Count",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_int(
            candidate_b_overview.get(
                "Verified_Source_Count",
                0,
            )
        ),
    },
    {
        "Metric": "Policy Diversity",
        CANDIDATE_A_NAME: safe_int(
            candidate_a_features.get(
                "Policy_Diversity",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_int(
            candidate_b_features.get(
                "Policy_Diversity",
                0,
            )
        ),
    },
    {
        "Metric": "Total Policy Weight",
        CANDIDATE_A_NAME: safe_float(
            candidate_a_features.get(
                "Total_Policy_Weight",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_float(
            candidate_b_features.get(
                "Total_Policy_Weight",
                0,
            )
        ),
    },
    {
        "Metric": "Dominant Policy",
        CANDIDATE_A_NAME: clean_text(
            candidate_a_features.get(
                "Dominant_Policy",
                "",
            )
        ),
        CANDIDATE_B_NAME: clean_text(
            candidate_b_features.get(
                "Dominant_Policy",
                "",
            )
        ),
    },
    {
        "Metric": "Dominant Policy Percent",
        CANDIDATE_A_NAME: safe_float(
            candidate_a_features.get(
                "Dominant_Policy_Percent",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_float(
            candidate_b_features.get(
                "Dominant_Policy_Percent",
                0,
            )
        ),
    },
    {
        "Metric": "Average Policy Weight",
        CANDIDATE_A_NAME: safe_float(
            candidate_a_features.get(
                "Average_Policy_Weight",
                0,
            )
        ),
        CANDIDATE_B_NAME: safe_float(
            candidate_b_features.get(
                "Average_Policy_Weight",
                0,
            )
        ),
    },
    {
        "Metric": "Top Five Policy Topics",
        CANDIDATE_A_NAME: "; ".join(
            top_topics_a
        ),
        CANDIDATE_B_NAME: "; ".join(
            top_topics_b
        ),
    },
]

feature_comparison = pd.DataFrame(
    feature_rows
)


# ============================================================
# BUILD SIMILARITY SUMMARY
# ============================================================

shared_topics = split_semicolon_text(
    pair_row.get(
        "Shared_Policy_Topics",
        "",
    )
)

unique_to_a = split_semicolon_text(
    pair_row.get(
        "Unique_to_Candidate_A",
        "",
    )
)

unique_to_b = split_semicolon_text(
    pair_row.get(
        "Unique_to_Candidate_B",
        "",
    )
)

if (
    clean_text(
        pair_row.get(
            "Candidate_A",
            "",
        )
    ).lower()
    != CANDIDATE_A_NAME.lower()
):
    unique_to_a, unique_to_b = (
        unique_to_b,
        unique_to_a,
    )

similarity_summary = pd.DataFrame(
    [
        {
            "Metric": "Candidate A",
            "Value": CANDIDATE_A_NAME,
        },
        {
            "Metric": "Candidate B",
            "Value": CANDIDATE_B_NAME,
        },
        {
            "Metric": "Cosine Similarity Percent",
            "Value": safe_float(
                pair_row.get(
                    "Cosine_Similarity_Percent",
                    0,
                )
            ),
        },
        {
            "Metric": "Jaccard Similarity Percent",
            "Value": safe_float(
                pair_row.get(
                    "Jaccard_Similarity_Percent",
                    0,
                )
            ),
        },
        {
            "Metric": "Combined Similarity Percent",
            "Value": safe_float(
                pair_row.get(
                    "Combined_Similarity_Percent",
                    0,
                )
            ),
        },
        {
            "Metric": "Shared Topic Count",
            "Value": len(
                shared_topics
            ),
        },
        {
            "Metric": "Shared Policy Topics",
            "Value": "; ".join(
                shared_topics
            ),
        },
        {
            "Metric": (
                f"Unique Topics — "
                f"{CANDIDATE_A_NAME}"
            ),
            "Value": "; ".join(
                unique_to_a
            ),
        },
        {
            "Metric": (
                f"Unique Topics — "
                f"{CANDIDATE_B_NAME}"
            ),
            "Value": "; ".join(
                unique_to_b
            ),
        },
        {
            "Metric": "Comparison Created",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
    ]
)


# ============================================================
# BUILD COMPARISON SUMMARY
# ============================================================

combined_similarity = safe_float(
    pair_row.get(
        "Combined_Similarity_Percent",
        0,
    )
)

if combined_similarity >= 75:
    similarity_interpretation = (
        "The candidates have strongly overlapping "
        "policy profiles in the current pilot dataset."
    )

elif combined_similarity >= 50:
    similarity_interpretation = (
        "The candidates have moderately similar "
        "policy profiles with several shared priorities."
    )

else:
    similarity_interpretation = (
        "The candidates have relatively different "
        "policy profiles in the current pilot dataset."
    )

comparison_summary = pd.DataFrame(
    [
        {
            "Section": "Comparison",
            "Result": (
                f"{CANDIDATE_A_NAME} vs. "
                f"{CANDIDATE_B_NAME}"
            ),
        },
        {
            "Section": "Combined Similarity",
            "Result": (
                f"{combined_similarity:.2f}%"
            ),
        },
        {
            "Section": "Interpretation",
            "Result": (
                similarity_interpretation
            ),
        },
        {
            "Section": "Shared Topics",
            "Result": "; ".join(
                shared_topics
            ),
        },
        {
            "Section": (
                f"Unique to {CANDIDATE_A_NAME}"
            ),
            "Result": "; ".join(
                unique_to_a
            ),
        },
        {
            "Section": (
                f"Unique to {CANDIDATE_B_NAME}"
            ),
            "Result": "; ".join(
                unique_to_b
            ),
        },
        {
            "Section": (
                f"{CANDIDATE_A_NAME} Top Topics"
            ),
            "Result": "; ".join(
                top_topics_a
            ),
        },
        {
            "Section": (
                f"{CANDIDATE_B_NAME} Top Topics"
            ),
            "Result": "; ".join(
                top_topics_b
            ),
        },
        {
            "Section": "Methodological Note",
            "Result": (
                "Similarity is based on policy-topic "
                "keyword weights in the current pilot statements. "
                "It does not measure ideology, policy feasibility, "
                "or candidate quality."
            ),
        },
    ]
)


# ============================================================
# SAVE EXCEL WORKBOOK
# ============================================================

OUTPUT_EXCEL.parent.mkdir(
    parents=True,
    exist_ok=True,
)

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl",
) as writer:

    comparison_summary.to_excel(
        writer,
        sheet_name="ComparisonSummary",
        index=False,
    )

    feature_comparison.to_excel(
        writer,
        sheet_name="FeatureComparison",
        index=False,
    )

    policy_comparison.to_excel(
        writer,
        sheet_name="PolicyComparison",
        index=False,
    )

    similarity_summary.to_excel(
        writer,
        sheet_name="SimilarityComparison",
        index=False,
    )

    workbook = writer.book

    for worksheet in workbook.worksheets:
        style_worksheet(
            worksheet
        )

    feature_sheet = workbook[
        "FeatureComparison"
    ]

    for row_number in range(
        2,
        feature_sheet.max_row + 1,
    ):
        metric_name = clean_text(
            feature_sheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if "Percent" in metric_name:
            for column_number in [
                2,
                3,
            ]:
                feature_sheet.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = '0.00"%"'

    similarity_sheet = workbook[
        "SimilarityComparison"
    ]

    for row_number in range(
        2,
        similarity_sheet.max_row + 1,
    ):
        metric_name = clean_text(
            similarity_sheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if "Similarity Percent" in metric_name:
            similarity_sheet.cell(
                row=row_number,
                column=2,
            ).number_format = '0.00"%"'


# ============================================================
# CREATE MARKDOWN REPORT
# ============================================================

shared_topic_text = (
    ", ".join(shared_topics)
    if shared_topics
    else "None identified"
)

unique_a_text = (
    ", ".join(unique_to_a)
    if unique_to_a
    else "None identified"
)

unique_b_text = (
    ", ".join(unique_to_b)
    if unique_to_b
    else "None identified"
)

markdown_content = f"""# Candidate Comparison

## {CANDIDATE_A_NAME} vs. {CANDIDATE_B_NAME}

**Combined similarity:** {combined_similarity:.2f}%  
**Cosine similarity:** {safe_float(pair_row.get('Cosine_Similarity_Percent', 0)):.2f}%  
**Jaccard similarity:** {safe_float(pair_row.get('Jaccard_Similarity_Percent', 0)):.2f}%

{similarity_interpretation}

## Shared policy topics

{shared_topic_text}

## Topics unique to {CANDIDATE_A_NAME}

{unique_a_text}

## Topics unique to {CANDIDATE_B_NAME}

{unique_b_text}

## Candidate feature comparison

| Metric | {CANDIDATE_A_NAME} | {CANDIDATE_B_NAME} |
|---|---:|---:|
| Research completeness | {safe_float(candidate_a_overview.get('Research_Completeness_Percent', 0)):.2f}% | {safe_float(candidate_b_overview.get('Research_Completeness_Percent', 0)):.2f}% |
| Source count | {safe_int(candidate_a_overview.get('Source_Count', 0))} | {safe_int(candidate_b_overview.get('Source_Count', 0))} |
| Policy diversity | {safe_int(candidate_a_features.get('Policy_Diversity', 0))} | {safe_int(candidate_b_features.get('Policy_Diversity', 0))} |
| Total policy weight | {safe_float(candidate_a_features.get('Total_Policy_Weight', 0)):.2f} | {safe_float(candidate_b_features.get('Total_Policy_Weight', 0)):.2f} |
| Dominant policy | {clean_text(candidate_a_features.get('Dominant_Policy', ''))} | {clean_text(candidate_b_features.get('Dominant_Policy', ''))} |
| Dominant policy percent | {safe_float(candidate_a_features.get('Dominant_Policy_Percent', 0)):.2f}% | {safe_float(candidate_b_features.get('Dominant_Policy_Percent', 0)):.2f}% |

## Top policy topics

**{CANDIDATE_A_NAME}:** {", ".join(top_topics_a)}

**{CANDIDATE_B_NAME}:** {", ".join(top_topics_b)}

## Methodological note

This comparison is based on policy-topic keyword weights extracted from the current pilot statement dataset. Similarity does not measure ideology, candidate quality, legislative feasibility, or the importance voters assign to each topic. All files remain local and were not publicly deployed.
"""

OUTPUT_MARKDOWN.write_text(
    markdown_content,
    encoding="utf-8",
)


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("CANDIDATE COMPARISON CREATED")
print("=" * 72)

print(
    f"Candidate A: {CANDIDATE_A_NAME}"
)

print(
    f"Candidate B: {CANDIDATE_B_NAME}"
)

print()

print(f"Excel output:\n{OUTPUT_EXCEL}")
print()

print(
    f"Markdown output:\n{OUTPUT_MARKDOWN}"
)

print()

print(
    "Combined similarity: "
    f"{combined_similarity:.2f}%"
)

print(
    f"Shared policy topics: "
    f"{len(shared_topics)}"
)

print(
    f"Topics unique to {CANDIDATE_A_NAME}: "
    f"{len(unique_to_a)}"
)

print(
    f"Topics unique to {CANDIDATE_B_NAME}: "
    f"{len(unique_to_b)}"
)

print()

print(
    "Next step: open candidate_comparison.xlsx "
    "and review ComparisonSummary, FeatureComparison, "
    "PolicyComparison, and SimilarityComparison."
)