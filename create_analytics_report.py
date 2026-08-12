from pathlib import Path
from datetime import datetime
import math

import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "political_analytics_dashboard_data.xlsx"
)

REPORTS_FOLDER = (
    PROJECT_FOLDER
    / "reports"
)

OUTPUT_EXCEL = (
    REPORTS_FOLDER
    / "analytics_summary.xlsx"
)

OUTPUT_MARKDOWN = (
    REPORTS_FOLDER
    / "analytics_summary.md"
)


# ============================================================
# CONFIGURATION
# ============================================================

TOP_TOPIC_COUNT = 15

ANALYSIS_DATE = datetime.now().strftime(
    "%Y-%m-%d %H:%M:%S"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

KPI_FILL = PatternFill(
    fill_type="solid",
    fgColor="EDF2F7",
)

GOOD_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="D9E2F3",
    ),
    right=Side(
        style="thin",
        color="D9E2F3",
    ),
    top=Side(
        style="thin",
        color="D9E2F3",
    ),
    bottom=Side(
        style="thin",
        color="D9E2F3",
    ),
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop with a clear message when a required file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to blank text and collapse whitespace.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def numeric_series(series):
    """
    Convert a pandas Series to numeric values.
    Missing and invalid values become zero.
    """
    return pd.to_numeric(
        series,
        errors="coerce",
    ).fillna(0)


def safe_float(value, default=0.0):
    """
    Convert one value to float safely.
    """
    try:
        result = pd.to_numeric(
            value,
            errors="coerce",
        )

        if pd.isna(result):
            return default

        return float(result)

    except Exception:
        return default


def safe_int(value, default=0):
    """
    Convert one value to integer safely.
    """
    return int(
        round(
            safe_float(
                value,
                default,
            )
        )
    )


def safe_percentage(
    numerator,
    denominator,
):
    """
    Calculate a percentage safely.
    """
    denominator_value = safe_float(
        denominator
    )

    if denominator_value == 0:
        return 0.0

    return round(
        safe_float(numerator)
        / denominator_value
        * 100,
        2,
    )


def get_summary_value(
    summary_dataframe,
    metric_name,
    default=0,
):
    """
    Retrieve one metric from the DashboardSummary sheet.
    """
    if summary_dataframe.empty:
        return default

    if (
        "Metric" not in summary_dataframe.columns
        or "Value" not in summary_dataframe.columns
    ):
        return default

    matching_rows = summary_dataframe[
        summary_dataframe["Metric"]
        .astype(str)
        .str.strip()
        == metric_name
    ]

    if matching_rows.empty:
        return default

    return matching_rows.iloc[0]["Value"]


def write_dataframe(
    worksheet,
    dataframe,
    start_row=1,
    start_column=1,
    include_header=True,
):
    """
    Write a pandas DataFrame into an openpyxl worksheet.

    Returns:
        first_data_row,
        last_data_row,
        last_column
    """
    current_row = start_row

    if include_header:
        for column_offset, column_name in enumerate(
            dataframe.columns,
            start=start_column,
        ):
            cell = worksheet.cell(
                row=current_row,
                column=column_offset,
                value=column_name,
            )

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = THIN_BORDER

        current_row += 1

    first_data_row = current_row

    for _, dataframe_row in dataframe.iterrows():
        for column_offset, column_name in enumerate(
            dataframe.columns,
            start=start_column,
        ):
            value = dataframe_row[column_name]

            if pd.isna(value):
                value = ""

            cell = worksheet.cell(
                row=current_row,
                column=column_offset,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = THIN_BORDER

        current_row += 1

    last_data_row = max(
        first_data_row,
        current_row - 1,
    )

    last_column = (
        start_column
        + len(dataframe.columns)
        - 1
    )

    return (
        first_data_row,
        last_data_row,
        last_column,
    )


def set_reasonable_widths(
    worksheet,
    minimum=10,
    maximum=42,
):
    """
    Set readable but bounded Excel column widths.
    """
    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                minimum,
            ),
            maximum,
        )


def style_table_sheet(
    worksheet,
    freeze_cell="A2",
):
    """
    Apply basic styling to a table-based worksheet.
    """
    worksheet.freeze_panes = freeze_cell

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = THIN_BORDER

    set_reasonable_widths(
        worksheet
    )


def add_section_title(
    worksheet,
    row_number,
    title,
    end_column=8,
):
    """
    Add a merged dashboard section title.
    """
    worksheet.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=end_column,
    )

    cell = worksheet.cell(
        row=row_number,
        column=1,
        value=title,
    )

    cell.fill = SECTION_FILL
    cell.font = Font(
        bold=True,
        size=13,
        color="1F4E78",
    )
    cell.alignment = Alignment(
        vertical="center",
    )

    worksheet.row_dimensions[
        row_number
    ].height = 25


def add_kpi_card(
    worksheet,
    start_row,
    start_column,
    label,
    value,
    number_format=None,
):
    """
    Create one merged KPI card on the dashboard.
    """
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row + 1,
        end_column=start_column + 1,
    )

    value_cell = worksheet.cell(
        row=start_row,
        column=start_column,
        value=value,
    )

    value_cell.fill = KPI_FILL
    value_cell.font = Font(
        bold=True,
        size=17,
        color="1F4E78",
    )
    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    value_cell.border = THIN_BORDER

    if number_format:
        value_cell.number_format = (
            number_format
        )

    worksheet.merge_cells(
        start_row=start_row + 2,
        start_column=start_column,
        end_row=start_row + 2,
        end_column=start_column + 1,
    )

    label_cell = worksheet.cell(
        row=start_row + 2,
        column=start_column,
        value=label,
    )

    label_cell.fill = KPI_FILL
    label_cell.font = Font(
        size=9,
        color="4A5568",
    )
    label_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    label_cell.border = THIN_BORDER


def add_bar_chart(
    worksheet,
    data_min_col,
    data_max_col,
    data_min_row,
    data_max_row,
    category_col,
    category_min_row,
    category_max_row,
    title,
    y_axis_title,
    x_axis_title,
    anchor,
    width=12,
    height=7,
    horizontal=False,
    show_legend=True,
):
    """
    Add a native Excel bar or column chart.
    """
    chart = BarChart()

    if horizontal:
        chart.type = "bar"
        chart.style = 10

    else:
        chart.type = "col"
        chart.style = 10

    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.width = width
    chart.height = height

    data = Reference(
        worksheet,
        min_col=data_min_col,
        max_col=data_max_col,
        min_row=data_min_row,
        max_row=data_max_row,
    )

    categories = Reference(
        worksheet,
        min_col=category_col,
        min_row=category_min_row,
        max_row=category_max_row,
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(
        categories
    )

    chart.legend = (
        chart.legend
        if show_legend
        else None
    )

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    worksheet.add_chart(
        chart,
        anchor,
    )


# ============================================================
# LOAD INPUT DATA
# ============================================================

require_file(
    INPUT_FILE,
    "Dashboard dataset workbook",
)

REPORTS_FOLDER.mkdir(
    parents=True,
    exist_ok=True,
)

excel_file = pd.ExcelFile(
    INPUT_FILE
)

required_sheets = [
    "DashboardSummary",
    "CandidateOverview",
    "CandidateHighlights",
    "CandidatePolicyMatrix",
    "CandidateTopicSummary",
    "TopicFrequency",
    "PolicyClassifications",
    "ClassifierOverallMetrics",
    "ClassifierTopicMetrics",
    "ClassifierConfidence",
    "ClassifierLabels",
]

missing_sheets = [
    sheet_name
    for sheet_name in required_sheets
    if sheet_name not in excel_file.sheet_names
]

if missing_sheets:
    raise ValueError(
        "The dashboard workbook is missing these sheets: "
        + ", ".join(missing_sheets)
    )

dashboard_summary = pd.read_excel(
    INPUT_FILE,
    sheet_name="DashboardSummary",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_overview = pd.read_excel(
    INPUT_FILE,
    sheet_name="CandidateOverview",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_highlights = pd.read_excel(
    INPUT_FILE,
    sheet_name="CandidateHighlights",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_policy_matrix = pd.read_excel(
    INPUT_FILE,
    sheet_name="CandidatePolicyMatrix",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_topic_summary = pd.read_excel(
    INPUT_FILE,
    sheet_name="CandidateTopicSummary",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

topic_frequency = pd.read_excel(
    INPUT_FILE,
    sheet_name="TopicFrequency",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

policy_classifications = pd.read_excel(
    INPUT_FILE,
    sheet_name="PolicyClassifications",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

overall_metrics = pd.read_excel(
    INPUT_FILE,
    sheet_name="ClassifierOverallMetrics",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

topic_metrics = pd.read_excel(
    INPUT_FILE,
    sheet_name="ClassifierTopicMetrics",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

confidence_metrics = pd.read_excel(
    INPUT_FILE,
    sheet_name="ClassifierConfidence",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

label_distribution = pd.read_excel(
    INPUT_FILE,
    sheet_name="ClassifierLabels",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# STANDARDIZE NUMERIC COLUMNS
# ============================================================

candidate_numeric_columns = [
    "Research_Completeness_Percent",
    "Source_Count",
    "Verified_Source_Count",
    "High_Reliability_Source_Count",
    "Fact_Count",
    "Statement_Count",
    "Social_Media_Count",
    "Unique_Policy_Topics",
    "Total_Policy_Records",
    "Total_Keyword_Matches",
    "Average_Classification_Score",
    "High_Confidence_Record_Count",
    "Medium_Confidence_Record_Count",
    "Low_Confidence_Record_Count",
    "Classification_Coverage_Percent",
]

for column in candidate_numeric_columns:
    if column in candidate_overview.columns:
        candidate_overview[column] = (
            numeric_series(
                candidate_overview[column]
            )
        )

topic_numeric_columns = [
    "Candidate_Count",
    "Statement_Count",
    "Classification_Record_Count",
    "Total_Keyword_Matches",
    "Average_Classification_Score",
]

for column in topic_numeric_columns:
    if column in topic_frequency.columns:
        topic_frequency[column] = (
            numeric_series(
                topic_frequency[column]
            )
        )

classification_numeric_columns = [
    "Keyword_Match_Count",
    "Classification_Score",
]

for column in classification_numeric_columns:
    if column in policy_classifications.columns:
        policy_classifications[column] = (
            numeric_series(
                policy_classifications[column]
            )
        )

if "Record_Count" in label_distribution.columns:
    label_distribution[
        "Record_Count"
    ] = numeric_series(
        label_distribution[
            "Record_Count"
        ]
    )


# ============================================================
# SUMMARY METRICS
# ============================================================

candidate_count = safe_int(
    get_summary_value(
        dashboard_summary,
        "Candidate Count",
        len(candidate_overview),
    )
)

district_count = safe_int(
    get_summary_value(
        dashboard_summary,
        "District Count",
        0,
    )
)

policy_topic_count = safe_int(
    get_summary_value(
        dashboard_summary,
        "Policy Topics Detected",
        0,
    )
)

policy_record_count = safe_int(
    get_summary_value(
        dashboard_summary,
        "Policy Classification Records",
        len(policy_classifications),
    )
)

average_research_completeness = safe_float(
    get_summary_value(
        dashboard_summary,
        "Average Research Completeness Percent",
        0,
    )
)

strict_precision = safe_float(
    get_summary_value(
        dashboard_summary,
        "Human-Reviewed Strict Precision Percent",
        0,
    )
)

relevant_rate = safe_float(
    get_summary_value(
        dashboard_summary,
        "Relevant or Partial Rate Percent",
        0,
    )
)

weighted_precision = safe_float(
    get_summary_value(
        dashboard_summary,
        "Weighted Precision Percent",
        0,
    )
)


# ============================================================
# PREPARE REPORT TABLES
# ============================================================

research_completeness_data = (
    candidate_overview[
        [
            "Ballot_Name",
            "Research_Completeness_Percent",
        ]
    ]
    .sort_values(
        by="Research_Completeness_Percent",
        ascending=False,
    )
    .reset_index(drop=True)
)

source_quality_data = (
    candidate_overview[
        [
            "Ballot_Name",
            "Source_Count",
            "Verified_Source_Count",
            "High_Reliability_Source_Count",
        ]
    ]
    .sort_values(
        by="Source_Count",
        ascending=False,
    )
    .reset_index(drop=True)
)

policy_frequency_data = (
    topic_frequency
    .sort_values(
        by="Total_Keyword_Matches",
        ascending=False,
    )
    .head(TOP_TOPIC_COUNT)
    .reset_index(drop=True)
)

policy_diversity_data = (
    candidate_overview[
        [
            "Ballot_Name",
            "Unique_Policy_Topics",
        ]
    ]
    .sort_values(
        by="Unique_Policy_Topics",
        ascending=False,
    )
    .reset_index(drop=True)
)

confidence_order = [
    "High",
    "Medium",
    "Low",
    "Unknown",
]

confidence_distribution = (
    policy_classifications[
        "Classification_Confidence"
    ]
    .fillna("Unknown")
    .astype(str)
    .value_counts()
    .reindex(
        confidence_order,
        fill_value=0,
    )
    .rename_axis(
        "Classification_Confidence"
    )
    .reset_index(
        name="Record_Count"
    )
)

label_order = [
    "Correct",
    "Partially Correct",
    "Incorrect",
]

classifier_label_data = (
    label_distribution
    .set_index("Manual_Label")
    .reindex(label_order)
    .fillna(0)
    .reset_index()
)

summary_metrics = pd.DataFrame(
    [
        {
            "Metric": "Candidates Analyzed",
            "Value": candidate_count,
        },
        {
            "Metric": "Districts Included",
            "Value": district_count,
        },
        {
            "Metric": "Policy Topics Detected",
            "Value": policy_topic_count,
        },
        {
            "Metric": "Policy Classification Records",
            "Value": policy_record_count,
        },
        {
            "Metric": (
                "Average Research Completeness Percent"
            ),
            "Value": round(
                average_research_completeness,
                2,
            ),
        },
        {
            "Metric": (
                "Human-Reviewed Strict Precision Percent"
            ),
            "Value": round(
                strict_precision,
                2,
            ),
        },
        {
            "Metric": (
                "Relevant or Partial Rate Percent"
            ),
            "Value": round(
                relevant_rate,
                2,
            ),
        },
        {
            "Metric": (
                "Weighted Precision Percent"
            ),
            "Value": round(
                weighted_precision,
                2,
            ),
        },
        {
            "Metric": "Report Created",
            "Value": ANALYSIS_DATE,
        },
        {
            "Metric": "Deployment Status",
            "Value": "Local only",
        },
    ]
)

top_topics = (
    topic_frequency
    .sort_values(
        by="Total_Keyword_Matches",
        ascending=False,
    )
    .head(5)
)

most_complete_candidate = (
    candidate_overview
    .sort_values(
        by="Research_Completeness_Percent",
        ascending=False,
    )
    .iloc[0]
)

most_diverse_candidate = (
    candidate_overview
    .sort_values(
        by="Unique_Policy_Topics",
        ascending=False,
    )
    .iloc[0]
)

key_findings = pd.DataFrame(
    [
        {
            "Finding": (
                "Most frequent policy topic"
            ),
            "Result": clean_text(
                top_topics.iloc[0][
                    "Policy_Topic"
                ]
            ),
            "Evidence": (
                f"{safe_int(top_topics.iloc[0]['Total_Keyword_Matches'])} "
                "keyword matches"
            ),
        },
        {
            "Finding": (
                "Highest research completeness"
            ),
            "Result": clean_text(
                most_complete_candidate[
                    "Ballot_Name"
                ]
            ),
            "Evidence": (
                f"{safe_float(most_complete_candidate['Research_Completeness_Percent']):.2f}%"
            ),
        },
        {
            "Finding": (
                "Highest policy diversity"
            ),
            "Result": clean_text(
                most_diverse_candidate[
                    "Ballot_Name"
                ]
            ),
            "Evidence": (
                f"{safe_int(most_diverse_candidate['Unique_Policy_Topics'])} "
                "unique topics"
            ),
        },
        {
            "Finding": (
                "Classifier pilot strict precision"
            ),
            "Result": (
                f"{strict_precision:.2f}%"
            ),
            "Evidence": (
                "47 human-reviewed classifications"
            ),
        },
        {
            "Finding": (
                "Relevant or partially relevant rate"
            ),
            "Result": (
                f"{relevant_rate:.2f}%"
            ),
            "Evidence": (
                "Correct and partial labels combined"
            ),
        },
    ]
)


# ============================================================
# CREATE WORKBOOK
# ============================================================

workbook = Workbook()

dashboard_sheet = workbook.active
dashboard_sheet.title = "Dashboard"

candidate_sheet = workbook.create_sheet(
    "CandidateOverview"
)

topic_sheet = workbook.create_sheet(
    "TopicFrequency"
)

heatmap_sheet = workbook.create_sheet(
    "CandidatePolicyHeatmap"
)

classifier_sheet = workbook.create_sheet(
    "ClassifierMetrics"
)

highlight_sheet = workbook.create_sheet(
    "CandidateHighlights"
)

topic_summary_sheet = workbook.create_sheet(
    "CandidateTopicSummary"
)

summary_sheet = workbook.create_sheet(
    "SummaryMetrics"
)

findings_sheet = workbook.create_sheet(
    "KeyFindings"
)


# ============================================================
# DASHBOARD TITLE
# ============================================================

dashboard_sheet.sheet_view.showGridLines = False

dashboard_sheet.merge_cells(
    "A1:L2"
)

title_cell = dashboard_sheet["A1"]

title_cell.value = (
    "Political Candidate Analytics Dashboard"
)

title_cell.font = Font(
    bold=True,
    size=21,
    color="FFFFFF",
)

title_cell.fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

title_cell.alignment = Alignment(
    horizontal="center",
    vertical="center",
)

dashboard_sheet.row_dimensions[1].height = 30
dashboard_sheet.row_dimensions[2].height = 20

dashboard_sheet.merge_cells(
    "A3:L3"
)

subtitle_cell = dashboard_sheet["A3"]

subtitle_cell.value = (
    "Local-only pilot analysis — candidate research, "
    "policy NLP classification, and human evaluation"
)

subtitle_cell.font = Font(
    italic=True,
    size=10,
    color="4A5568",
)

subtitle_cell.alignment = Alignment(
    horizontal="center",
    vertical="center",
)


# ============================================================
# KPI CARDS
# ============================================================

add_kpi_card(
    dashboard_sheet,
    5,
    1,
    "Candidates",
    candidate_count,
)

add_kpi_card(
    dashboard_sheet,
    5,
    3,
    "Districts",
    district_count,
)

add_kpi_card(
    dashboard_sheet,
    5,
    5,
    "Policy Topics",
    policy_topic_count,
)

add_kpi_card(
    dashboard_sheet,
    5,
    7,
    "Policy Records",
    policy_record_count,
)

add_kpi_card(
    dashboard_sheet,
    5,
    9,
    "Research Completeness",
    average_research_completeness / 100,
    "0.00%",
)

add_kpi_card(
    dashboard_sheet,
    5,
    11,
    "Strict Precision",
    strict_precision / 100,
    "0.00%",
)


# ============================================================
# DASHBOARD SUPPORTING DATA
# ============================================================

add_section_title(
    dashboard_sheet,
    10,
    "Research Completeness",
    12,
)

research_start_row = 11

research_first_row, research_last_row, _ = (
    write_dataframe(
        dashboard_sheet,
        research_completeness_data,
        start_row=research_start_row,
        start_column=1,
    )
)

add_section_title(
    dashboard_sheet,
    10,
    "Research Completeness",
    12,
)

add_bar_chart(
    worksheet=dashboard_sheet,
    data_min_col=2,
    data_max_col=2,
    data_min_row=research_start_row,
    data_max_row=research_last_row,
    category_col=1,
    category_min_row=research_first_row,
    category_max_row=research_last_row,
    title="Research Completeness by Candidate",
    y_axis_title="Candidate",
    x_axis_title="Completeness (%)",
    anchor="D11",
    width=14,
    height=7,
    horizontal=True,
    show_legend=False,
)

policy_section_row = max(
    research_last_row + 3,
    27,
)

add_section_title(
    dashboard_sheet,
    policy_section_row,
    "Leading Policy Topics",
    12,
)

policy_table_header_row = (
    policy_section_row + 1
)

policy_first_row, policy_last_row, _ = (
    write_dataframe(
        dashboard_sheet,
        policy_frequency_data[
            [
                "Policy_Topic",
                "Candidate_Count",
                "Statement_Count",
                "Total_Keyword_Matches",
            ]
        ],
        start_row=policy_table_header_row,
        start_column=1,
    )
)

add_bar_chart(
    worksheet=dashboard_sheet,
    data_min_col=4,
    data_max_col=4,
    data_min_row=policy_table_header_row,
    data_max_row=policy_last_row,
    category_col=1,
    category_min_row=policy_first_row,
    category_max_row=policy_last_row,
    title="Top Policy Topics",
    y_axis_title="Policy Topic",
    x_axis_title="Keyword Matches",
    anchor=f"F{policy_table_header_row}",
    width=14,
    height=9,
    horizontal=True,
    show_legend=False,
)

classifier_section_row = max(
    policy_last_row + 3,
    policy_section_row + 20,
)

add_section_title(
    dashboard_sheet,
    classifier_section_row,
    "Classifier Evaluation",
    12,
)

classifier_table_row = (
    classifier_section_row + 1
)

classifier_first_row, classifier_last_row, _ = (
    write_dataframe(
        dashboard_sheet,
        classifier_label_data[
            [
                "Manual_Label",
                "Record_Count",
            ]
        ],
        start_row=classifier_table_row,
        start_column=1,
    )
)

add_bar_chart(
    worksheet=dashboard_sheet,
    data_min_col=2,
    data_max_col=2,
    data_min_row=classifier_table_row,
    data_max_row=classifier_last_row,
    category_col=1,
    category_min_row=classifier_first_row,
    category_max_row=classifier_last_row,
    title="Human Review Labels",
    y_axis_title="Records",
    x_axis_title="Review Label",
    anchor=f"D{classifier_table_row}",
    width=11,
    height=7,
    horizontal=False,
    show_legend=False,
)

confidence_table_row = (
    classifier_section_row + 11
)

confidence_first_row, confidence_last_row, _ = (
    write_dataframe(
        dashboard_sheet,
        confidence_distribution,
        start_row=confidence_table_row,
        start_column=1,
    )
)

add_bar_chart(
    worksheet=dashboard_sheet,
    data_min_col=2,
    data_max_col=2,
    data_min_row=confidence_table_row,
    data_max_row=confidence_last_row,
    category_col=1,
    category_min_row=confidence_first_row,
    category_max_row=confidence_last_row,
    title="Classification Confidence",
    y_axis_title="Records",
    x_axis_title="Confidence",
    anchor=f"D{confidence_table_row}",
    width=11,
    height=7,
    horizontal=False,
    show_legend=False,
)


# ============================================================
# CANDIDATE OVERVIEW SHEET
# ============================================================

write_dataframe(
    candidate_sheet,
    candidate_overview,
)

style_table_sheet(
    candidate_sheet
)

percentage_candidate_columns = [
    "Research_Completeness_Percent",
    "Verified_Source_Rate_Percent",
    "High_Reliability_Source_Rate_Percent",
    "Classification_Coverage_Percent",
]

candidate_headers = {
    candidate_sheet.cell(
        row=1,
        column=column_number,
    ).value: column_number
    for column_number in range(
        1,
        candidate_sheet.max_column + 1,
    )
}

for column_name in percentage_candidate_columns:
    if column_name not in candidate_headers:
        continue

    column_number = candidate_headers[
        column_name
    ]

    for row_number in range(
        2,
        candidate_sheet.max_row + 1,
    ):
        candidate_sheet.cell(
            row=row_number,
            column=column_number,
        ).number_format = '0.00"%"'


# ============================================================
# TOPIC FREQUENCY SHEET
# ============================================================

write_dataframe(
    topic_sheet,
    topic_frequency,
)

style_table_sheet(
    topic_sheet
)

topic_chart = BarChart()
topic_chart.type = "bar"
topic_chart.style = 10
topic_chart.title = (
    "Policy Topics by Keyword Matches"
)
topic_chart.x_axis.title = (
    "Keyword Matches"
)
topic_chart.y_axis.title = (
    "Policy Topic"
)
topic_chart.width = 16
topic_chart.height = 10
topic_chart.dLbls = DataLabelList()
topic_chart.dLbls.showVal = True

topic_headers = {
    topic_sheet.cell(
        row=1,
        column=column_number,
    ).value: column_number
    for column_number in range(
        1,
        topic_sheet.max_column + 1,
    )
}

if (
    "Policy_Topic" in topic_headers
    and "Total_Keyword_Matches"
    in topic_headers
):
    topic_data = Reference(
        topic_sheet,
        min_col=topic_headers[
            "Total_Keyword_Matches"
        ],
        min_row=1,
        max_row=min(
            TOP_TOPIC_COUNT + 1,
            topic_sheet.max_row,
        ),
    )

    topic_categories = Reference(
        topic_sheet,
        min_col=topic_headers[
            "Policy_Topic"
        ],
        min_row=2,
        max_row=min(
            TOP_TOPIC_COUNT + 1,
            topic_sheet.max_row,
        ),
    )

    topic_chart.add_data(
        topic_data,
        titles_from_data=True,
    )

    topic_chart.set_categories(
        topic_categories
    )

    topic_chart.legend = None

    topic_sheet.add_chart(
        topic_chart,
        "H2",
    )


# ============================================================
# POLICY HEATMAP SHEET
# ============================================================

write_dataframe(
    heatmap_sheet,
    candidate_policy_matrix,
)

style_table_sheet(
    heatmap_sheet
)

heatmap_sheet.freeze_panes = "C2"

identifier_columns = [
    column
    for column in [
        "Candidate_ID",
        "Ballot_Name",
    ]
    if column
    in candidate_policy_matrix.columns
]

policy_columns = [
    column
    for column
    in candidate_policy_matrix.columns
    if column not in identifier_columns
]

if policy_columns:
    first_policy_column = (
        candidate_policy_matrix.columns.get_loc(
            policy_columns[0]
        )
        + 1
    )

    last_policy_column = (
        candidate_policy_matrix.columns.get_loc(
            policy_columns[-1]
        )
        + 1
    )

    heatmap_range = (
        f"{get_column_letter(first_policy_column)}2:"
        f"{get_column_letter(last_policy_column)}"
        f"{heatmap_sheet.max_row}"
    )

    heatmap_sheet.conditional_formatting.add(
        heatmap_range,
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="5B9BD5",
        ),
    )

    for column_number in range(
        first_policy_column,
        last_policy_column + 1,
    ):
        heatmap_sheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = 15

    for row_number in range(
        2,
        heatmap_sheet.max_row + 1,
    ):
        for column_number in range(
            first_policy_column,
            last_policy_column + 1,
        ):
            heatmap_sheet.cell(
                row=row_number,
                column=column_number,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


# ============================================================
# CLASSIFIER METRICS SHEET
# ============================================================

classifier_sheet.sheet_view.showGridLines = False

classifier_sheet["A1"] = (
    "Classifier Evaluation Summary"
)

classifier_sheet["A1"].font = Font(
    bold=True,
    size=18,
    color="FFFFFF",
)

classifier_sheet["A1"].fill = HEADER_FILL

classifier_sheet.merge_cells(
    "A1:H2"
)

classifier_sheet["A1"].alignment = Alignment(
    horizontal="center",
    vertical="center",
)

add_kpi_card(
    classifier_sheet,
    4,
    1,
    "Strict Precision",
    strict_precision / 100,
    "0.00%",
)

add_kpi_card(
    classifier_sheet,
    4,
    3,
    "Relevant or Partial",
    relevant_rate / 100,
    "0.00%",
)

add_kpi_card(
    classifier_sheet,
    4,
    5,
    "Weighted Precision",
    weighted_precision / 100,
    "0.00%",
)

add_kpi_card(
    classifier_sheet,
    4,
    7,
    "Reviewed Records",
    safe_int(
        classifier_label_data[
            "Record_Count"
        ].sum()
    ),
)

write_dataframe(
    classifier_sheet,
    classifier_label_data[
        [
            "Manual_Label",
            "Record_Count",
        ]
    ],
    start_row=9,
    start_column=1,
)

write_dataframe(
    classifier_sheet,
    confidence_distribution,
    start_row=9,
    start_column=5,
)

add_bar_chart(
    worksheet=classifier_sheet,
    data_min_col=2,
    data_max_col=2,
    data_min_row=9,
    data_max_row=12,
    category_col=1,
    category_min_row=10,
    category_max_row=12,
    title="Human Review Labels",
    y_axis_title="Records",
    x_axis_title="Label",
    anchor="A15",
    width=10,
    height=7,
    horizontal=False,
    show_legend=False,
)

add_bar_chart(
    worksheet=classifier_sheet,
    data_min_col=6,
    data_max_col=6,
    data_min_row=9,
    data_max_row=13,
    category_col=5,
    category_min_row=10,
    category_max_row=13,
    title="Classification Confidence",
    y_axis_title="Records",
    x_axis_title="Confidence",
    anchor="F15",
    width=10,
    height=7,
    horizontal=False,
    show_legend=False,
)

classifier_topic_start_row = 31

classifier_sheet.cell(
    row=classifier_topic_start_row,
    column=1,
    value="Performance by Policy Topic",
)

classifier_sheet.cell(
    row=classifier_topic_start_row,
    column=1,
).font = Font(
    bold=True,
    size=13,
    color="1F4E78",
)

write_dataframe(
    classifier_sheet,
    topic_metrics,
    start_row=classifier_topic_start_row + 1,
    start_column=1,
)

set_reasonable_widths(
    classifier_sheet,
    maximum=35,
)


# ============================================================
# SUPPORTING TABLE SHEETS
# ============================================================

write_dataframe(
    highlight_sheet,
    candidate_highlights,
)

style_table_sheet(
    highlight_sheet
)

write_dataframe(
    topic_summary_sheet,
    candidate_topic_summary,
)

style_table_sheet(
    topic_summary_sheet
)

write_dataframe(
    summary_sheet,
    summary_metrics,
)

style_table_sheet(
    summary_sheet
)

write_dataframe(
    findings_sheet,
    key_findings,
)

style_table_sheet(
    findings_sheet
)


# ============================================================
# DASHBOARD FINAL FORMATTING
# ============================================================

dashboard_sheet.freeze_panes = "A4"

for column_number in range(
    1,
    13,
):
    dashboard_sheet.column_dimensions[
        get_column_letter(
            column_number
        )
    ].width = 14

dashboard_sheet.column_dimensions[
    "A"
].width = 25

dashboard_sheet.column_dimensions[
    "B"
].width = 18

dashboard_sheet.column_dimensions[
    "C"
].width = 18

dashboard_sheet.column_dimensions[
    "D"
].width = 18

for row in dashboard_sheet.iter_rows(
    min_row=4,
):
    for cell in row:
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

dashboard_sheet["A70"] = (
    "Privacy note: all reports and charts are stored "
    "locally. Nothing was uploaded or publicly deployed."
)

dashboard_sheet["A70"].font = Font(
    italic=True,
    color="4A5568",
)

dashboard_sheet.merge_cells(
    "A70:L70"
)


# ============================================================
# SAVE WORKBOOK
# ============================================================

workbook.save(
    OUTPUT_EXCEL
)


# ============================================================
# CREATE MARKDOWN SUMMARY
# ============================================================

top_topic_lines = []

for index, (_, row) in enumerate(
    top_topics.iterrows(),
    start=1,
):
    top_topic_lines.append(
        (
            f"{index}. **{clean_text(row['Policy_Topic'])}** — "
            f"{safe_int(row['Candidate_Count'])} candidates, "
            f"{safe_int(row['Statement_Count'])} statements, "
            f"{safe_int(row['Total_Keyword_Matches'])} keyword matches"
        )
    )

candidate_lines = []

for _, row in candidate_highlights.iterrows():
    candidate_lines.append(
        (
            f"- **{clean_text(row.get('Ballot_Name'))}** — "
            f"{safe_float(row.get('Research_Completeness_Percent')):.2f}% "
            f"research completeness; "
            f"{safe_int(row.get('Source_Count'))} structured sources; "
            f"{safe_int(row.get('Unique_Policy_Topics'))} policy topics. "
            f"Top topics: {clean_text(row.get('Top_5_Policy_Topics'))}."
        )
    )

markdown_content = f"""# Political Candidate Analytics Summary

**Created:** {ANALYSIS_DATE}  
**Status:** Local-only analysis. Nothing was uploaded or publicly deployed.

## Project overview

This pilot project transforms candidate research into a normalized, source-tracked analytical dataset. Python scripts validate database relationships, clean candidate records, classify campaign statements into policy topics, and evaluate the classifier using human-reviewed labels.

## Dataset summary

- Candidates analyzed: **{candidate_count}**
- Districts included: **{district_count}**
- Policy topics detected: **{policy_topic_count}**
- Policy classification records: **{policy_record_count}**
- Average research completeness: **{average_research_completeness:.2f}%**

## Classifier evaluation

- Strict precision: **{strict_precision:.2f}%**
- Relevant or partially relevant rate: **{relevant_rate:.2f}%**
- Weighted precision: **{weighted_precision:.2f}%**
- Evaluation sample: **47 human-reviewed classifications**

These results apply only to the current pilot sample and should not yet be generalized to all candidates or all political text.

## Leading policy topics

{chr(10).join(top_topic_lines)}

## Candidate highlights

{chr(10).join(candidate_lines)}

## Key findings

- **{clean_text(most_complete_candidate['Ballot_Name'])}** had the highest research completeness at **{safe_float(most_complete_candidate['Research_Completeness_Percent']):.2f}%**.
- **{clean_text(most_diverse_candidate['Ballot_Name'])}** had the highest detected policy diversity with **{safe_int(most_diverse_candidate['Unique_Policy_Topics'])} policy topics**.
- **{clean_text(top_topics.iloc[0]['Policy_Topic'])}** was the most frequently detected policy topic.
- The rule-based classifier retained the exact keywords responsible for every classification.
- The pilot validation pipeline found no critical structural or referential-integrity errors.

## Excel workbook

The local Excel report contains:

- `Dashboard`
- `CandidateOverview`
- `TopicFrequency`
- `CandidatePolicyHeatmap`
- `ClassifierMetrics`
- `CandidateHighlights`
- `CandidateTopicSummary`
- `SummaryMetrics`
- `KeyFindings`

The charts in `analytics_summary.xlsx` are native Excel charts. The candidate-policy heatmap uses Excel conditional formatting.

## Limitations

- The pilot includes only five candidates and two districts.
- Some candidate statements are AI-assisted summaries based on verified sources rather than direct quotations.
- Keyword frequency measures emphasis in the available text, not the feasibility or strength of a policy commitment.
- The human review used a single first-pass annotation rather than multiple independent reviewers.
- Similarity analysis will become more meaningful after expanding the candidate database.

## Next steps

1. Build candidate policy vectors.
2. Calculate cosine and Jaccard similarity.
3. Create a candidate-similarity matrix.
4. Build a local-only Streamlit dashboard.
5. Expand the database to 20–30 candidates.
6. Prepare final portfolio and résumé documentation.
"""

OUTPUT_MARKDOWN.write_text(
    markdown_content,
    encoding="utf-8",
)


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("LOCAL EXCEL ANALYTICS REPORT CREATED")
print("=" * 72)

print(f"Input workbook:\n{INPUT_FILE}")
print()

print(f"Excel dashboard:\n{OUTPUT_EXCEL}")
print()

print(f"Markdown summary:\n{OUTPUT_MARKDOWN}")
print()

print("Excel report sheets:")

for sheet_name in workbook.sheetnames:
    print(f"  - {sheet_name}")

print()

print(
    f"Candidates analyzed: {candidate_count}"
)

print(
    f"Policy topics analyzed: {policy_topic_count}"
)

print(
    f"Policy classification records: "
    f"{policy_record_count}"
)

print(
    "Human-reviewed strict precision: "
    f"{strict_precision:.2f}%"
)

print()

print(
    "No Matplotlib or ReportLab was used. "
    "All charts are native Excel charts and remain local."
)

print()

print(
    "Next step: open analytics_summary.xlsx and inspect "
    "the Dashboard, CandidatePolicyHeatmap, and "
    "ClassifierMetrics sheets."
)