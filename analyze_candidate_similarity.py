from pathlib import Path
from itertools import combinations

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

INPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "political_analytics_dashboard_data.xlsx"
)

OUTPUT_EXCEL = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_similarity_analysis.xlsx"
)

OUTPUT_COSINE_CSV = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_cosine_similarity.csv"
)

OUTPUT_JACCARD_CSV = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_jaccard_similarity.csv"
)

OUTPUT_COMBINED_CSV = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "candidate_combined_similarity.csv"
)


# ============================================================
# CONFIGURATION
# ============================================================

IDENTIFIER_COLUMNS = [
    "Candidate_ID",
    "Ballot_Name",
]

SIMILARITY_DECIMALS = 4

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop with a clear error if a required file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to blank text and normalize whitespace.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def cosine_similarity(vector_a, vector_b):
    """
    Calculate cosine similarity between two numeric vectors.

    Returns zero when either vector has no magnitude.
    """
    vector_a = np.asarray(
        vector_a,
        dtype=float,
    )

    vector_b = np.asarray(
        vector_b,
        dtype=float,
    )

    denominator = (
        np.linalg.norm(vector_a)
        * np.linalg.norm(vector_b)
    )

    if denominator == 0:
        return 0.0

    return float(
        np.dot(vector_a, vector_b)
        / denominator
    )


def jaccard_similarity(vector_a, vector_b):
    """
    Calculate Jaccard similarity using topic presence or absence.

    A policy topic is considered present when its value is greater than zero.
    """
    binary_a = np.asarray(
        vector_a,
        dtype=float,
    ) > 0

    binary_b = np.asarray(
        vector_b,
        dtype=float,
    ) > 0

    intersection = np.logical_and(
        binary_a,
        binary_b,
    ).sum()

    union = np.logical_or(
        binary_a,
        binary_b,
    ).sum()

    if union == 0:
        return 0.0

    return float(
        intersection / union
    )


def safe_percentage(value):
    """
    Convert a zero-to-one similarity score into a percentage.
    """
    try:
        return round(
            float(value) * 100,
            2,
        )

    except Exception:
        return 0.0


def get_shared_topics(
    row_a,
    row_b,
    policy_columns,
):
    """
    Return policy topics present for both candidates.
    """
    shared_topics = []

    for topic in policy_columns:
        value_a = float(
            row_a.get(topic, 0)
        )

        value_b = float(
            row_b.get(topic, 0)
        )

        if value_a > 0 and value_b > 0:
            shared_topics.append(topic)

    return shared_topics


def get_unique_topics(
    row_a,
    row_b,
    policy_columns,
):
    """
    Return topics present for candidate A but absent for candidate B.
    """
    unique_topics = []

    for topic in policy_columns:
        value_a = float(
            row_a.get(topic, 0)
        )

        value_b = float(
            row_b.get(topic, 0)
        )

        if value_a > 0 and value_b == 0:
            unique_topics.append(topic)

    return unique_topics


def style_table_sheet(worksheet):
    """
    Apply consistent formatting to a table worksheet.
    """
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            55,
        )


def add_similarity_heatmap(
    worksheet,
    start_row=2,
    start_column=2,
):
    """
    Apply conditional formatting to a similarity matrix.
    """
    if (
        worksheet.max_row < start_row
        or worksheet.max_column < start_column
    ):
        return

    start_cell = (
        f"{get_column_letter(start_column)}"
        f"{start_row}"
    )

    end_cell = (
        f"{get_column_letter(worksheet.max_column)}"
        f"{worksheet.max_row}"
    )

    worksheet.conditional_formatting.add(
        f"{start_cell}:{end_cell}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.5,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )

    for row_number in range(
        start_row,
        worksheet.max_row + 1,
    ):
        for column_number in range(
            start_column,
            worksheet.max_column + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            cell.number_format = "0.00%"
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


# ============================================================
# LOAD POLICY MATRIX
# ============================================================

require_file(
    INPUT_FILE,
    "Dashboard dataset workbook",
)

excel_file = pd.ExcelFile(
    INPUT_FILE
)

required_sheet = "CandidatePolicyMatrix"

if required_sheet not in excel_file.sheet_names:
    raise ValueError(
        f"The workbook does not contain the "
        f"'{required_sheet}' sheet."
    )

policy_matrix = pd.read_excel(
    INPUT_FILE,
    sheet_name=required_sheet,
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

missing_identifier_columns = [
    column
    for column in IDENTIFIER_COLUMNS
    if column not in policy_matrix.columns
]

if missing_identifier_columns:
    raise ValueError(
        "CandidatePolicyMatrix is missing required columns: "
        + ", ".join(
            missing_identifier_columns
        )
    )

policy_columns = [
    column
    for column in policy_matrix.columns
    if column not in IDENTIFIER_COLUMNS
]

if not policy_columns:
    raise ValueError(
        "No policy-topic columns were found."
    )


# ============================================================
# CLEAN AND STANDARDIZE MATRIX
# ============================================================

for identifier_column in IDENTIFIER_COLUMNS:
    policy_matrix[
        identifier_column
    ] = policy_matrix[
        identifier_column
    ].apply(clean_text)

for policy_column in policy_columns:
    policy_matrix[
        policy_column
    ] = pd.to_numeric(
        policy_matrix[
            policy_column
        ],
        errors="coerce",
    ).fillna(0)

policy_matrix = policy_matrix[
    policy_matrix["Candidate_ID"] != ""
].copy()

policy_matrix.reset_index(
    drop=True,
    inplace=True,
)

candidate_count = len(
    policy_matrix
)

if candidate_count < 2:
    raise ValueError(
        "At least two candidates are required "
        "for similarity analysis."
    )

candidate_ids = policy_matrix[
    "Candidate_ID"
].tolist()

candidate_names = policy_matrix[
    "Ballot_Name"
].tolist()

policy_vectors = policy_matrix[
    policy_columns
].to_numpy(
    dtype=float
)


# ============================================================
# CALCULATE SIMILARITY MATRICES
# ============================================================

cosine_matrix = np.zeros(
    (
        candidate_count,
        candidate_count,
    ),
    dtype=float,
)

jaccard_matrix = np.zeros(
    (
        candidate_count,
        candidate_count,
    ),
    dtype=float,
)

combined_matrix = np.zeros(
    (
        candidate_count,
        candidate_count,
    ),
    dtype=float,
)

for row_index in range(
    candidate_count
):
    for column_index in range(
        candidate_count
    ):
        cosine_value = cosine_similarity(
            policy_vectors[row_index],
            policy_vectors[column_index],
        )

        jaccard_value = jaccard_similarity(
            policy_vectors[row_index],
            policy_vectors[column_index],
        )

        combined_value = (
            cosine_value
            + jaccard_value
        ) / 2

        cosine_matrix[
            row_index,
            column_index,
        ] = cosine_value

        jaccard_matrix[
            row_index,
            column_index,
        ] = jaccard_value

        combined_matrix[
            row_index,
            column_index,
        ] = combined_value


# ============================================================
# CREATE MATRIX DATAFRAMES
# ============================================================

cosine_dataframe = pd.DataFrame(
    cosine_matrix,
    index=candidate_names,
    columns=candidate_names,
).round(
    SIMILARITY_DECIMALS
)

jaccard_dataframe = pd.DataFrame(
    jaccard_matrix,
    index=candidate_names,
    columns=candidate_names,
).round(
    SIMILARITY_DECIMALS
)

combined_dataframe = pd.DataFrame(
    combined_matrix,
    index=candidate_names,
    columns=candidate_names,
).round(
    SIMILARITY_DECIMALS
)

cosine_export = (
    cosine_dataframe
    .reset_index()
    .rename(
        columns={
            "index": "Candidate"
        }
    )
)

jaccard_export = (
    jaccard_dataframe
    .reset_index()
    .rename(
        columns={
            "index": "Candidate"
        }
    )
)

combined_export = (
    combined_dataframe
    .reset_index()
    .rename(
        columns={
            "index": "Candidate"
        }
    )
)


# ============================================================
# CREATE CANDIDATE-PAIR ANALYSIS
# ============================================================

pair_rows = []

for index_a, index_b in combinations(
    range(candidate_count),
    2,
):
    row_a = policy_matrix.iloc[
        index_a
    ]

    row_b = policy_matrix.iloc[
        index_b
    ]

    shared_topics = get_shared_topics(
        row_a,
        row_b,
        policy_columns,
    )

    unique_to_a = get_unique_topics(
        row_a,
        row_b,
        policy_columns,
    )

    unique_to_b = get_unique_topics(
        row_b,
        row_a,
        policy_columns,
    )

    pair_rows.append(
        {
            "Candidate_A_ID": candidate_ids[
                index_a
            ],
            "Candidate_A": candidate_names[
                index_a
            ],
            "Candidate_B_ID": candidate_ids[
                index_b
            ],
            "Candidate_B": candidate_names[
                index_b
            ],
            "Cosine_Similarity": round(
                cosine_matrix[
                    index_a,
                    index_b,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Cosine_Similarity_Percent": (
                safe_percentage(
                    cosine_matrix[
                        index_a,
                        index_b,
                    ]
                )
            ),
            "Jaccard_Similarity": round(
                jaccard_matrix[
                    index_a,
                    index_b,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Jaccard_Similarity_Percent": (
                safe_percentage(
                    jaccard_matrix[
                        index_a,
                        index_b,
                    ]
                )
            ),
            "Combined_Similarity": round(
                combined_matrix[
                    index_a,
                    index_b,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Combined_Similarity_Percent": (
                safe_percentage(
                    combined_matrix[
                        index_a,
                        index_b,
                    ]
                )
            ),
            "Shared_Topic_Count": len(
                shared_topics
            ),
            "Shared_Policy_Topics": "; ".join(
                shared_topics
            ),
            "Unique_to_Candidate_A": "; ".join(
                unique_to_a
            ),
            "Unique_to_Candidate_B": "; ".join(
                unique_to_b
            ),
        }
    )

candidate_pairs = pd.DataFrame(
    pair_rows
)

candidate_pairs.sort_values(
    by=[
        "Combined_Similarity",
        "Cosine_Similarity",
        "Jaccard_Similarity",
    ],
    ascending=[
        False,
        False,
        False,
    ],
    inplace=True,
)

candidate_pairs.reset_index(
    drop=True,
    inplace=True,
)

candidate_pairs.insert(
    0,
    "Similarity_Rank",
    range(
        1,
        len(candidate_pairs) + 1,
    ),
)


# ============================================================
# FIND EACH CANDIDATE'S CLOSEST MATCH
# ============================================================

closest_match_rows = []

for candidate_index in range(
    candidate_count
):
    similarity_values = (
        combined_matrix[
            candidate_index
        ].copy()
    )

    similarity_values[
        candidate_index
    ] = -1

    closest_index = int(
        np.argmax(
            similarity_values
        )
    )

    row_a = policy_matrix.iloc[
        candidate_index
    ]

    row_b = policy_matrix.iloc[
        closest_index
    ]

    shared_topics = get_shared_topics(
        row_a,
        row_b,
        policy_columns,
    )

    closest_match_rows.append(
        {
            "Candidate_ID": candidate_ids[
                candidate_index
            ],
            "Candidate": candidate_names[
                candidate_index
            ],
            "Closest_Match_ID": candidate_ids[
                closest_index
            ],
            "Closest_Match": candidate_names[
                closest_index
            ],
            "Cosine_Similarity": round(
                cosine_matrix[
                    candidate_index,
                    closest_index,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Cosine_Similarity_Percent": (
                safe_percentage(
                    cosine_matrix[
                        candidate_index,
                        closest_index,
                    ]
                )
            ),
            "Jaccard_Similarity": round(
                jaccard_matrix[
                    candidate_index,
                    closest_index,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Jaccard_Similarity_Percent": (
                safe_percentage(
                    jaccard_matrix[
                        candidate_index,
                        closest_index,
                    ]
                )
            ),
            "Combined_Similarity": round(
                combined_matrix[
                    candidate_index,
                    closest_index,
                ],
                SIMILARITY_DECIMALS,
            ),
            "Combined_Similarity_Percent": (
                safe_percentage(
                    combined_matrix[
                        candidate_index,
                        closest_index,
                    ]
                )
            ),
            "Shared_Topic_Count": len(
                shared_topics
            ),
            "Shared_Policy_Topics": "; ".join(
                shared_topics
            ),
        }
    )

closest_matches = pd.DataFrame(
    closest_match_rows
)

closest_matches.sort_values(
    by="Candidate",
    inplace=True,
)

closest_matches.reset_index(
    drop=True,
    inplace=True,
)


# ============================================================
# CREATE CANDIDATE VECTOR SUMMARY
# ============================================================

vector_summary_rows = []

for candidate_index, row in (
    policy_matrix.iterrows()
):
    candidate_topic_values = {
        topic: float(
            row[topic]
        )
        for topic in policy_columns
    }

    sorted_topics = sorted(
        candidate_topic_values.items(),
        key=lambda item: item[1],
        reverse=True,
    )

    nonzero_topics = [
        (
            topic,
            value,
        )
        for topic, value in sorted_topics
        if value > 0
    ]

    top_five_topics = nonzero_topics[
        :5
    ]

    vector_summary_rows.append(
        {
            "Candidate_ID": row[
                "Candidate_ID"
            ],
            "Candidate": row[
                "Ballot_Name"
            ],
            "Policy_Topic_Count": len(
                nonzero_topics
            ),
            "Total_Policy_Weight": round(
                sum(
                    value
                    for _, value
                    in nonzero_topics
                ),
                2,
            ),
            "Top_1_Topic": (
                top_five_topics[0][0]
                if len(top_five_topics) >= 1
                else ""
            ),
            "Top_1_Weight": (
                top_five_topics[0][1]
                if len(top_five_topics) >= 1
                else 0
            ),
            "Top_2_Topic": (
                top_five_topics[1][0]
                if len(top_five_topics) >= 2
                else ""
            ),
            "Top_2_Weight": (
                top_five_topics[1][1]
                if len(top_five_topics) >= 2
                else 0
            ),
            "Top_3_Topic": (
                top_five_topics[2][0]
                if len(top_five_topics) >= 3
                else ""
            ),
            "Top_3_Weight": (
                top_five_topics[2][1]
                if len(top_five_topics) >= 3
                else 0
            ),
            "Top_4_Topic": (
                top_five_topics[3][0]
                if len(top_five_topics) >= 4
                else ""
            ),
            "Top_4_Weight": (
                top_five_topics[3][1]
                if len(top_five_topics) >= 4
                else 0
            ),
            "Top_5_Topic": (
                top_five_topics[4][0]
                if len(top_five_topics) >= 5
                else ""
            ),
            "Top_5_Weight": (
                top_five_topics[4][1]
                if len(top_five_topics) >= 5
                else 0
            ),
        }
    )

vector_summary = pd.DataFrame(
    vector_summary_rows
)


# ============================================================
# CREATE RUN SUMMARY
# ============================================================

most_similar_pair = (
    candidate_pairs.iloc[0]
    if not candidate_pairs.empty
    else None
)

least_similar_pair = (
    candidate_pairs.iloc[-1]
    if not candidate_pairs.empty
    else None
)

summary_rows = [
    {
        "Metric": "Candidates Analyzed",
        "Value": candidate_count,
    },
    {
        "Metric": "Policy Features",
        "Value": len(
            policy_columns
        ),
    },
    {
        "Metric": "Candidate Pairs",
        "Value": len(
            candidate_pairs
        ),
    },
    {
        "Metric": "Similarity Methods",
        "Value": (
            "Cosine; Jaccard; Combined Average"
        ),
    },
]

if most_similar_pair is not None:
    summary_rows.extend(
        [
            {
                "Metric": "Most Similar Pair",
                "Value": (
                    f"{most_similar_pair['Candidate_A']} "
                    f"and "
                    f"{most_similar_pair['Candidate_B']}"
                ),
            },
            {
                "Metric": (
                    "Highest Combined Similarity Percent"
                ),
                "Value": (
                    most_similar_pair[
                        "Combined_Similarity_Percent"
                    ]
                ),
            },
        ]
    )

if least_similar_pair is not None:
    summary_rows.extend(
        [
            {
                "Metric": "Least Similar Pair",
                "Value": (
                    f"{least_similar_pair['Candidate_A']} "
                    f"and "
                    f"{least_similar_pair['Candidate_B']}"
                ),
            },
            {
                "Metric": (
                    "Lowest Combined Similarity Percent"
                ),
                "Value": (
                    least_similar_pair[
                        "Combined_Similarity_Percent"
                    ]
                ),
            },
        ]
    )

run_summary = pd.DataFrame(
    summary_rows
)


# ============================================================
# SAVE CSV OUTPUTS
# ============================================================

OUTPUT_EXCEL.parent.mkdir(
    parents=True,
    exist_ok=True,
)

cosine_export.to_csv(
    OUTPUT_COSINE_CSV,
    index=False,
    encoding="utf-8-sig",
)

jaccard_export.to_csv(
    OUTPUT_JACCARD_CSV,
    index=False,
    encoding="utf-8-sig",
)

combined_export.to_csv(
    OUTPUT_COMBINED_CSV,
    index=False,
    encoding="utf-8-sig",
)


# ============================================================
# SAVE EXCEL WORKBOOK
# ============================================================

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl",
) as writer:

    run_summary.to_excel(
        writer,
        sheet_name="RunSummary",
        index=False,
    )

    closest_matches.to_excel(
        writer,
        sheet_name="ClosestMatches",
        index=False,
    )

    candidate_pairs.to_excel(
        writer,
        sheet_name="CandidatePairs",
        index=False,
    )

    cosine_export.to_excel(
        writer,
        sheet_name="CosineSimilarity",
        index=False,
    )

    jaccard_export.to_excel(
        writer,
        sheet_name="JaccardSimilarity",
        index=False,
    )

    combined_export.to_excel(
        writer,
        sheet_name="CombinedSimilarity",
        index=False,
    )

    vector_summary.to_excel(
        writer,
        sheet_name="VectorSummary",
        index=False,
    )

    policy_matrix.to_excel(
        writer,
        sheet_name="PolicyVectors",
        index=False,
    )

    output_workbook = writer.book

    for worksheet in output_workbook.worksheets:
        style_table_sheet(
            worksheet
        )

    for matrix_sheet_name in [
        "CosineSimilarity",
        "JaccardSimilarity",
        "CombinedSimilarity",
    ]:
        matrix_sheet = output_workbook[
            matrix_sheet_name
        ]

        add_similarity_heatmap(
            matrix_sheet,
            start_row=2,
            start_column=2,
        )

        matrix_sheet.freeze_panes = "B2"

        matrix_sheet.column_dimensions[
            "A"
        ].width = 24

        for column_number in range(
            2,
            matrix_sheet.max_column + 1,
        ):
            matrix_sheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = 18


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("CANDIDATE SIMILARITY ANALYSIS COMPLETE")
print("=" * 72)

print(f"Input workbook:\n{INPUT_FILE}")
print()

print(f"Excel output:\n{OUTPUT_EXCEL}")
print()

print(f"Cosine similarity CSV:\n{OUTPUT_COSINE_CSV}")
print()

print(f"Jaccard similarity CSV:\n{OUTPUT_JACCARD_CSV}")
print()

print(f"Combined similarity CSV:\n{OUTPUT_COMBINED_CSV}")
print()

print(
    f"Candidates analyzed: {candidate_count}"
)

print(
    f"Policy features used: {len(policy_columns)}"
)

print(
    f"Candidate pairs compared: {len(candidate_pairs)}"
)

if most_similar_pair is not None:
    print()

    print(
        "Most similar pair: "
        f"{most_similar_pair['Candidate_A']} and "
        f"{most_similar_pair['Candidate_B']}"
    )

    print(
        "Combined similarity: "
        f"{most_similar_pair['Combined_Similarity_Percent']:.2f}%"
    )

    print(
        "Shared policy topics: "
        f"{most_similar_pair['Shared_Topic_Count']}"
    )

print()

print(
    "Next step: open candidate_similarity_analysis.xlsx "
    "and review ClosestMatches, CandidatePairs, and "
    "CombinedSimilarity."
)