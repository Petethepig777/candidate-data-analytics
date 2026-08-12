from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = (
    Path.home()
    / "Desktop"
    / "POLITICAL_CANDIDATE_ANALYTICS"
)

CANDIDATE_MASTER_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "candidate_master_clean.xlsx"
)

POLICY_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "policy_classification_results_expanded.xlsx"
)

EVALUATION_FILE = (
    PROJECT_FOLDER
    / "data"
    / "processed"
    / "classifier_evaluation.xlsx"
)

OUTPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "output"
    / "political_analytics_dashboard_data.xlsx"
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def require_file(file_path, label):
    """
    Stop the script with a clear message when an input file is missing.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"{label} was not found:\n{file_path}"
        )


def clean_text(value):
    """
    Convert missing values to blank text and remove extra whitespace.
    """
    if pd.isna(value):
        return ""

    return " ".join(
        str(value).strip().split()
    )


def safe_divide(numerator, denominator):
    """
    Divide safely and return 0 when the denominator is zero.
    """
    if denominator == 0:
        return 0.0

    return numerator / denominator


def style_workbook(workbook):
    """
    Apply consistent formatting to the output workbook.
    """
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 32

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            column_letter = get_column_letter(
                column_number
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                65,
            )


# ============================================================
# VERIFY INPUT FILES
# ============================================================

require_file(
    CANDIDATE_MASTER_FILE,
    "Candidate master workbook",
)

require_file(
    POLICY_FILE,
    "Policy classification workbook",
)

require_file(
    EVALUATION_FILE,
    "Classifier evaluation workbook",
)


# ============================================================
# LOAD CANDIDATE MASTER DATA
# ============================================================

candidate_overview = pd.read_excel(
    CANDIDATE_MASTER_FILE,
    sheet_name="CandidateMaster",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

required_candidate_columns = [
    "Candidate_ID",
    "Ballot_Name",
    "Party",
    "Office",
    "State",
    "District",
    "District_Label",
    "Incumbent",
    "Research_Completeness_Percent",
    "Source_Count",
    "Verified_Source_Count",
    "High_Reliability_Source_Count",
    "Fact_Count",
    "Statement_Count",
    "Policy_Record_Count",
    "Social_Media_Count",
]

missing_candidate_columns = [
    column
    for column in required_candidate_columns
    if column not in candidate_overview.columns
]

if missing_candidate_columns:
    raise ValueError(
        "CandidateMaster is missing required columns: "
        + ", ".join(missing_candidate_columns)
    )


# ============================================================
# LOAD POLICY DATA
# ============================================================

policy_classifications = pd.read_excel(
    POLICY_FILE,
    sheet_name="PolicyClassifications",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_topic_summary = pd.read_excel(
    POLICY_FILE,
    sheet_name="CandidateTopicSummary",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

topic_frequency = pd.read_excel(
    POLICY_FILE,
    sheet_name="TopicFrequency",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

candidate_policy_matrix = pd.read_excel(
    POLICY_FILE,
    sheet_name="CandidatePolicyMatrix",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

statement_coverage = pd.read_excel(
    POLICY_FILE,
    sheet_name="StatementCoverage",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# LOAD CLASSIFIER EVALUATION DATA
# ============================================================

overall_metrics = pd.read_excel(
    EVALUATION_FILE,
    sheet_name="OverallMetrics",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

topic_metrics = pd.read_excel(
    EVALUATION_FILE,
    sheet_name="TopicMetrics",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

confidence_metrics = pd.read_excel(
    EVALUATION_FILE,
    sheet_name="ConfidenceMetrics",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)

label_distribution = pd.read_excel(
    EVALUATION_FILE,
    sheet_name="LabelDistribution",
    dtype=object,
).dropna(
    how="all"
).reset_index(
    drop=True
)


# ============================================================
# STANDARDIZE NUMERIC COLUMNS
# ============================================================

candidate_numeric_columns = [
    "Research_Completeness_Percent",
    "Source_Count",
    "Verified_Source_Count",
    "High_Reliability_Source_Count",
    "Fact_Count",
    "Statement_Count",
    "Policy_Record_Count",
    "Social_Media_Count",
]

for column in candidate_numeric_columns:
    candidate_overview[column] = pd.to_numeric(
        candidate_overview[column],
        errors="coerce",
    ).fillna(0)

policy_numeric_columns = [
    "Keyword_Match_Count",
    "Classification_Score",
]

for column in policy_numeric_columns:
    if column in policy_classifications.columns:
        policy_classifications[column] = pd.to_numeric(
            policy_classifications[column],
            errors="coerce",
        ).fillna(0)


# ============================================================
# CREATE POLICY DIVERSITY METRICS
# ============================================================

if policy_classifications.empty:
    policy_diversity = pd.DataFrame(
        columns=[
            "Candidate_ID",
            "Unique_Policy_Topics",
            "Total_Policy_Records",
            "Total_Keyword_Matches",
            "Average_Classification_Score",
            "High_Confidence_Record_Count",
            "Medium_Confidence_Record_Count",
            "Low_Confidence_Record_Count",
        ]
    )

else:
    working = policy_classifications.copy()

    confidence_text = (
        working["Classification_Confidence"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    working["Is_High_Confidence"] = (
        confidence_text == "High"
    ).astype(int)

    working["Is_Medium_Confidence"] = (
        confidence_text == "Medium"
    ).astype(int)

    working["Is_Low_Confidence"] = (
        confidence_text == "Low"
    ).astype(int)

    policy_diversity = (
        working
        .groupby(
            "Candidate_ID",
            as_index=False,
        )
        .agg(
            Unique_Policy_Topics=(
                "Policy_Topic",
                "nunique",
            ),
            Total_Policy_Records=(
                "Policy_Record_ID",
                "count",
            ),
            Total_Keyword_Matches=(
                "Keyword_Match_Count",
                "sum",
            ),
            Average_Classification_Score=(
                "Classification_Score",
                "mean",
            ),
            High_Confidence_Record_Count=(
                "Is_High_Confidence",
                "sum",
            ),
            Medium_Confidence_Record_Count=(
                "Is_Medium_Confidence",
                "sum",
            ),
            Low_Confidence_Record_Count=(
                "Is_Low_Confidence",
                "sum",
            ),
        )
    )

    policy_diversity[
        "Average_Classification_Score"
    ] = policy_diversity[
        "Average_Classification_Score"
    ].round(2)


# ============================================================
# MERGE CANDIDATE OVERVIEW WITH POLICY METRICS
# ============================================================

candidate_dashboard = candidate_overview.merge(
    policy_diversity,
    on="Candidate_ID",
    how="left",
)

policy_metric_columns = [
    "Unique_Policy_Topics",
    "Total_Policy_Records",
    "Total_Keyword_Matches",
    "Average_Classification_Score",
    "High_Confidence_Record_Count",
    "Medium_Confidence_Record_Count",
    "Low_Confidence_Record_Count",
]

for column in policy_metric_columns:
    if column in candidate_dashboard.columns:
        candidate_dashboard[column] = pd.to_numeric(
            candidate_dashboard[column],
            errors="coerce",
        ).fillna(0)

if not statement_coverage.empty:
    coverage_columns = [
        "Candidate_ID",
        "Total_Statements",
        "Classified_Statements",
        "Unclassified_Statements",
        "Classification_Coverage_Percent",
    ]

    available_coverage_columns = [
        column
        for column in coverage_columns
        if column in statement_coverage.columns
    ]

    candidate_dashboard = candidate_dashboard.merge(
        statement_coverage[
            available_coverage_columns
        ],
        on="Candidate_ID",
        how="left",
    )


# ============================================================
# CREATE SOURCE QUALITY METRICS
# ============================================================

candidate_dashboard[
    "Verified_Source_Rate_Percent"
] = candidate_dashboard.apply(
    lambda row: round(
        safe_divide(
            row["Verified_Source_Count"],
            row["Source_Count"],
        )
        * 100,
        2,
    ),
    axis=1,
)

candidate_dashboard[
    "High_Reliability_Source_Rate_Percent"
] = candidate_dashboard.apply(
    lambda row: round(
        safe_divide(
            row["High_Reliability_Source_Count"],
            row["Source_Count"],
        )
        * 100,
        2,
    ),
    axis=1,
)

candidate_dashboard[
    "Average_Keywords_Per_Policy_Record"
] = candidate_dashboard.apply(
    lambda row: round(
        safe_divide(
            row["Total_Keyword_Matches"],
            row["Total_Policy_Records"],
        ),
        2,
    ),
    axis=1,
)


# ============================================================
# CREATE CANDIDATE RANKINGS
# ============================================================

candidate_dashboard[
    "Research_Completeness_Rank"
] = candidate_dashboard[
    "Research_Completeness_Percent"
].rank(
    method="dense",
    ascending=False,
).astype(int)

candidate_dashboard[
    "Policy_Diversity_Rank"
] = candidate_dashboard[
    "Unique_Policy_Topics"
].rank(
    method="dense",
    ascending=False,
).astype(int)

candidate_dashboard[
    "Source_Count_Rank"
] = candidate_dashboard[
    "Source_Count"
].rank(
    method="dense",
    ascending=False,
).astype(int)


# ============================================================
# CREATE DASHBOARD SUMMARY
# ============================================================

candidate_count = len(
    candidate_dashboard
)

district_count = candidate_dashboard[
    "District_Label"
].nunique()

topic_count = (
    policy_classifications[
        "Policy_Topic"
    ].nunique()
    if not policy_classifications.empty
    else 0
)

policy_record_count = len(
    policy_classifications
)

average_completeness = round(
    candidate_dashboard[
        "Research_Completeness_Percent"
    ].mean(),
    2,
)

average_topic_diversity = round(
    candidate_dashboard[
        "Unique_Policy_Topics"
    ].mean(),
    2,
)

strict_precision = 0.0
relevant_rate = 0.0
weighted_precision = 0.0

if not overall_metrics.empty:
    strict_precision = float(
        pd.to_numeric(
            overall_metrics.loc[
                0,
                "Strict_Precision_Percent",
            ],
            errors="coerce",
        )
    )

    relevant_rate = float(
        pd.to_numeric(
            overall_metrics.loc[
                0,
                "Relevant_or_Partial_Percent",
            ],
            errors="coerce",
        )
    )

    weighted_precision = float(
        pd.to_numeric(
            overall_metrics.loc[
                0,
                "Weighted_Precision_Percent",
            ],
            errors="coerce",
        )
    )

dashboard_summary = pd.DataFrame(
    [
        {
            "Metric": "Candidate Count",
            "Value": candidate_count,
        },
        {
            "Metric": "District Count",
            "Value": district_count,
        },
        {
            "Metric": "Policy Topics Detected",
            "Value": topic_count,
        },
        {
            "Metric": "Policy Classification Records",
            "Value": policy_record_count,
        },
        {
            "Metric": "Average Research Completeness Percent",
            "Value": average_completeness,
        },
        {
            "Metric": "Average Policy Topic Diversity",
            "Value": average_topic_diversity,
        },
        {
            "Metric": "Human-Reviewed Strict Precision Percent",
            "Value": strict_precision,
        },
        {
            "Metric": "Relevant or Partial Rate Percent",
            "Value": relevant_rate,
        },
        {
            "Metric": "Weighted Precision Percent",
            "Value": weighted_precision,
        },
        {
            "Metric": "Dashboard Dataset Created",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
    ]
)


# ============================================================
# CREATE TOP-CANDIDATE SUMMARY
# ============================================================

top_candidate_rows = []

for _, row in candidate_dashboard.iterrows():
    top_topics = (
        candidate_topic_summary[
            candidate_topic_summary[
                "Candidate_ID"
            ]
            == row["Candidate_ID"]
        ]
        .sort_values(
            by="Total_Keyword_Matches",
            ascending=False,
        )
        .head(5)
    )

    top_topic_text = "; ".join(
        top_topics[
            "Policy_Topic"
        ].astype(str).tolist()
    )

    top_candidate_rows.append(
        {
            "Candidate_ID": row["Candidate_ID"],
            "Ballot_Name": row["Ballot_Name"],
            "District_Label": row["District_Label"],
            "Research_Completeness_Percent": (
                row["Research_Completeness_Percent"]
            ),
            "Source_Count": row["Source_Count"],
            "Unique_Policy_Topics": (
                row["Unique_Policy_Topics"]
            ),
            "Top_5_Policy_Topics": top_topic_text,
            "Classification_Coverage_Percent": (
                row.get(
                    "Classification_Coverage_Percent",
                    0,
                )
            ),
        }
    )

top_candidate_summary = pd.DataFrame(
    top_candidate_rows
)


# ============================================================
# REORDER CANDIDATE DASHBOARD COLUMNS
# ============================================================

preferred_columns = [
    "Candidate_ID",
    "Ballot_Name",
    "Full_Name",
    "Party",
    "Office",
    "State",
    "District",
    "District_Label",
    "Incumbent",
    "Candidate_Status",
    "Residence_City",
    "Campaign_Website",
    "Research_Completeness_Percent",
    "Research_Completeness_Rank",
    "Source_Count",
    "Verified_Source_Count",
    "Verified_Source_Rate_Percent",
    "High_Reliability_Source_Count",
    "High_Reliability_Source_Rate_Percent",
    "Source_Count_Rank",
    "Fact_Count",
    "Statement_Count",
    "Social_Media_Count",
    "Unique_Policy_Topics",
    "Policy_Diversity_Rank",
    "Total_Policy_Records",
    "Total_Keyword_Matches",
    "Average_Keywords_Per_Policy_Record",
    "Average_Classification_Score",
    "High_Confidence_Record_Count",
    "Medium_Confidence_Record_Count",
    "Low_Confidence_Record_Count",
    "Total_Statements",
    "Classified_Statements",
    "Unclassified_Statements",
    "Classification_Coverage_Percent",
]

existing_preferred_columns = [
    column
    for column in preferred_columns
    if column in candidate_dashboard.columns
]

remaining_columns = [
    column
    for column in candidate_dashboard.columns
    if column not in existing_preferred_columns
]

candidate_dashboard = candidate_dashboard[
    existing_preferred_columns
    + remaining_columns
].copy()


# ============================================================
# SAVE DASHBOARD WORKBOOK
# ============================================================

OUTPUT_FILE.parent.mkdir(
    parents=True,
    exist_ok=True,
)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl",
) as writer:

    dashboard_summary.to_excel(
        writer,
        sheet_name="DashboardSummary",
        index=False,
    )

    candidate_dashboard.to_excel(
        writer,
        sheet_name="CandidateOverview",
        index=False,
    )

    top_candidate_summary.to_excel(
        writer,
        sheet_name="CandidateHighlights",
        index=False,
    )

    candidate_policy_matrix.to_excel(
        writer,
        sheet_name="CandidatePolicyMatrix",
        index=False,
    )

    candidate_topic_summary.to_excel(
        writer,
        sheet_name="CandidateTopicSummary",
        index=False,
    )

    topic_frequency.to_excel(
        writer,
        sheet_name="TopicFrequency",
        index=False,
    )

    policy_classifications.to_excel(
        writer,
        sheet_name="PolicyClassifications",
        index=False,
    )

    overall_metrics.to_excel(
        writer,
        sheet_name="ClassifierOverallMetrics",
        index=False,
    )

    topic_metrics.to_excel(
        writer,
        sheet_name="ClassifierTopicMetrics",
        index=False,
    )

    confidence_metrics.to_excel(
        writer,
        sheet_name="ClassifierConfidence",
        index=False,
    )

    label_distribution.to_excel(
        writer,
        sheet_name="ClassifierLabels",
        index=False,
    )

    output_workbook = writer.book

    style_workbook(
        output_workbook
    )


# ============================================================
# TERMINAL OUTPUT
# ============================================================

print("=" * 72)
print("DASHBOARD DATASET CREATED")
print("=" * 72)

print(f"Candidate master:\n{CANDIDATE_MASTER_FILE}")
print()

print(f"Policy results:\n{POLICY_FILE}")
print()

print(f"Classifier evaluation:\n{EVALUATION_FILE}")
print()

print(f"Dashboard workbook:\n{OUTPUT_FILE}")
print()

print(
    f"Candidates included: {candidate_count}"
)

print(
    f"Districts included: {district_count}"
)

print(
    f"Policy topics included: {topic_count}"
)

print(
    f"Policy classification records: {policy_record_count}"
)

print(
    "Average research completeness: "
    f"{average_completeness:.2f}%"
)

print(
    "Human-reviewed strict precision: "
    f"{strict_precision:.2f}%"
)

print()

print(
    "Next step: create exploratory charts from "
    "political_analytics_dashboard_data.xlsx."
)