from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# FILE PATHS
# ============================================================

PROJECT_FOLDER = Path.home() / "Desktop" / "POLITICAL_CANDIDATE_ANALYTICS"
OUTPUT_FILE = (
    PROJECT_FOLDER
    / "data"
    / "raw"
    / "political_candidate_database.xlsx"
)


# ============================================================
# WORKBOOK STYLING
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_GRAY_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def style_header(worksheet, row_number=1):
    """Apply consistent formatting to a worksheet header row."""
    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_GRAY_BORDER

    worksheet.row_dimensions[row_number].height = 35


def set_column_widths(worksheet, widths):
    """Set column widths using a dictionary such as {'A': 15, 'B': 25}."""
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def add_excel_table(worksheet, table_name, end_column, end_row=2):
    """
    Convert the initial range into an Excel table.

    The second row is intentionally included as a blank data-entry row.
    """
    table_reference = f"A1:{end_column}{end_row}"

    table = Table(
        displayName=table_name,
        ref=table_reference,
    )

    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = table_style
    worksheet.add_table(table)


def add_list_validation(worksheet, cell_range, options):
    """Add an Excel dropdown list to a range."""
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )

    validation.error = "Please select a value from the dropdown list."
    validation.errorTitle = "Invalid Entry"
    validation.prompt = "Select one of the available values."
    validation.promptTitle = "Allowed Values"

    worksheet.add_data_validation(validation)
    validation.add(cell_range)


# ============================================================
# CREATE WORKBOOK
# ============================================================

workbook = Workbook()

# Remove the default worksheet.
default_sheet = workbook.active
workbook.remove(default_sheet)


# ============================================================
# SHEET 1: CANDIDATES
# One row per candidate
# ============================================================

candidates_sheet = workbook.create_sheet("Candidates")

candidate_columns = [
    "Candidate_ID",
    "Ballot_Name",
    "Full_Name",
    "First_Name",
    "Middle_Name",
    "Last_Name",
    "Suffix",
    "Party",
    "Office",
    "State",
    "District",
    "Incumbent",
    "Candidate_Status",
    "Filing_Date",
    "DOB",
    "Birth_Year",
    "Birthplace",
    "Residence_City",
    "Residence_State",
    "Campaign_Website",
    "Campaign_Email",
    "Campaign_Phone",
    "Campaign_Address",
    "Research_Status",
    "Researcher",
    "Date_First_Researched",
    "Date_Last_Updated",
    "Candidate_Notes",
]

candidates_sheet.append(candidate_columns)
candidates_sheet.append([""] * len(candidate_columns))

style_header(candidates_sheet)
candidates_sheet.freeze_panes = "A2"
candidates_sheet.auto_filter.ref = f"A1:AB2"

set_column_widths(
    candidates_sheet,
    {
        "A": 15,
        "B": 24,
        "C": 28,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 10,
        "H": 22,
        "I": 25,
        "J": 14,
        "K": 12,
        "L": 12,
        "M": 18,
        "N": 15,
        "O": 15,
        "P": 12,
        "Q": 22,
        "R": 20,
        "S": 18,
        "T": 40,
        "U": 32,
        "V": 18,
        "W": 38,
        "X": 18,
        "Y": 20,
        "Z": 20,
        "AA": 20,
        "AB": 45,
    },
)

add_list_validation(
    candidates_sheet,
    "H2:H1000",
    [
        "Democratic-Farmer-Labor",
        "Republican",
        "Independent",
        "Libertarian",
        "Green",
        "Nonpartisan",
        "Other",
        "Unknown",
    ],
)

add_list_validation(
    candidates_sheet,
    "L2:L1000",
    ["Yes", "No", "Unknown"],
)

add_list_validation(
    candidates_sheet,
    "M2:M1000",
    [
        "Filed",
        "Primary Candidate",
        "General Election Candidate",
        "Withdrawn",
        "Disqualified",
        "Elected",
        "Defeated",
        "Unknown",
    ],
)

add_list_validation(
    candidates_sheet,
    "X2:X1000",
    [
        "Not Started",
        "In Progress",
        "Ready for Review",
        "Complete",
        "Needs Update",
    ],
)

candidates_sheet["A2"].fill = REQUIRED_FILL
candidates_sheet["B2"].fill = REQUIRED_FILL
candidates_sheet["H2"].fill = REQUIRED_FILL
candidates_sheet["I2"].fill = REQUIRED_FILL
candidates_sheet["J2"].fill = REQUIRED_FILL
candidates_sheet["K2"].fill = REQUIRED_FILL

for row in candidates_sheet.iter_rows(min_row=2, max_row=1000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

candidates_sheet["N2"].number_format = "yyyy-mm-dd"
candidates_sheet["O2"].number_format = "yyyy-mm-dd"
candidates_sheet["Z2"].number_format = "yyyy-mm-dd"
candidates_sheet["AA2"].number_format = "yyyy-mm-dd"

add_excel_table(
    candidates_sheet,
    table_name="CandidatesTable",
    end_column="AB",
)


# ============================================================
# SHEET 2: SOURCES
# One row per online source
# ============================================================

sources_sheet = workbook.create_sheet("Sources")

source_columns = [
    "Source_ID",
    "Candidate_ID",
    "Source_Title",
    "Source_Type",
    "Publisher_or_Organization",
    "Domain",
    "URL",
    "Publication_Date",
    "Access_Date",
    "Author",
    "Reliability_Level",
    "Primary_or_Secondary",
    "Link_Status",
    "Verified",
    "Archive_URL",
    "Source_Notes",
]

sources_sheet.append(source_columns)
sources_sheet.append([""] * len(source_columns))

style_header(sources_sheet)
sources_sheet.freeze_panes = "A2"

set_column_widths(
    sources_sheet,
    {
        "A": 14,
        "B": 15,
        "C": 42,
        "D": 24,
        "E": 30,
        "F": 24,
        "G": 60,
        "H": 17,
        "I": 17,
        "J": 24,
        "K": 18,
        "L": 20,
        "M": 16,
        "N": 12,
        "O": 50,
        "P": 45,
    },
)

add_list_validation(
    sources_sheet,
    "D2:D5000",
    [
        "Government",
        "Campaign Website",
        "Official Social Media",
        "Professional Biography",
        "News",
        "Election Database",
        "Academic",
        "Nonprofit",
        "Commercial Database",
        "Other",
    ],
)

add_list_validation(
    sources_sheet,
    "K2:K5000",
    ["High", "Medium", "Low", "Unverified"],
)

add_list_validation(
    sources_sheet,
    "L2:L5000",
    ["Primary", "Secondary", "Unknown"],
)

add_list_validation(
    sources_sheet,
    "M2:M5000",
    ["Active", "Redirected", "Broken", "Blocked", "Not Checked"],
)

add_list_validation(
    sources_sheet,
    "N2:N5000",
    ["Yes", "No", "Partially"],
)

sources_sheet["A2"].fill = REQUIRED_FILL
sources_sheet["B2"].fill = REQUIRED_FILL
sources_sheet["D2"].fill = REQUIRED_FILL
sources_sheet["G2"].fill = REQUIRED_FILL

for row in sources_sheet.iter_rows(min_row=2, max_row=5000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

sources_sheet["H2"].number_format = "yyyy-mm-dd"
sources_sheet["I2"].number_format = "yyyy-mm-dd"

add_excel_table(
    sources_sheet,
    table_name="SourcesTable",
    end_column="P",
)


# ============================================================
# SHEET 3: CANDIDATE FACTS
# One row per factual statement
# ============================================================

facts_sheet = workbook.create_sheet("CandidateFacts")

fact_columns = [
    "Fact_ID",
    "Candidate_ID",
    "Fact_Category",
    "Fact_Field",
    "Fact_Value",
    "Source_ID",
    "Verification_Status",
    "Confidence_Level",
    "Is_Inferred",
    "Conflicting_Source_Flag",
    "Missing_Value_Flag",
    "Date_Verified",
    "Fact_Notes",
]

facts_sheet.append(fact_columns)
facts_sheet.append([""] * len(fact_columns))

style_header(facts_sheet)
facts_sheet.freeze_panes = "A2"

set_column_widths(
    facts_sheet,
    {
        "A": 14,
        "B": 15,
        "C": 24,
        "D": 28,
        "E": 65,
        "F": 14,
        "G": 20,
        "H": 18,
        "I": 14,
        "J": 22,
        "K": 18,
        "L": 17,
        "M": 50,
    },
)

add_list_validation(
    facts_sheet,
    "C2:C10000",
    [
        "Identity",
        "Contact",
        "Biography",
        "Family",
        "Education",
        "Professional Experience",
        "Civic Involvement",
        "Political Experience",
        "Religion",
        "Military Service",
        "Awards and Accomplishments",
        "Campaign",
        "Election",
        "Other",
    ],
)

add_list_validation(
    facts_sheet,
    "G2:G10000",
    [
        "Verified",
        "Partially Verified",
        "Unverified",
        "Conflicting",
        "Not Found",
    ],
)

add_list_validation(
    facts_sheet,
    "H2:H10000",
    ["High", "Medium", "Low", "Unknown"],
)

add_list_validation(
    facts_sheet,
    "I2:I10000",
    ["Yes", "No"],
)

add_list_validation(
    facts_sheet,
    "J2:J10000",
    ["Yes", "No"],
)

add_list_validation(
    facts_sheet,
    "K2:K10000",
    ["Yes", "No"],
)

facts_sheet["A2"].fill = REQUIRED_FILL
facts_sheet["B2"].fill = REQUIRED_FILL
facts_sheet["C2"].fill = REQUIRED_FILL
facts_sheet["D2"].fill = REQUIRED_FILL
facts_sheet["E2"].fill = REQUIRED_FILL

for row in facts_sheet.iter_rows(min_row=2, max_row=10000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

facts_sheet["L2"].number_format = "yyyy-mm-dd"

add_excel_table(
    facts_sheet,
    table_name="CandidateFactsTable",
    end_column="M",
)


# ============================================================
# SHEET 4: CANDIDATE STATEMENTS
# One row per long-form answer from the candidate profile
# ============================================================

statements_sheet = workbook.create_sheet("CandidateStatements")

statement_columns = [
    "Statement_ID",
    "Candidate_ID",
    "Statement_Type",
    "Statement_Text",
    "Primary_Source_ID",
    "Additional_Source_IDs",
    "Statement_Date",
    "Verified",
    "Manual_or_Generated",
    "Review_Status",
    "Reviewer_Notes",
]

statements_sheet.append(statement_columns)
statements_sheet.append([""] * len(statement_columns))

style_header(statements_sheet)
statements_sheet.freeze_panes = "A2"

set_column_widths(
    statements_sheet,
    {
        "A": 16,
        "B": 15,
        "C": 30,
        "D": 100,
        "E": 18,
        "F": 30,
        "G": 17,
        "H": 12,
        "I": 20,
        "J": 20,
        "K": 50,
    },
)

add_list_validation(
    statements_sheet,
    "C2:C5000",
    [
        "General Philosophy",
        "Personal and Family",
        "Professional Experience",
        "Civic Involvement",
        "Political Experience",
        "Religious Affiliation",
        "Accomplishments and Awards",
        "Educational Background",
        "Military Service",
        "Why Running",
        "Goals If Elected",
        "Achievements If Elected",
        "Areas of Concentration",
        "Entering Public Service",
        "Other Candidates",
        "Information Not Found",
        "Other",
    ],
)

add_list_validation(
    statements_sheet,
    "H2:H5000",
    ["Yes", "No", "Partially"],
)

add_list_validation(
    statements_sheet,
    "I2:I5000",
    ["Manual", "AI-Assisted", "Imported"],
)

add_list_validation(
    statements_sheet,
    "J2:J5000",
    [
        "Not Reviewed",
        "Needs Review",
        "Approved",
        "Rejected",
        "Needs Update",
    ],
)

for row in statements_sheet.iter_rows(min_row=2, max_row=5000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

statements_sheet["G2"].number_format = "yyyy-mm-dd"

add_excel_table(
    statements_sheet,
    table_name="CandidateStatementsTable",
    end_column="K",
)


# ============================================================
# SHEET 5: POLICY TOPICS
# One row per candidate-policy relationship
# ============================================================

policy_sheet = workbook.create_sheet("PolicyTopics")

policy_columns = [
    "Policy_Record_ID",
    "Candidate_ID",
    "Statement_ID",
    "Policy_Topic",
    "Policy_Subtopic",
    "Position_Summary",
    "Stance",
    "Classification_Method",
    "Keyword_or_Model_Output",
    "Classification_Confidence",
    "Human_Reviewed",
    "Reviewer_Notes",
]

policy_sheet.append(policy_columns)
policy_sheet.append([""] * len(policy_columns))

style_header(policy_sheet)
policy_sheet.freeze_panes = "A2"

set_column_widths(
    policy_sheet,
    {
        "A": 18,
        "B": 15,
        "C": 17,
        "D": 25,
        "E": 30,
        "F": 70,
        "G": 20,
        "H": 25,
        "I": 45,
        "J": 24,
        "K": 18,
        "L": 50,
    },
)

add_list_validation(
    policy_sheet,
    "D2:D10000",
    [
        "Economy",
        "Education",
        "Healthcare",
        "Housing",
        "Public Safety",
        "Criminal Justice",
        "Environment",
        "Energy",
        "Transportation",
        "Immigration",
        "Civil Rights",
        "LGBTQ+ Rights",
        "Reproductive Rights",
        "Labor and Workers",
        "Agriculture",
        "Technology and AI",
        "Government and Democracy",
        "Veterans",
        "Childcare and Families",
        "Taxes and Budget",
        "Other",
    ],
)

add_list_validation(
    policy_sheet,
    "G2:G10000",
    [
        "Support",
        "Oppose",
        "Mixed",
        "Neutral",
        "Priority",
        "Unclear",
    ],
)

add_list_validation(
    policy_sheet,
    "H2:H10000",
    [
        "Manual",
        "Rule-Based NLP",
        "Machine Learning",
        "AI-Assisted",
    ],
)

add_list_validation(
    policy_sheet,
    "J2:J10000",
    ["High", "Medium", "Low", "Unknown"],
)

add_list_validation(
    policy_sheet,
    "K2:K10000",
    ["Yes", "No"],
)

for row in policy_sheet.iter_rows(min_row=2, max_row=10000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

add_excel_table(
    policy_sheet,
    table_name="PolicyTopicsTable",
    end_column="L",
)


# ============================================================
# SHEET 6: SOCIAL MEDIA
# One row per candidate-platform account
# ============================================================

social_sheet = workbook.create_sheet("SocialMedia")

social_columns = [
    "Social_ID",
    "Candidate_ID",
    "Platform",
    "Account_Name",
    "Profile_URL",
    "Official_Status",
    "Account_Type",
    "Active_Status",
    "Last_Checked",
    "Follower_Count",
    "Notes",
]

social_sheet.append(social_columns)
social_sheet.append([""] * len(social_columns))

style_header(social_sheet)
social_sheet.freeze_panes = "A2"

set_column_widths(
    social_sheet,
    {
        "A": 14,
        "B": 15,
        "C": 18,
        "D": 30,
        "E": 60,
        "F": 18,
        "G": 20,
        "H": 18,
        "I": 17,
        "J": 18,
        "K": 45,
    },
)

add_list_validation(
    social_sheet,
    "C2:C5000",
    [
        "Facebook",
        "Instagram",
        "X or Twitter",
        "LinkedIn",
        "YouTube",
        "TikTok",
        "BlueSky",
        "Threads",
        "Medium",
        "Flickr",
        "Rumble",
        "Gettr",
        "Pinterest",
        "Blogger",
        "Wikipedia",
        "Ballotpedia",
        "Podcast",
        "GoFundMe",
        "Vimeo",
        "Other",
    ],
)

add_list_validation(
    social_sheet,
    "F2:F5000",
    ["Verified Official", "Likely Official", "Unverified", "Not Official"],
)

add_list_validation(
    social_sheet,
    "G2:G5000",
    [
        "Campaign",
        "Official Government",
        "Personal Public",
        "Professional",
        "Third Party",
    ],
)

add_list_validation(
    social_sheet,
    "H2:H5000",
    ["Active", "Inactive", "Deleted", "Blocked", "Unknown"],
)

for row in social_sheet.iter_rows(min_row=2, max_row=5000):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

social_sheet["I2"].number_format = "yyyy-mm-dd"
social_sheet["J2"].number_format = "#,##0"

add_excel_table(
    social_sheet,
    table_name="SocialMediaTable",
    end_column="K",
)


# ============================================================
# SHEET 7: DATA DICTIONARY
# Explains every table and field
# ============================================================

dictionary_sheet = workbook.create_sheet("DataDictionary")

dictionary_columns = [
    "Table_Name",
    "Column_Name",
    "Data_Type",
    "Required",
    "Description",
    "Example",
]

dictionary_sheet.append(dictionary_columns)

dictionary_rows = []

table_definitions = {
    "Candidates": {
        "Candidate_ID": (
            "Text",
            "Yes",
            "Unique candidate identifier used across all tables.",
            "MN_SD50_001",
        ),
        "Ballot_Name": (
            "Text",
            "Yes",
            "Candidate name exactly as shown on the ballot.",
            "Karla Hult",
        ),
        "Party": (
            "Category",
            "Yes",
            "Candidate political party.",
            "Democratic-Farmer-Labor",
        ),
        "Office": (
            "Text",
            "Yes",
            "Office being sought.",
            "State Senator",
        ),
        "District": (
            "Text",
            "Yes",
            "Electoral district.",
            "50",
        ),
        "DOB": (
            "Date",
            "No",
            "Verified complete date of birth. Leave blank when unknown.",
            "1986-03-12",
        ),
        "Research_Status": (
            "Category",
            "Yes",
            "Current status of candidate research.",
            "Complete",
        ),
    },
    "Sources": {
        "Source_ID": (
            "Text",
            "Yes",
            "Unique source identifier.",
            "SRC000001",
        ),
        "Candidate_ID": (
            "Text",
            "Yes",
            "Candidate associated with the source.",
            "MN_SD50_001",
        ),
        "URL": (
            "URL",
            "Yes",
            "Complete source URL.",
            "https://example.com/candidate",
        ),
        "Reliability_Level": (
            "Category",
            "Yes",
            "Researcher assessment of source reliability.",
            "High",
        ),
        "Verified": (
            "Category",
            "Yes",
            "Whether the link and content were reviewed.",
            "Yes",
        ),
    },
    "CandidateFacts": {
        "Fact_ID": (
            "Text",
            "Yes",
            "Unique identifier for one factual claim.",
            "FACT000001",
        ),
        "Fact_Field": (
            "Text",
            "Yes",
            "Specific field represented by the fact.",
            "Date of Birth",
        ),
        "Fact_Value": (
            "Long Text",
            "Yes",
            "Value or factual statement collected.",
            "March 12, 1986",
        ),
        "Source_ID": (
            "Text",
            "No",
            "Primary source supporting the fact.",
            "SRC000001",
        ),
        "Verification_Status": (
            "Category",
            "Yes",
            "Degree to which the fact has been verified.",
            "Verified",
        ),
    },
    "CandidateStatements": {
        "Statement_ID": (
            "Text",
            "Yes",
            "Unique identifier for a long-form response.",
            "STMT000001",
        ),
        "Statement_Type": (
            "Category",
            "Yes",
            "Type of question or profile section.",
            "Why Running",
        ),
        "Statement_Text": (
            "Long Text",
            "Yes",
            "Copy-ready candidate profile response.",
            "I am running because...",
        ),
    },
    "PolicyTopics": {
        "Policy_Record_ID": (
            "Text",
            "Yes",
            "Unique policy classification record.",
            "POL000001",
        ),
        "Policy_Topic": (
            "Category",
            "Yes",
            "High-level policy category.",
            "Healthcare",
        ),
        "Stance": (
            "Category",
            "No",
            "Candidate position toward the policy.",
            "Support",
        ),
        "Classification_Method": (
            "Category",
            "Yes",
            "Method used to classify the statement.",
            "Rule-Based NLP",
        ),
    },
    "SocialMedia": {
        "Social_ID": (
            "Text",
            "Yes",
            "Unique social-media record.",
            "SOC000001",
        ),
        "Platform": (
            "Category",
            "Yes",
            "Social-media or web platform.",
            "Instagram",
        ),
        "Profile_URL": (
            "URL",
            "Yes",
            "Full profile URL.",
            "https://www.instagram.com/example/",
        ),
        "Official_Status": (
            "Category",
            "Yes",
            "Assessment of whether the account is candidate-owned.",
            "Verified Official",
        ),
    },
}

for table_name, columns in table_definitions.items():
    for column_name, definition in columns.items():
        data_type, required, description, example = definition

        dictionary_rows.append(
            [
                table_name,
                column_name,
                data_type,
                required,
                description,
                example,
            ]
        )

for row in dictionary_rows:
    dictionary_sheet.append(row)

style_header(dictionary_sheet)
dictionary_sheet.freeze_panes = "A2"

set_column_widths(
    dictionary_sheet,
    {
        "A": 24,
        "B": 30,
        "C": 18,
        "D": 12,
        "E": 70,
        "F": 45,
    },
)

for row in dictionary_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_GRAY_BORDER

add_excel_table(
    dictionary_sheet,
    table_name="DataDictionaryTable",
    end_column="F",
    end_row=len(dictionary_rows) + 1,
)


# ============================================================
# SHEET 8: README
# ============================================================

readme_sheet = workbook.create_sheet("README")

readme_sheet["A1"] = "Political Candidate Analytics Database"
readme_sheet["A1"].font = Font(size=18, bold=True, color="1F4E78")

readme_sheet["A3"] = "Project Purpose"
readme_sheet["A3"].font = Font(size=13, bold=True)
readme_sheet["A4"] = (
    "This workbook stores structured, source-verified political candidate "
    "information for data cleaning, quality analysis, natural-language "
    "processing, and dashboard development."
)

readme_sheet["A6"] = "Core Tables"
readme_sheet["A6"].font = Font(size=13, bold=True)

readme_content = [
    ["Candidates", "One row per political candidate."],
    ["Sources", "One row per online source used during research."],
    ["CandidateFacts", "One row per factual candidate attribute."],
    ["CandidateStatements", "One row per long-form copy-ready response."],
    ["PolicyTopics", "One row per candidate and policy-topic classification."],
    ["SocialMedia", "One row per social-media or public profile."],
    ["DataDictionary", "Definitions and examples for database fields."],
]

for row_number, row_data in enumerate(readme_content, start=7):
    readme_sheet.cell(row=row_number, column=1, value=row_data[0])
    readme_sheet.cell(row=row_number, column=2, value=row_data[1])

readme_sheet["A16"] = "ID Conventions"
readme_sheet["A16"].font = Font(size=13, bold=True)

id_examples = [
    ["Candidate_ID", "MN_SD50_001"],
    ["Source_ID", "SRC000001"],
    ["Fact_ID", "FACT000001"],
    ["Statement_ID", "STMT000001"],
    ["Policy_Record_ID", "POL000001"],
    ["Social_ID", "SOC000001"],
]

for row_number, row_data in enumerate(id_examples, start=17):
    readme_sheet.cell(row=row_number, column=1, value=row_data[0])
    readme_sheet.cell(row=row_number, column=2, value=row_data[1])

readme_sheet["A25"] = "Important Research Rules"
readme_sheet["A25"].font = Font(size=13, bold=True)

research_rules = [
    "Do not guess missing facts.",
    "Record the full source URL.",
    "Use government and first-party sources whenever possible.",
    "Distinguish verified facts from inferences.",
    "Record conflicting sources rather than silently selecting one.",
    "Leave DOB blank when only an unverified commercial record is available.",
    "Do not publish private residential addresses.",
]

for row_number, rule in enumerate(research_rules, start=26):
    readme_sheet.cell(row=row_number, column=1, value=f"• {rule}")

readme_sheet["A35"] = "Workbook Created"
readme_sheet["B35"] = datetime.now()
readme_sheet["B35"].number_format = "yyyy-mm-dd hh:mm"

readme_sheet.column_dimensions["A"].width = 35
readme_sheet.column_dimensions["B"].width = 95

for row in readme_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)


# ============================================================
# SAVE WORKBOOK
# ============================================================

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT_FILE)

print("=" * 70)
print("POLITICAL CANDIDATE DATABASE TEMPLATE CREATED")
print("=" * 70)
print(f"File saved to:\n{OUTPUT_FILE}")
print()
print("Sheets created:")
for worksheet in workbook.worksheets:
    print(f"  - {worksheet.title}")
print()
print("Next step: open the workbook and confirm that all sheets appear.")